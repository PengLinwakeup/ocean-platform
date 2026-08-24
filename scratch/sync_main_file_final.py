import openpyxl
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
purple_file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest_PURPLE_DSW.xlsx"
backup8_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest_backup8.xlsx"

# 1. Create backup8
shutil.copyfile(file_path, backup8_path)
print(f"Backup created: {backup8_path}")

# 2. Copy purple file directly to main file
shutil.copyfile(purple_file_path, file_path)
print(f"Successfully copied purple-highlighted file to main file: {file_path}")

# 3. Verify main file
wb = openpyxl.load_workbook(file_path, data_only=True)
ws_master = wb["All_Columns_Sequence_QC_Master"]

purple_dsw_count = 0
below_39_count = 0

wb_style = openpyxl.load_workbook(file_path, data_only=False)
ws_master_style = wb_style["All_Columns_Sequence_QC_Master"]

for r in range(6, ws_master.max_row + 1):
    c1_val = ws_master.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master.cell(r, 2).value or "").strip()
        cat_type = str(ws_master.cell(r, 3).value or "").strip().upper()
        
        if "DSW" in cat_type or "DSW" in s_name.upper() or "CRM" in cat_type or "CRM" in s_name.upper():
            qc_doc = ws_master.cell(r, 14).value
            doc_num = float(qc_doc) if qc_doc is not None and isinstance(qc_doc, (int, float)) else 0.0
            
            cell_f = ws_master_style.cell(r, 14).fill
            fill_hex = cell_f.fgColor.value if (cell_f and cell_f.fgColor) else ""
            
            if fill_hex == "FFE9D5FF":
                purple_dsw_count += 1
                
            if doc_num < 39.0 and doc_num > 0:
                below_39_count += 1

print(f"\nVerification Results on Main File:")
print(f"Total DSW rows highlighted in PURPLE (#E9D5FF): {purple_dsw_count}")
print(f"Total DSW rows with QC Dynamic DOC < 39.0: {below_39_count} (Expected: 0)")
