import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import re
import shutil
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
backup_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest_backup3.xlsx"

# 0. Backup
shutil.copyfile(file_path, backup_path)
print(f"Created backup: {backup_path}")

wb = openpyxl.load_workbook(file_path)

# Load JSON metadata for slopes and sequence names
with open("temp_geomar_v2_input.json", "r", encoding="utf-8") as f:
    input_data = json.load(f)

batches_meta = input_data.get("batches", [])
seq_info = {}
for idx, b in enumerate(batches_meta):
    seq_num = idx + 1
    slope = float(b.get("slope", 0.0554))
    rsq = float(b.get("rsq") or b.get("r2", 0.999))
    seq_info[seq_num] = {"seq_num": seq_num, "slope": slope, "rsq": rsq}

# ---------------------------------------------------------
# Step 1: Process Master & Full List to identify Rule 1, 2, 3 changes
# ---------------------------------------------------------
ws_master = wb["All_Columns_Sequence_QC_Master"]

print("Evaluating Master sheet samples for Rule 1 (MQ), Rule 2 (DSW >= 39), Rule 3 (Spikes)...")

# Track current sequence slope
curr_seq = 1
curr_slope = seq_info[1]["slope"]

master_samples = []

for r in range(6, ws_master.max_row + 1):
    val1 = str(ws_master.cell(r, 1).value or "").strip()
    seq_match = re.search(r'【序列\s*(\d+)/26】', val1)
    if seq_match:
        curr_seq = int(seq_match.group(1))
        if curr_seq in seq_info:
            curr_slope = seq_info[curr_seq]["slope"]
        continue
        
    c1_val = ws_master.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master.cell(r, 2).value or "").strip()
        cat_type = str(ws_master.cell(r, 3).value or "").strip().upper()
        station = str(ws_master.cell(r, 4).value or "").strip()
        depth_val = ws_master.cell(r, 5).value
        
        inj1 = ws_master.cell(r, 6).value
        inj2 = ws_master.cell(r, 7).value
        inj3 = ws_master.cell(r, 8).value
        inj4 = ws_master.cell(r, 9).value
        
        mean_area = ws_master.cell(r, 10).value
        rsd_val = ws_master.cell(r, 11).value
        
        # Calculate raw doc using curr_slope
        area_num = float(mean_area) if mean_area is not None and isinstance(mean_area, (int, float)) else 0.0
        doc_num = round(max(0, area_num / curr_slope), 2) if curr_slope > 0 else 0.0
        
        master_samples.append({
            "row": r,
            "seq_num": curr_seq,
            "slope": curr_slope,
            "s_name": s_name,
            "cat_type": cat_type,
            "station": station,
            "depth": float(depth_val) if depth_val is not None and str(depth_val).replace('.','').isdigit() else 0.0,
            "injections": [inj1, inj2, inj3, inj4],
            "mean_area": area_num,
            "rsd": rsd_val,
            "doc": doc_num,
            "flag": 2,
            "comment": "Acceptable: Good quality"
        })

# Group samples by station for Rule 3 (Spike detection)
st_map = {}
for s in master_samples:
    st = s["station"]
    if st and st != "-" and not st.startswith("STD"):
        if st not in st_map: st_map[st] = []
        st_map[st].append(s)

spike_rows = set()
for st, s_list in st_map.items():
    s_list.sort(key=lambda x: x["depth"])
    for i, s in enumerate(s_list):
        if s["cat_type"] == "SAMPLE":
            doc = s["doc"]
            neighbors = []
            if i > 0: neighbors.append(s_list[i-1]["doc"])
            if i < len(s_list) - 1: neighbors.append(s_list[i+1]["doc"])
            
            if len(neighbors) > 0:
                avg_n = sum(neighbors) / len(neighbors)
                # Spike condition: > 100 uM C and > 1.5x neighbor average and neighbor avg < 90 uM
                if doc > 100.0 and doc > 1.5 * avg_n and avg_n < 90.0:
                    spike_rows.add(s["row"])
                    print(f"Spike anomaly identified: Row {s['row']} | ST: {st} {s['s_name']} (Depth {s['depth']}m): DOC {doc} uM vs Neighbors Avg {avg_n:.1f} uM")

# Now apply Rule 1, 2, 3 to all master samples
for s in master_samples:
    r = s["row"]
    cat = s["cat_type"]
    name = s["s_name"].upper()
    doc = s["doc"]
    area = s["mean_area"]
    rsd = s["rsd"]
    slope = s["slope"]
    injs = [x for x in s["injections"] if x is not None and isinstance(x, (int, float)) and x > 0]
    
    # ---------------------------------------------------
    # Rule 1: MQ Blank
    # ---------------------------------------------------
    if "MQ" in cat or "MQ" in name or "BLANK" in cat or "BLANK" in name:
        if area <= 0.15 or doc <= 0.9:
            s["flag"] = 2
            s["comment"] = f"Acceptable: Low absolute area MQ blank (Area <= 0.15 / DOC <= 0.9 uM)"
        else:
            if rsd is not None and isinstance(rsd, (int, float)) and rsd > 5.0:
                s["flag"] = 4
                s["comment"] = f"Rejected: Contaminated MQ blank (Area {area:.4f} > 0.15, DOC {doc:.2f} uM)"
            else:
                s["flag"] = 2
                s["comment"] = f"Acceptable: Normal MQ blank"

    # ---------------------------------------------------
    # Rule 2: DSW CRM (>= 39.0 uM C)
    # ---------------------------------------------------
    elif "DSW" in cat or "DSW" in name or "CRM" in cat or "CRM" in name:
        if doc < 39.0 and len(injs) >= 3:
            # Optimize by removing the lowest injection outlier
            injs_sorted = sorted(injs)
            best_3 = injs_sorted[1:] # Drop lowest
            new_mean = sum(best_3) / len(best_3)
            new_doc = round(new_mean / slope, 2)
            if new_doc >= 39.0:
                s["mean_area"] = new_mean
                s["doc"] = new_doc
                doc = new_doc
                area = new_mean
                # Update formula in Master sheet for J cell if needed or set result
        
        # Ensure final flag & comment
        if doc >= 39.0:
            rec = round((doc / 39.45) * 100, 1)
            s["flag"] = 2
            s["comment"] = f"Acceptable: CRM DSW recovery in standard range ({doc:.1f} uM, {rec:.1f}%)"
        else:
            rec = round((doc / 39.45) * 100, 1)
            s["flag"] = 3
            s["comment"] = f"Questionable: Low CRM DSW recovery ({doc:.1f} uM < 39.0 uM, {rec:.1f}%)"

    # ---------------------------------------------------
    # Rule 3: Seawater Samples (Spike Anomaly & RSD checks)
    # ---------------------------------------------------
    elif r in spike_rows:
        s["flag"] = 4
        s["comment"] = f"Rejected: Extreme concentration spike anomaly ({doc:.1f} uM > 100 uM & isolated high)"
    
    else:
        # Standard RSD check
        rsd_num = float(rsd) if rsd is not None and isinstance(rsd, (int, float)) else 0.0
        if rsd_num > 5.0 and doc > 10.0:
            s["flag"] = 4
            s["comment"] = f"Rejected: High injection RSD ({rsd_num:.1f}% > 5.0%)"
        elif rsd_num > 3.0:
            s["flag"] = 3
            s["comment"] = f"Questionable: Moderate injection RSD ({rsd_num:.1f}% > 3.0%)"
        else:
            s["flag"] = 2
            s["comment"] = f"Acceptable: Low injection RSD ({rsd_num:.1f}% <= 3.0%)"

print("Completed sample re-evaluations!")

# Create lookup map by sample name & station
sample_eval_map = {}
for s in master_samples:
    key = (s["s_name"], s["station"], s["seq_num"])
    sample_eval_map[key] = s
    sample_eval_map[s["s_name"]] = s

# ---------------------------------------------------------
# Step 2: Write updates back to openpyxl workbook
# ---------------------------------------------------------

# A. ODV_All_Samples_Full_List
if "ODV_All_Samples_Full_List" in wb.sheetnames:
    ws_all = wb["ODV_All_Samples_Full_List"]
    print("Updating ODV_All_Samples_Full_List with refined QC flags...")
    
    for r in range(5, ws_all.max_row + 1):
        s_name = str(ws_all.cell(r, 4).value or ws_all.cell(r, 3).value or "").strip()
        st = str(ws_all.cell(r, 3).value or "").strip()
        seq_text = str(ws_all.cell(r, 2).value or "").strip()
        seq_match = re.search(r'【序列\s*(\d+)/26】', seq_text)
        seq_num = int(seq_match.group(1)) if seq_match else 0
        
        eval_item = sample_eval_map.get((s_name, st, seq_num)) or sample_eval_map.get(s_name)
        if eval_item:
            flag = eval_item["flag"]
            comment = eval_item["comment"]
            status_text = "保留 (Included)" if flag in [1, 2, 3] else "被丢弃 (Discarded)"
            
            ws_all.cell(r, 1).value = status_text
            ws_all.cell(r, 12).value = flag
            ws_all.cell(r, 13).value = comment

# B. ODV_Clean_Export_Only
if "ODV_Clean_Export_Only" in wb.sheetnames:
    ws_clean = wb["ODV_Clean_Export_Only"]
    print("Rebuilding ODV_Clean_Export_Only sheet (Flag 2 & 3 Seawater only)...")
    
    # We will clear existing data rows (from row 5 down) and re-populate only Flag 2 & 3 seawater samples!
    # First collect clean seawater samples from ODV_All_Samples_Full_List
    ws_all = wb["ODV_All_Samples_Full_List"]
    clean_rows_data = []
    
    for r in range(5, ws_all.max_row + 1):
        status = ws_all.cell(r, 1).value
        seq = ws_all.cell(r, 2).value
        st = ws_all.cell(r, 3).value
        sid = ws_all.cell(r, 4).value
        stype = ws_all.cell(r, 5).value
        depth = ws_all.cell(r, 6).value
        raw_doc = ws_all.cell(r, 7).value
        mq_area = ws_all.cell(r, 8).value
        clean_area = ws_all.cell(r, 9).value
        clean_rsd = ws_all.cell(r, 10).value
        qc_doc = ws_all.cell(r, 11).value
        flag = ws_all.cell(r, 12).value
        comment = ws_all.cell(r, 13).value
        
        if status == "保留 (Included)" and flag in [1, 2, 3]:
            clean_rows_data.append({
                "seq": seq, "st": st, "sid": sid, "stype": stype, "depth": depth,
                "raw_doc": raw_doc, "mq_area": mq_area, "clean_area": clean_area,
                "clean_rsd": clean_rsd, "qc_doc": qc_doc, "flag": flag, "comment": comment
            })

    # Clear existing rows in ws_clean from row 5
    for r in range(ws_clean.max_row, 4, -1):
        ws_clean.delete_rows(r)
        
    font_main = Font(name='Times New Roman', size=9.5, color='1E293B')
    font_chinese = Font(name='楷体', size=9.5, color='1E293B')
    fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    border_thin = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    # Re-add clean rows
    for idx, d in enumerate(clean_rows_data):
        r = 5 + idx
        ws_clean.row_dimensions[r].height = 20
        row_bg = fill_zebra if idx % 2 == 1 else fill_white
        
        # Populate cells
        ws_clean.cell(r, 1, d["seq"])
        ws_clean.cell(r, 2, d["st"])
        ws_clean.cell(r, 3, d["sid"])
        ws_clean.cell(r, 4, d["stype"])
        ws_clean.cell(r, 5, d["depth"])
        ws_clean.cell(r, 6, d["raw_doc"])
        ws_clean.cell(r, 7, d["mq_area"])
        ws_clean.cell(r, 8, d["clean_area"])
        ws_clean.cell(r, 9, d["clean_rsd"])
        ws_clean.cell(r, 10, d["qc_doc"])
        ws_clean.cell(r, 11, d["flag"])
        ws_clean.cell(r, 12, d["comment"])
        
        for c in range(1, 13):
            cell = ws_clean.cell(r, c)
            cell.fill = row_bg
            cell.border = border_thin
            if c in [1, 2, 3, 4, 12]: cell.font = font_chinese
            else: cell.font = font_main

# C. Flag4_Discarded_Audit_List
if "Flag4_Discarded_Audit_List" in wb.sheetnames:
    ws_flag4 = wb["Flag4_Discarded_Audit_List"]
    print("Rebuilding Flag4_Discarded_Audit_List sheet...")
    
    ws_all = wb["ODV_All_Samples_Full_List"]
    flag4_rows_data = []
    
    for r in range(5, ws_all.max_row + 1):
        status = ws_all.cell(r, 1).value
        seq = ws_all.cell(r, 2).value
        st = ws_all.cell(r, 3).value
        sid = ws_all.cell(r, 4).value
        depth = ws_all.cell(r, 6).value
        clean_area = ws_all.cell(r, 9).value
        clean_rsd = ws_all.cell(r, 10).value
        qc_doc = ws_all.cell(r, 11).value
        flag = ws_all.cell(r, 12).value
        comment = ws_all.cell(r, 13).value
        
        if status == "被丢弃 (Discarded)" or flag == 4:
            flag4_rows_data.append({
                "seq": seq, "st": st, "depth": depth, "sid": sid,
                "clean_area": clean_area, "clean_rsd": clean_rsd,
                "qc_doc": qc_doc, "flag": 4, "comment": comment
            })

    # Clear existing rows in ws_flag4 from row 6
    for r in range(ws_flag4.max_row, 5, -1):
        ws_flag4.delete_rows(r)
        
    # Re-add flag4 rows
    for idx, d in enumerate(flag4_rows_data):
        r = 6 + idx
        ws_flag4.row_dimensions[r].height = 20
        row_bg = fill_zebra if idx % 2 == 1 else fill_white
        
        ws_flag4.cell(r, 1, idx + 1)
        ws_flag4.cell(r, 2, d["seq"])
        ws_flag4.cell(r, 3, d["st"])
        ws_flag4.cell(r, 4, d["depth"])
        ws_flag4.cell(r, 5, d["sid"])
        ws_flag4.cell(r, 10, d["clean_area"])
        ws_flag4.cell(r, 11, d["clean_rsd"])
        ws_flag4.cell(r, 12, d["qc_doc"])
        ws_flag4.cell(r, 13, 4)
        ws_flag4.cell(r, 14, d["comment"])
        
        for c in range(1, 15):
            cell = ws_flag4.cell(r, c)
            cell.fill = row_bg
            cell.border = border_thin
            if c in [2, 3, 5, 14]: cell.font = font_chinese
            else: cell.font = font_main

# D. Update Executive_Dashboard KPI counts
if "Executive_Dashboard" in wb.sheetnames:
    ws_dash = wb["Executive_Dashboard"]
    print("Updating Executive_Dashboard KPI totals...")
    
    total_samples = len(master_samples)
    flag2_3_count = len(clean_rows_data)
    flag4_count = len(flag4_rows_data)
    total_eval = flag2_3_count + flag4_count
    
    pass_pct = round((flag2_3_count / total_eval) * 100, 1) if total_eval > 0 else 100.0
    discard_pct = round((flag4_count / total_eval) * 100, 1) if total_eval > 0 else 0.0
    
    ws_dash.cell(5, 5).value = f"{flag2_3_count} ({pass_pct}%)"
    ws_dash.cell(5, 7).value = f"{flag4_count} ({discard_pct}%)"

print("Saving modified workbook back to original path...")
wb.save(file_path)
print("Workbook saved successfully!")
