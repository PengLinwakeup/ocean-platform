import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path)

ws_master = wb["All_Columns_Sequence_QC_Master"]

print("Rows in Master where Flag is 2, 3, 4:")
mismatches = []
for r in range(6, ws_master.max_row + 1):
    c1 = ws_master.cell(r, 1)
    co = ws_master.cell(r, 15)
    flag_val = co.value
    
    if flag_val in [1, 2, 3, 4]:
        c1_fill = c1.fill.fgColor.value if (c1.fill and c1.fill.fgColor) else None
        co_fill = co.fill.fgColor.value if (co.fill and co.fill.fgColor) else None
        
        # Flag 2 should have green fill (DCFCE7), Flag 3 yellow (FEF08A or FEF9C3), Flag 4 red (FEE2E2 or FFC7CE)
        expected_fill = "FFDCFCE7" if flag_val in [1, 2] else ("FFFEF08A" if flag_val == 3 else "FFFEE2E2")
        
        if c1_fill != expected_fill or co_fill != expected_fill:
            mismatches.append((r, flag_val, c1_fill, co_fill, expected_fill))

print(f"Total rows with mismatched background fill: {len(mismatches)}")
for r, flag_val, c1_fill, co_fill, expected_fill in mismatches[:20]:
    print(f"Row {r:4d} | Flag: {flag_val} | Col 1 Fill: {c1_fill} | Col 15 Fill: {co_fill} | Expected: {expected_fill}")
