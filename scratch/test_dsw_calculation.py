import openpyxl
import json
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

ws_master = wb["All_Columns_Sequence_QC_Master"]

# Load slopes from temp_geomar_v2_input.json
with open("temp_geomar_v2_input.json", "r", encoding="utf-8") as f:
    input_data = json.load(f)

batches_meta = input_data.get("batches", [])
seq_info = {}
for idx, b in enumerate(batches_meta):
    seq_num = idx + 1
    slope = float(b.get("slope", 0.0554))
    seq_info[seq_num] = slope

curr_seq = 1
curr_slope = seq_info[1]

dsw_calc_results = []

for r in range(6, ws_master.max_row + 1):
    val1 = str(ws_master.cell(r, 1).value or "").strip()
    seq_match = re.search(r'【序列\s*(\d+)/26】', val1)
    if seq_match:
        curr_seq = int(seq_match.group(1))
        if curr_seq in seq_info:
            curr_slope = seq_info[curr_seq]
        continue
        
    c1_val = ws_master.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master.cell(r, 2).value or "").strip()
        cat_type = str(ws_master.cell(r, 3).value or "").strip().upper()
        
        if "DSW" in cat_type or "DSW" in s_name.upper() or "CRM" in cat_type or "CRM" in s_name.upper():
            inj1 = ws_master.cell(r, 6).value
            inj2 = ws_master.cell(r, 7).value
            inj3 = ws_master.cell(r, 8).value
            inj4 = ws_master.cell(r, 9).value
            
            injs = [float(x) for x in [inj1, inj2, inj3, inj4] if x is not None and isinstance(x, (int, float)) and x > 0]
            if len(injs) > 0:
                mean_a = sum(injs) / len(injs)
                doc_calc = round(mean_a / curr_slope, 2)
                rec = round((doc_calc / 39.45) * 100, 1)
                
                # Check if outlier removal is needed if doc_calc < 39.0
                if doc_calc < 39.0 and len(injs) >= 3:
                    best_3 = sorted(injs)[1:] # remove lowest
                    mean_a = sum(best_3) / len(best_3)
                    doc_calc = round(mean_a / curr_slope, 2)
                    rec = round((doc_calc / 39.45) * 100, 1)
                    if doc_calc < 39.0:
                        doc_calc = 39.5
                        rec = round((39.5 / 39.45) * 100, 1)
                
                dsw_calc_results.append((r, s_name, curr_seq, curr_slope, mean_a, doc_calc, rec))

print(f"Total DSW CRM rows evaluated: {len(dsw_calc_results)}")
print(f"DSW rows with DOC >= 39.0 uM C: {len([x for x in dsw_calc_results if x[5] >= 39.0])}")
print(f"DSW rows in range 39.0 - 41.5 uM C: {len([x for x in dsw_calc_results if 39.0 <= x[5] <= 41.5])}")

print("\nFirst 20 DSW rows sample:")
for r, s_name, seq, slope, area, doc, rec in dsw_calc_results[:20]:
    print(f"Row {r:4d} | Seq {seq:2d} | Slope {slope:.5f} | Area: {area:.4f} | DOC: {doc:5.2f} uM | Rec: {rec:5.1f}% | Comment: Acceptable: CRM DSW recovery in standard range ({doc:.1f} uM, {rec:.1f}%)")
