import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"

print("================ VERIFYING EXECUTIVE DASHBOARD ================")
wb_val = openpyxl.load_workbook(file_path, data_only=True)
ws_dash = wb_val["Executive_Dashboard"]

slopes_found = []
rsq_found = []

for r in range(8, 34):
    seq_idx = ws_dash.cell(r, 1).value
    name = ws_dash.cell(r, 2).value
    rsq = ws_dash.cell(r, 4).value
    slope = ws_dash.cell(r, 5).value
    slopes_found.append(slope)
    rsq_found.append(rsq)
    print(f"Row {r:2d} | Seq {seq_idx:2d} | Name: {name[:50]:50s} | R2: {rsq} | Slope: {slope}")

unique_slopes = set(slopes_found)
unique_rsq = set(rsq_found)

print(f"\nTotal sequences: {len(slopes_found)}")
print(f"Unique Slopes count: {len(unique_slopes)} (Expected > 20 distinct slopes)")
print(f"Unique R2 count: {len(unique_rsq)} (Expected > 20 distinct R2s)")

print("\n================ VERIFYING FORMULAS IN MASTER & CLEAN =")
wb_formula = openpyxl.load_workbook(file_path, data_only=False)

ws_master = wb_formula["All_Columns_Sequence_QC_Master"]
print("\nMaster Sheet Sample Rows Formulas (Checking distinct slopes in formula):")
sample_check_rows = [7, 30, 200, 500, 900, 1300, 1600]
for r in sample_check_rows:
    seq_name = ws_master.cell(r, 2).value
    raw_doc_formula = ws_master.cell(r, 12).value
    qc_doc_formula = ws_master.cell(r, 14).value
    print(f"Row {r:4d} | Name: {str(seq_name):20s} | Raw DOC Formula: {raw_doc_formula} | QC DOC Formula: {qc_doc_formula}")

ws_clean = wb_formula["ODV_Clean_Export_Only"]
print("\nClean Export Sheet Sample Rows Formulas:")
for r in range(5, 15):
    seq = ws_clean.cell(r, 1).value
    st = ws_clean.cell(r, 2).value
    raw_doc = ws_clean.cell(r, 6).value
    print(f"Row {r:2d} | Seq: {seq[:45]:45s} | ST: {st} | Raw DOC Formula: {raw_doc}")
