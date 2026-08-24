import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

ws = wb["Executive_Dashboard"] if "Executive_Dashboard" in wb.sheetnames else wb.worksheets[0]

print(f"Sheet Name: {ws.title}")
print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")

for r in range(1, ws.max_row + 1):
    row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if any(v is not None for v in row_vals):
        print(f"Row {r:2d}: {row_vals}")
