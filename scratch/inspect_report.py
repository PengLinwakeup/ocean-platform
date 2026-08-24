import openpyxl
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

print("Sheet names in workbook:")
for name in wb.sheetnames:
    print(f" - {name}")

ws = wb["Executive_Dashboard"] if "Executive_Dashboard" in wb.sheetnames else None
if ws:
    print("\n--- Executive_Dashboard contents ---")
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}: {row_vals}")

ws_master = wb["All_Columns_Sequence_QC_Master"] if "All_Columns_Sequence_QC_Master" in wb.sheetnames else None
if ws_master:
    print("\n--- All_Columns_Sequence_QC_Master titles and slopes ---")
    for r in range(1, ws_master.max_row + 1):
        v1 = ws_master.cell(r, 1).value
        if v1 and ("【序列" in str(v1) or "柱" in str(v1) or "R2" in str(v1) or "斜率" in str(v1) or "Slope" in str(v1)):
            print(f"Row {r:3d}: {v1}")

ws_clean = wb["ODV_Clean_Export_Only"] if "ODV_Clean_Export_Only" in wb.sheetnames else None
if ws_clean:
    print("\n--- ODV_Clean_Export_Only Sample rows (First 20) ---")
    for r in range(1, 25):
        row_vals = [ws_clean.cell(r, c).value for c in range(1, ws_clean.max_column + 1)]
        print(f"Row {r:2d}: {row_vals}")
