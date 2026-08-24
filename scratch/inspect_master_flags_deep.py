import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

ws_master = wb["All_Columns_Sequence_QC_Master"]

flag_counts = {1: 0, 2: 0, 3: 0, 4: 0, "other": 0}
sample_flags = []

for r in range(6, ws_master.max_row + 1):
    c1_val = ws_master.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master.cell(r, 2).value or "").strip()
        cat_type = str(ws_master.cell(r, 3).value or "").strip().upper()
        flag = ws_master.cell(r, 15).value if ws_master.max_column >= 15 else None
        comment = ws_master.cell(r, 16).value if ws_master.max_column >= 16 else None
        
        if flag in [1, 2, 3, 4]:
            flag_counts[flag] += 1
        else:
            flag_counts["other"] += 1
            
        sample_flags.append((r, s_name, cat_type, flag, comment))

print("Flag counts in All_Columns_Sequence_QC_Master:")
print(flag_counts)

print("\nSample of rows in Master:")
for r, s_name, cat_type, flag, comment in sample_flags[:30]:
    print(f"Row {r:4d} | Name: {s_name:25s} | Type: {cat_type:10s} | Flag: {flag} | Comment: {comment}")

ws_all = wb["ODV_All_Samples_Full_List"]
flag_counts_all = {1: 0, 2: 0, 3: 0, 4: 0, "other": 0}
for r in range(5, ws_all.max_row + 1):
    flag = ws_all.cell(r, 12).value
    if flag in [1, 2, 3, 4]:
        flag_counts_all[flag] += 1
    else:
        flag_counts_all["other"] += 1

print("\nFlag counts in ODV_All_Samples_Full_List:")
print(flag_counts_all)
