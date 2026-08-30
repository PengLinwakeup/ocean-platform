import openpyxl
import os
import glob
import re

target_excel = r'F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829_Final_QC_Completed.xlsx'
sample_list_file = r'F:\印度洋测样\Indian Ocean_SO308_DOC_Sample List(1) 的副本.xlsx'

wb_target = openpyxl.load_workbook(target_excel, data_only=True)
ws_master = wb_target['All_Columns_Sequence_QC_Master']

wb_sample = openpyxl.load_workbook(sample_list_file, data_only=True)
ws_sample = wb_sample['Sheet1']

# Build mapping from Sample List: (Station, Depth) -> Lable ID_DOC
# Row 1 header: ['Longitude (E)', 'Latitude (N)', 'Station', 'Time (UTC)', 'Bot.Depth (m)', 'Depth (m)', 'Lable ID_DOC']
sample_map = {}
for r in range(2, ws_sample.max_row + 1):
    st = ws_sample.cell(r, 3).value
    dp = ws_sample.cell(r, 6).value
    lbl = ws_sample.cell(r, 7).value
    if st and dp is not None and lbl:
        st_norm = str(st).strip().upper().replace(' ', '')
        # handle depth int/float
        try:
            dp_norm = int(round(float(dp)))
        except:
            dp_norm = str(dp).strip()
        sample_map[(st_norm, dp_norm)] = str(lbl).strip()

out_lines = []
out_lines.append(f"Loaded {len(sample_map)} records from Sample List.")

out_lines.append("\n=== INSPECTING 40xxx SAMPLE NAMES IN TARGET MASTER ===")
err_40xxx = []
zero_injs = []

for r in range(7, ws_master.max_row + 1):
    seq = ws_master.cell(r, 1).value
    sname = ws_master.cell(r, 2).value
    ctype = ws_master.cell(r, 3).value
    st = ws_master.cell(r, 4).value
    dp = ws_master.cell(r, 5).value
    inj1 = ws_master.cell(r, 6).value
    inj2 = ws_master.cell(r, 7).value
    inj3 = ws_master.cell(r, 8).value
    inj4 = ws_master.cell(r, 9).value
    mean_area = ws_master.cell(r, 10).value
    flag = ws_master.cell(r, 15).value
    comment = ws_master.cell(r, 16).value
    
    if sname and isinstance(sname, str):
        # check 40xxx pattern
        # e.g., SO308-40123 or 40123 or SO308-41183
        m = re.search(r'40\d{3}', sname)
        if m or '40' in sname:
            st_norm = str(st).strip().upper().replace(' ', '') if st else ''
            try:
                dp_norm = int(round(float(dp))) if dp is not None else None
            except:
                dp_norm = dp
            
            correct_id = sample_map.get((st_norm, dp_norm))
            out_lines.append(f"Row {r:4d} | Seq={seq} | Current={sname} | ST={st} | Depth={dp} | Map Match => {correct_id} | Flag={flag}")
            err_40xxx.append((r, sname, st, dp, correct_id))
            
    # Check 0.00 / missing injections
    injs = [inj1, inj2, inj3, inj4]
    valid_injs = [x for x in injs if x is not None and isinstance(x, (int, float)) and x > 0.0001]
    if ctype == 'SAMPLE' and len(valid_injs) == 0:
        out_lines.append(f"ZERO INJ ROW {r:4d} | Seq={seq} | Name={sname} | ST={st} | Depth={dp} | Injs={injs} | Flag={flag} | Comment={comment}")
        zero_injs.append((r, sname, st, dp, injs, flag, comment))

out_lines.append(f"\nTotal 40xxx rows inspected: {len(err_40xxx)}")
out_lines.append(f"Total zero injection SAMPLE rows inspected: {len(zero_injs)}")

with open('scratch/inspect_40xxx_results.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out_lines))

print("Saved to scratch/inspect_40xxx_results.txt")
