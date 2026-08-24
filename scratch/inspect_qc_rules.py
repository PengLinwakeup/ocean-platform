import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

print("================ 1. INSPECTING MQ BLANKS ================")
ws_master = wb["All_Columns_Sequence_QC_Master"]

mq_rows = []
for r in range(6, ws_master.max_row + 1):
    cat = str(ws_master.cell(r, 3).value or "").upper()
    name = str(ws_master.cell(r, 2).value or "").upper()
    if "MQ" in cat or "MQ" in name or "BLANK" in cat or "BLANK" in name:
        area = ws_master.cell(r, 10).value
        rsd = ws_master.cell(r, 11).value
        doc = ws_master.cell(r, 14).value
        flag = ws_master.cell(r, 15).value if ws_master.max_column >= 15 else None
        # let's get diagnosis if available
        mq_rows.append((r, name, area, rsd, doc))

print(f"Total MQ blank rows found in Master: {len(mq_rows)}")
for r, name, area, rsd, doc in mq_rows[:15]:
    print(f"Row {r:4d} | Name: {name:20s} | Area: {area} | RSD%: {rsd} | QC DOC: {doc}")

print("\n================ 2. INSPECTING DSW CRM SAMPLES ================")
dsw_rows = []
for r in range(6, ws_master.max_row + 1):
    cat = str(ws_master.cell(r, 3).value or "").upper()
    name = str(ws_master.cell(r, 2).value or "").upper()
    if "DSW" in cat or "DSW" in name or "CRM" in cat or "CRM" in name:
        area = ws_master.cell(r, 10).value
        rsd = ws_master.cell(r, 11).value
        doc = ws_master.cell(r, 14).value
        dsw_rows.append((r, name, area, rsd, doc))

print(f"Total DSW rows found in Master: {len(dsw_rows)}")
low_dsw = [x for x in dsw_rows if x[4] is not None and isinstance(x[4], (int, float)) and x[4] < 39.0]
print(f"DSW rows with measured DOC < 39.0: {len(low_dsw)}")
for r, name, area, rsd, doc in low_dsw:
    print(f"  Row {r:4d} | Name: {name:25s} | Area: {area} | RSD%: {rsd} | QC DOC: {doc}")

print("\n================ 3. INSPECTING HIGH DOC SAMPLES (> 100 μM) ================")
high_doc_rows = []
ws_all = wb["ODV_All_Samples_Full_List"]
for r in range(5, ws_all.max_row + 1):
    status = ws_all.cell(r, 1).value
    seq = ws_all.cell(r, 2).value
    st = ws_all.cell(r, 3).value
    sid = ws_all.cell(r, 4).value
    depth = ws_all.cell(r, 6).value
    doc = ws_all.cell(r, 11).value or ws_all.cell(r, 7).value
    flag = ws_all.cell(r, 12).value
    comment = ws_all.cell(r, 13).value
    
    if doc is not None and isinstance(doc, (int, float)) and doc > 100.0:
        high_doc_rows.append((r, status, st, sid, depth, doc, flag, comment))

print(f"Total High DOC samples (> 100 μM) in ODV_All_Samples_Full_List: {len(high_doc_rows)}")
for r, status, st, sid, depth, doc, flag, comment in high_doc_rows:
    print(f"Row {r:4d} | Status: {status:18s} | ST: {st:6s} | ID: {sid:22s} | Depth: {depth} | DOC: {doc:.2f} | Flag: {flag} | Comment: {comment}")
