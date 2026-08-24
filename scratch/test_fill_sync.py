import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path)

ws_master = wb["All_Columns_Sequence_QC_Master"]

stale_red_rows = []
for r in range(6, ws_master.max_row + 1):
    c1 = ws_master.cell(r, 1)
    co = ws_master.cell(r, 15)
    flag_val = co.value
    
    c1_fill = c1.fill.fgColor.value if (c1.fill and c1.fill.fgColor) else None
    co_fill = co.fill.fgColor.value if (co.fill and co.fill.fgColor) else None
    
    if flag_val == 2 and (c1_fill == "FFFEE2E2" or co_fill == "FFFEE2E2"):
        stale_red_rows.append((r, ws_master.cell(r, 2).value, ws_master.cell(r, 16).value))

print(f"Total Flag 2 rows in Master sheet with leftover RED fill: {len(stale_red_rows)}")
for r, name, comment in stale_red_rows[:15]:
    print(f"Row {r:4d} | Name: {name:20s} | Comment: {comment}")
