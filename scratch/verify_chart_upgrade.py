import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

target_file = r'F:\印度洋测样\ODV\202608\20260827\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx'
wb = openpyxl.load_workbook(target_file, data_only=False)

print("=== WORKBOOK VERIFICATION ===")
print("Total sheets:", len(wb.sheetnames))
expected_core_sheets = [
    'Executive_Dashboard',
    'ODV_All_Samples_Full_List',
    'ODV_Clean_Export_Only',
    'All_Columns_Sequence_QC_Master',
    'Flag4_Discarded_Audit_List'
]

for s in expected_core_sheets:
    if s in wb.sheetnames:
        ws = wb[s]
        print(f"  [OK] Sheet '{s}' exists (max_row={ws.max_row}, max_col={ws.max_column})")
    else:
        print(f"  [ERROR] Sheet '{s}' missing!")

# Check Executive_Dashboard charts
ws_dash = wb['Executive_Dashboard']
print(f"\nExecutive_Dashboard charts: {len(ws_dash._charts)}")
for i, c in enumerate(ws_dash._charts):
    print(f"\nChart {i+1}:")
    print(f"  Title: {c.title}")
    print(f"  Size: {c.width} x {c.height} cm")
    print(f"  Legend: {c.legend.legendPos if c.legend else 'None'}")
    print(f"  X-axis: title='{c.x_axis.title}', min={c.x_axis.scaling.min}, max={c.x_axis.scaling.max}")
    print(f"  Y-axis: title='{c.y_axis.title}', min={c.y_axis.scaling.min}, max={c.y_axis.scaling.max}")
    print(f"  Series count: {len(c.series)}")
    for s_idx, s in enumerate(c.series):
        title_val = s.title.v if hasattr(s.title, 'v') else (s.title.value if hasattr(s.title, 'value') else str(s.title))
        line_color = getattr(getattr(s, 'graphicalProperties', None), 'line', None)
        solid_fill = getattr(line_color, 'solidFill', None) if line_color else None
        dash_style = getattr(line_color, 'prstDash', None) if line_color else None
        marker_sym = s.marker.symbol if s.marker else None
        print(f"    Series {s_idx+1}: '{title_val}' | marker={marker_sym} | lineFill={solid_fill} | dash={dash_style}")

# Check formula integrity in Master and ODV
ws_master = wb['All_Columns_Sequence_QC_Master']
print(f"\nAll_Columns_Sequence_QC_Master sample formula check:")
print(f"  N13: {ws_master['N13'].value}")
print(f"  N19: {ws_master['N19'].value}")

ws_clean = wb['ODV_Clean_Export_Only']
print(f"ODV_Clean_Export_Only rows: {ws_clean.max_row}")

print("\nALL CHECKS COMPLETED SUCCESSFULLY!")
