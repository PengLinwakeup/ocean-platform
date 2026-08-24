import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n================ SHEET: {sheet_name} ================")
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(15, ws.max_column + 1))]
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}: {row_vals}")
