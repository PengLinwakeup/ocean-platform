import openpyxl
import re
import os
import shutil

src_file = r"temp_geomar_v2_export.xlsx"
out_file = r"F:\印度洋测样\ODV\202608\20260822\Ocean_DOC_MultiColumn_QC_Report_GEOMAR-再处理版_latest.xlsx"

print(f"Loading user original file: {src_file}")
wb = openpyxl.load_workbook(src_file, data_only=False)

# 1. 从 temp_geomar_v2_input.json 或 All_Columns_Sequence_QC_Master 提取每个序列的精准斜率与 R2
batch_slopes = {}
batch_r2s = {}

if os.path.exists("temp_geomar_v2_input.json"):
    with open("temp_geomar_v2_input.json", "r", encoding="utf-8") as f:
        _jdata = json.load(f)
        for _idx, _b in enumerate(_jdata.get("batches", [])):
            _seq_n = _idx + 1
            _s = float(_b.get("slope", 0.0554))
            _r = float(_b.get("rsq") or _b.get("r2", 0.999))
            batch_slopes[_seq_n] = _s
            batch_r2s[_seq_n] = _r

if "All_Columns_Sequence_QC_Master" in wb.sheetnames:
    ws_m = wb["All_Columns_Sequence_QC_Master"]
    curr_seq_n = None
    for r in range(1, ws_m.max_row + 1):
        val1 = str(ws_m.cell(r, 1).value or "").strip()
        seq_m = re.search(r'【序列\s*(\d+)/26】', val1)
        if seq_m:
            curr_seq_n = int(seq_m.group(1))
        if "斜率:" in val1:
            slope_m = re.search(r'斜率:\s*([\d\.]+)', val1)
            if slope_m and curr_seq_n:
                batch_slopes[curr_seq_n] = float(slope_m.group(1))

# 2. 站位/深度与 Sample ID 精准解析与修复逻辑
def process_sample_name_id_st_depth(name_val, id_val, st_val, d_val, prev_id_num=None):
    clean_name = str(name_val or '').strip()
    clean_id = str(id_val or '').strip()
    target_str = clean_name if clean_name else clean_id
    
    parsed_st = st_val
    parsed_d = d_val
    curr_id_num = None

    # 末尾 -STxx-depth 智能识别
    st_match = re.search(r'(?:ST|ST-)(\d+)[-_](\d+)', target_str, flags=re.I)
    if st_match:
        st_num = int(st_match.group(1))
        parsed_st = f"ST-{st_num}"
        try: parsed_d = float(st_match.group(2))
        except ValueError: pass

    # 前缀 41163 -> 41063 错号纠正
    prefix_match = re.match(r'^([A-Za-z0-9]+-)(\d{4,6})(-.*)$', target_str, flags=re.I)
    if prefix_match:
        prefix_head = prefix_match.group(1)
        num_str = prefix_match.group(2)
        prefix_tail = prefix_match.group(3)
        raw_num = int(num_str)

        if prev_id_num is not None:
            if abs(raw_num - prev_id_num) > 5 and abs(raw_num - prev_id_num) < 500:
                fixed_num = prev_id_num + 1
            else:
                fixed_num = raw_num
        else:
            fixed_num = raw_num
            
        curr_id_num = fixed_num
        num_len = len(num_str)
        fixed_num_str = f"{fixed_num:0{num_len}d}"
        fixed_str = f"{prefix_head}{fixed_num_str}{prefix_tail}"
        
        if clean_name and re.match(r'^[A-Za-z0-9]+-\d{4,6}-', clean_name, flags=re.I):
            clean_name = fixed_str
        if clean_id and re.match(r'^[A-Za-z0-9]+-\d{4,6}-', clean_id, flags=re.I):
            clean_id = fixed_str
    else:
        id_num_match = re.search(r'\d{4,6}', target_str)
        if id_num_match:
            curr_id_num = int(id_num_match.group(0))

    return clean_name, clean_id, parsed_st, parsed_d, curr_id_num

# 3. 处理 ODV_All_Samples_Full_List 表
if "ODV_All_Samples_Full_List" in wb.sheetnames:
    ws = wb["ODV_All_Samples_Full_List"]
    last_id_num = None
    curr_slope = 0.0554
    
    for r in range(5, ws.max_row + 1):
        seq_name = str(ws.cell(r, 2).value or "").strip()
        st_cell = ws.cell(r, 3)
        id_cell = ws.cell(r, 4)
        d_cell = ws.cell(r, 6)
        raw_doc_cell = ws.cell(r, 7)
        
        # 寻找对应的 slope
        slope_v = 0.0554
        for k, v in batch_slopes.items():
            if seq_name in k or k in seq_name:
                slope_v = v
                break
                
        # 修补 Station & Depth & Sample ID
        name_v, id_v, fixed_st, fixed_d, curr_num = process_sample_name_id_st_depth(
            id_cell.value, id_cell.value, st_cell.value, d_cell.value, last_id_num
        )
        if curr_num: last_id_num = curr_num
        
        id_cell.value = id_v
        st_cell.value = fixed_st
        if fixed_d is not None: d_cell.value = fixed_d
        
        # 插入 Raw DOC 动态活公式
        # 列 7 = Raw DOC, 列 9 = Clean Mean Area
        raw_doc_cell.value = f"=IF({slope_v:.6f}>0, MAX(0, (I{r} - 0.000000) / {slope_v:.6f}), 0)"
        raw_doc_cell.number_format = '0.00'

# 4. 处理 ODV_Clean_Export_Only 表
if "ODV_Clean_Export_Only" in wb.sheetnames:
    ws = wb["ODV_Clean_Export_Only"]
    last_id_num = None
    for r in range(5, ws.max_row + 1):
        seq_name = str(ws.cell(r, 1).value or "").strip()
        st_cell = ws.cell(r, 2)
        id_cell = ws.cell(r, 3)
        d_cell = ws.cell(r, 5)
        raw_doc_cell = ws.cell(r, 6)
        
        slope_v = 0.0554
        for k, v in batch_slopes.items():
            if seq_name in k or k in seq_name:
                slope_v = v
                break
                
        name_v, id_v, fixed_st, fixed_d, curr_num = process_sample_name_id_st_depth(
            id_cell.value, id_cell.value, st_cell.value, d_cell.value, last_id_num
        )
        if curr_num: last_id_num = curr_num
        
        id_cell.value = id_v
        st_cell.value = fixed_st
        if fixed_d is not None: d_cell.value = fixed_d
        
        # 列 6 = Raw DOC, 列 8 = Clean Mean Area
        raw_doc_cell.value = f"=IF({slope_v:.6f}>0, MAX(0, (H{r} - 0.000000) / {slope_v:.6f}), 0)"
        raw_doc_cell.number_format = '0.00'

# 5. 处理 ODV_Discarded_Samples_Flag4 表
if "ODV_Discarded_Samples_Flag4" in wb.sheetnames:
    ws = wb["ODV_Discarded_Samples_Flag4"]
    last_id_num = None
    for r in range(5, ws.max_row + 1):
        seq_name = str(ws.cell(r, 1).value or "").strip()
        st_cell = ws.cell(r, 2)
        id_cell = ws.cell(r, 3)
        d_cell = ws.cell(r, 5)
        raw_doc_cell = ws.cell(r, 6)
        
        slope_v = 0.0554
        for k, v in batch_slopes.items():
            if seq_name in k or k in seq_name:
                slope_v = v
                break
                
        name_v, id_v, fixed_st, fixed_d, curr_num = process_sample_name_id_st_depth(
            id_cell.value, id_cell.value, st_cell.value, d_cell.value, last_id_num
        )
        if curr_num: last_id_num = curr_num
        
        id_cell.value = id_v
        st_cell.value = fixed_st
        if fixed_d is not None: d_cell.value = fixed_d
        
        raw_doc_cell.value = f"=IF({slope_v:.6f}>0, MAX(0, (H{r} - 0.000000) / {slope_v:.6f}), 0)"
        raw_doc_cell.number_format = '0.00'

# 6. 处理 All_Columns_Sequence_QC_Master 表
if "All_Columns_Sequence_QC_Master" in wb.sheetnames:
    ws = wb["All_Columns_Sequence_QC_Master"]
    curr_slope = 0.0554
    last_id_num = None
    for r in range(1, ws.max_row + 1):
        val1 = str(ws.cell(r, 1).value or "").strip()
        if "【序列" in val1:
            slope_m = re.search(r'斜率:\s*([\d\.]+)', val1)
            if slope_m: curr_slope = float(slope_m.group(1))
            last_id_num = None
            continue
            
        if ws.cell(r, 1).value is not None and isinstance(ws.cell(r, 1).value, (int, float)):
            name_cell = ws.cell(r, 2)
            st_cell = ws.cell(r, 4)
            d_cell = ws.cell(r, 5)
            raw_doc_cell = ws.cell(r, 12)
            
            name_v, id_v, fixed_st, fixed_d, curr_num = process_sample_name_id_st_depth(
                name_cell.value, name_cell.value, st_cell.value, d_cell.value, last_id_num
            )
            if curr_num: last_id_num = curr_num
            
            name_cell.value = name_v
            st_cell.value = fixed_st
            if fixed_d is not None: d_cell.value = fixed_d
            
            # 列 12 = Raw DOC, 列 10 = Clean Mean Area
            raw_doc_cell.value = f"=IF({curr_slope:.6f}>0, MAX(0, (J{r} - 0.000000) / {curr_slope:.6f}), 0)"
            raw_doc_cell.number_format = '0.00'

# 7. 保存到目标输出路径
try:
    wb.save(out_file)
    print(f"Successfully saved user original restored file with formulas to: {out_file}")
except PermissionError:
    alt_out = r"F:\印度洋测样\ODV\202608\20260822\Ocean_DOC_MultiColumn_QC_Report_GEOMAR-再处理版_user_restored.xlsx"
    wb.save(alt_out)
    print(f"Target file locked, saved to alternative path: {alt_out}")
