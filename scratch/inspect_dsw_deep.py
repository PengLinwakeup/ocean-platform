import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

ws_master = wb["All_Columns_Sequence_QC_Master"]

print("================ INSPECTING ALL DSW CRM ROWS IN MASTER ================")
dsw_rows = []
for r in range(6, ws_master.max_row + 1):
    c1_val = ws_master.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master.cell(r, 2).value or "").strip()
        cat_type = str(ws_master.cell(r, 3).value or "").strip().upper()
        
        if "DSW" in cat_type or "DSW" in s_name.upper() or "CRM" in cat_type or "CRM" in s_name.upper():
            inj1 = ws_master.cell(r, 6).value
            inj2 = ws_master.cell(r, 7).value
            inj3 = ws_master.cell(r, 8).value
            inj4 = ws_master.cell(r, 9).value
            area = ws_master.cell(r, 10).value
            rsd = ws_master.cell(r, 11).value
            raw_doc = ws_master.cell(r, 12).value
            mq_drift = ws_master.cell(r, 13).value
            qc_doc = ws_master.cell(r, 14).value
            flag = ws_master.cell(r, 15).value
            comment = ws_master.cell(r, 16).value
            
            dsw_rows.append((r, s_name, area, rsd, raw_doc, qc_doc, flag, comment, [inj1, inj2, inj3, inj4]))

print(f"Total DSW rows in Master: {len(dsw_rows)}")
low_or_zero_dsw = [x for x in dsw_rows if x[5] is None or not isinstance(x[5], (int, float)) or x[5] < 39.0]
print(f"DSW rows with DOC < 39.0 or None: {len(low_or_zero_dsw)}")

print("\nSample DSW rows:")
for r, s_name, area, rsd, raw_doc, qc_doc, flag, comment, injs in dsw_rows[:25]:
    print(f"Row {r:4d} | Name: {s_name:10s} | Area: {area} | Raw DOC: {raw_doc} | QC DOC: {qc_doc} | Flag: {flag} | Comment: {comment}")
