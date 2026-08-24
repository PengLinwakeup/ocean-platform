import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

ws_all = wb["ODV_All_Samples_Full_List"]
print(f"Max col in ODV_All_Samples_Full_List: {ws_all.max_column}")
print("Row 4 headers:")
print([ws_all.cell(4, c).value for c in range(1, ws_all.max_column + 1)])

print("\nSearching for SO308-41446 in ODV_All_Samples_Full_List:")
for r in range(5, ws_all.max_row + 1):
    vals = [ws_all.cell(r, c).value for c in range(1, ws_all.max_column + 1)]
    row_str = " ".join([str(v) for v in vals if v is not None])
    if "41446" in row_str:
        print(f"Row {r:4d}: {vals}")

ws_master = wb["All_Columns_Sequence_QC_Master"]
print("\nSearching for 41446 in Master:")
for r in range(6, ws_master.max_row + 1):
    vals = [ws_master.cell(r, c).value for c in range(1, ws_master.max_column + 1)]
    row_str = " ".join([str(v) for v in vals if v is not None])
    if "41446" in row_str:
        print(f"Row {r:4d}: {vals}")
