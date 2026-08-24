import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path)

ws_master = wb["All_Columns_Sequence_QC_Master"]

print("Conditional formatting rules in ws_master:")
print(ws_master.conditional_formatting)

print("\nInspecting Rows 37, 38, 39, 40 cell fills in ws_master:")
for r in [37, 38, 39, 40]:
    c1 = ws_master.cell(r, 1)
    co = ws_master.cell(r, 15)
    flag_val = co.value
    comment_val = ws_master.cell(r, 16).value
    print(f"Row {r:3d} | Flag: {flag_val} | Col 1 Fill: {c1.fill.fill_type}, fgColor: {c1.fill.fgColor.value if c1.fill else None} | Col O Fill: {co.fill.fill_type}, fgColor: {co.fill.fgColor.value if co.fill else None} | Comment: {comment_val}")
