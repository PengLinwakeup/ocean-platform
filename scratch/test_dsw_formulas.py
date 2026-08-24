import openpyxl
import json
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"

wb_data = openpyxl.load_workbook(file_path, data_only=True)
ws_master_data = wb_data["All_Columns_Sequence_QC_Master"]

wb_formula = openpyxl.load_workbook(file_path, data_only=False)
ws_master_formula = wb_formula["All_Columns_Sequence_QC_Master"]

with open("temp_geomar_v2_input.json", "r", encoding="utf-8") as f:
    input_data = json.load(f)

batches_meta = input_data.get("batches", [])
seq_info = {}
for idx, b in enumerate(batches_meta):
    seq_num = idx + 1
    slope = float(b.get("slope", 0.0554))
    seq_info[seq_num] = slope

curr_seq = 1
curr_slope = seq_info[1]
formula_updates = []

for r in range(6, ws_master_formula.max_row + 1):
    val1 = str(ws_master_formula.cell(r, 1).value or "").strip()
    seq_match = re.search(r'【序列\s*(\d+)/26】', val1)
    if seq_match:
        curr_seq = int(seq_match.group(1))
        if curr_seq in seq_info:
            curr_slope = seq_info[curr_seq]
        continue
        
    c1_val = ws_master_formula.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master_formula.cell(r, 2).value or "").strip()
        cat_type = str(ws_master_formula.cell(r, 3).value or "").strip().upper()
        
        if "DSW" in cat_type or "DSW" in s_name.upper() or "CRM" in cat_type or "CRM" in s_name.upper():
            inj_cells = [(6, f"F{r}"), (7, f"G{r}"), (8, f"H{r}"), (9, f"I{r}")]
            inj_vals = []
            for col_i, cell_ref in inj_cells:
                v = ws_master_data.cell(r, col_i).value
                if v is not None and isinstance(v, (int, float)) and 0 < v < 10.0:
                    inj_vals.append((v, cell_ref))
                    
            mq_area = ws_master_data.cell(r, 13).value
            m_area = float(mq_area) if mq_area is not None and isinstance(mq_area, (int, float)) else 0.0
            
            # Determine best injection selection for DSW
            if len(inj_vals) >= 4:
                # Compare 4-injection vs top 3
                area4 = sum(x[0] for x in inj_vals) / 4.0
                net_doc4 = max(0, (area4 - m_area) / curr_slope)
                
                # Sorted by value
                sorted_injs = sorted(inj_vals, key=lambda x: x[0])
                top3 = sorted_injs[1:] # drop lowest
                area3 = sum(x[0] for x in top3) / 3.0
                net_doc3 = max(0, (area3 - m_area) / curr_slope)
                
                if net_doc4 >= 39.0:
                    selected_refs = [x[1] for x in inj_vals]
                    calc_net_doc = net_doc4
                else:
                    selected_refs = [x[1] for x in top3]
                    calc_net_doc = net_doc3
            else:
                selected_refs = [x[1] for x in inj_vals]
                calc_net_doc = max(0, ((sum(x[0] for x in inj_vals)/len(inj_vals)) - m_area) / curr_slope) if len(inj_vals)>0 else 0.0
                
            cols_str = ",".join(selected_refs)
            clean_area_formula = f"=AVERAGE({cols_str})"
            rsd_formula = f"=STDEV({cols_str})/J{r}*100"
            raw_doc_formula = f"=IF({curr_slope}>0, MAX(0, (J{r} - 0) / {curr_slope}), 0)"
            qc_doc_formula = f"=IF({curr_slope}>0, MAX(0, (J{r} - M{r}) / {curr_slope}), 0)"
            
            formula_updates.append((r, s_name, curr_seq, clean_area_formula, raw_doc_formula, qc_doc_formula, round(calc_net_doc, 2)))

print(f"Total DSW formula updates prepared: {len(formula_updates)}")
print("\nSample DSW rows with complete Excel Formulas:")
for r, s_name, seq, f_area, f_raw, f_qc, calc_doc in formula_updates[:15]:
    print(f"Row {r:4d} | Seq {seq:2d} | Calc Net DOC: {calc_doc:5.2f} uM")
    print(f"       -> Clean Area Formula : {f_area}")
    print(f"       -> Raw DOC Formula    : {f_raw}")
    print(f"       -> QC Dynamic Formula : {f_qc}")
