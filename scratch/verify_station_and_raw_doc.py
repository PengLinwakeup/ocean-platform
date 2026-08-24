import sys
import os
sys.path.append(os.path.abspath('.'))

from run_geomar_qc_processor_20260820 import extract_station_depth_and_fix_id, parse_web_exported_excel, build_geomar_master_excel
import openpyxl

print("=== 1. Testing extract_station_depth_and_fix_id ===")
test_cases = [
    ("SO308-41060-ST35-5540", "SO308-41060-ST35-5540", None),
    ("SO308-41061-ST35-5400", "SO308-41061-ST35-5400", 41060),
    ("SO308-41062-ST35-4800", "SO308-41062-ST35-4800", 41061),
    ("SO308-41163-ST35-4000", "SO308-41163-ST35-4000", 41062), # Erroneous 41163!
    ("50308-41163-ST35-4000", "50308-41163-ST35-4000", 41062), # 50308 variant
]

for name, sid, prev_num in test_cases:
    fname, fid, st, d, curr_num = extract_station_depth_and_fix_id(name, sid, prev_num)
    print(f"Input: {name} (prev={prev_num}) -> Station: {st}, Depth: {d}, Fixed Name: {fname}, Fixed ID: {fid}, Current Num: {curr_num}")

print("\n=== 2. Testing Excel File Generation with Live Input ===")
input_excel = r"F:\印度洋测样\ODV\202608\20260822\Ocean_DOC_MultiColumn_QC_Report_GEOMAR-再处理版_latest.xlsx"
if not os.path.exists(input_excel):
    input_excel = r"F:\印度洋测样\ODV\202608\20260822\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_latest.xlsx"
if not os.path.exists(input_excel):
    print("Finding alternative excel files in F:\\印度洋测样\\ODV\\202608\\20260822...")
    import glob
    candidates = glob.glob(r"F:\印度洋测样\ODV\202608\20260822\*.xlsx")
    if candidates:
        input_excel = candidates[0]

print("Using input file:", input_excel)
if os.path.exists(input_excel):
    batches = parse_web_exported_excel(input_excel)
    print(f"Parsed {len(batches)} sequence batches.")
    
    out_file = "scratch/test_output_qc_report.xlsx"
    build_geomar_master_excel(batches, out_file)
    print("Successfully built output file:", out_file)
    
    wb = openpyxl.load_workbook(out_file, data_only=False)
    ws_master = wb["All_Columns_Sequence_QC_Master"]
    
    # Check row 7 (first data row in master) Col L (12)
    print("Formula in Master Col L row 7:", ws_master.cell(7, 12).value)
    
    ws_all = wb["ODV_All_Samples_Full_List"]
    print("Formula in ODV_All_Samples Col G row 5:", ws_all.cell(5, 7).value)
    
    ws_clean = wb["ODV_Clean_Export_Only"]
    print("Formula in ODV_Clean_Export Col F row 5:", ws_clean.cell(5, 6).value)

print("\nVerification Complete.")
