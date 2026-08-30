import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import glob
import re
from datetime import datetime

print("========================================================================")
print("     SO308 DOC QC Report Data Cleaning & Style Calibration Tool")
print("========================================================================")

target_excel = r'F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829_Final_QC_Completed.xlsx'
sample_list_file = r'F:\印度洋测样\Indian Ocean_SO308_DOC_Sample List(1) 的副本.xlsx'

if not os.path.exists(target_excel):
    raise FileNotFoundError(f"Target Excel file not found: {target_excel}")
if not os.path.exists(sample_list_file):
    raise FileNotFoundError(f"Sample list file not found: {sample_list_file}")

# 1. Build (Station, Depth) -> Correct_ID Map from Raw Sample List
wb_sample = openpyxl.load_workbook(sample_list_file, data_only=True)
ws_sample = wb_sample['Sheet1']

sample_map = {}
for r in range(2, ws_sample.max_row + 1):
    st = ws_sample.cell(r, 3).value
    dp = ws_sample.cell(r, 6).value
    lbl = ws_sample.cell(r, 7).value
    if st and dp is not None and lbl:
        st_str = str(st).strip().upper().replace(' ', '')
        if not st_str.startswith('ST-') and st_str.startswith('ST'):
            st_key = 'ST-' + st_str[2:]
        else:
            st_key = st_str
            
        try:
            dp_key = int(round(float(dp)))
        except:
            dp_key = str(dp).strip()
            
        lbl_clean = str(lbl).strip()
        sample_map[(st_key, dp_key)] = lbl_clean
        st_alt = st_key.replace('ST-', 'ST')
        sample_map[(st_alt, dp_key)] = lbl_clean

print(f"[*] Built composite map with {len(sample_map)} entries from Raw Sample List.")

# 2. Load Target Workbook with Formulas Preserved (data_only=False)
print(f"[*] Loading Target Workbook (formulas preserved): {target_excel}")
wb_target = openpyxl.load_workbook(target_excel, data_only=False)
ws_master = wb_target['All_Columns_Sequence_QC_Master']

# Standard Green Font for Flag 2 Seq Order (Col 1): ARGB FF166534
green_font = Font(name='Calibri', size=11, bold=True, color='FF166534')
no_fill = PatternFill(fill_type=None)
flag2_fill = PatternFill(fill_type='solid', fgColor='FFDCFCE7')

stats = {
    'style_cleaned': 0,
    'id_corrected': 0,
    'zero_inj_cleaned': 0,
    'total_rows_scanned': 0
}

logs = []

def get_clean_station_key(st_val):
    if not st_val:
        return ''
    s = str(st_val).strip().upper().replace(' ', '')
    if not s.startswith('ST-') and s.startswith('ST'):
        return 'ST-' + s[2:]
    return s

def get_clean_depth_key(dp_val):
    if dp_val is None:
        return None
    try:
        return int(round(float(dp_val)))
    except:
        return str(dp_val).strip()

print("[*] Processing Master Sheet rows...")

for r in range(7, ws_master.max_row + 1):
    seq_cell = ws_master.cell(r, 1)
    name_cell = ws_master.cell(r, 2)
    ctype_cell = ws_master.cell(r, 3)
    st_cell = ws_master.cell(r, 4)
    dp_cell = ws_master.cell(r, 5)
    
    inj1_cell = ws_master.cell(r, 6)
    inj2_cell = ws_master.cell(r, 7)
    inj3_cell = ws_master.cell(r, 8)
    inj4_cell = ws_master.cell(r, 9)
    
    flag_cell = ws_master.cell(r, 15)
    comment_cell = ws_master.cell(r, 16)
    
    if not any([seq_cell.value, name_cell.value, ctype_cell.value]):
        continue
        
    stats['total_rows_scanned'] += 1
    
    flag_val = str(flag_cell.value).strip() if flag_cell.value is not None else ''
    comment_val = str(comment_cell.value).strip() if comment_cell.value is not None else ''
    is_acceptable = (flag_val == '2' or 'Acceptable' in comment_val or flag_val == '2.0')
    
    # ------------------------------------------------------------------------
    # TASK 1: Style & Color Calibration (Flag 2 qualified rows)
    # ------------------------------------------------------------------------
    if is_acceptable:
        # Clear warning background fill for Col 1..14 and Col 16 (exclude Col 15 Flag)
        for c in range(1, 15):
            cell = ws_master.cell(r, c)
            if cell.fill and cell.fill.fill_type:
                cell.fill = no_fill
        
        if comment_cell.fill and comment_cell.fill.fill_type:
            comment_cell.fill = no_fill
            
        # Fix Col 1 Seq Order font and fill
        seq_cell.font = green_font
        seq_cell.fill = no_fill
        
        # Ensure Col 15 Flag retains light green fill
        flag_cell.fill = flag2_fill
        
        stats['style_cleaned'] += 1

    # ------------------------------------------------------------------------
    # TASK 2: Sample ID Correction (40xxx & Mismatches via Station & Depth)
    # ------------------------------------------------------------------------
    ctype_val = str(ctype_cell.value).strip().upper() if ctype_cell.value else ''
    if ctype_val == 'SAMPLE':
        st_key = get_clean_station_key(st_cell.value)
        dp_key = get_clean_depth_key(dp_cell.value)
        
        if st_key and dp_key is not None:
            correct_id = sample_map.get((st_key, dp_key))
            if correct_id:
                curr_name = str(name_cell.value).strip() if name_cell.value else ''
                
                parts = curr_name.split('-')
                if len(parts) >= 4 and (parts[0].startswith('SO') or parts[0].startswith('50')):
                    st_suffix = '-'.join(parts[2:])
                    new_name = f"{correct_id}-{st_suffix}"
                else:
                    new_name = correct_id
                    
                if curr_name != new_name:
                    name_cell.value = new_name
                    stats['id_corrected'] += 1
                    logs.append(f"[ID FIX] Row {r:4d} | Station: {st_key:<6s} Depth: {dp_key:<5d} | '{curr_name}' -> '{new_name}'")

    # ------------------------------------------------------------------------
    # TASK 3: Zero Injection / Dry Draw Data Safe Cleaning
    # ------------------------------------------------------------------------
    if ctype_val == 'SAMPLE':
        injs = [inj1_cell.value, inj2_cell.value, inj3_cell.value, inj4_cell.value]
        valid_injs = [x for x in injs if x is not None and isinstance(x, (int, float)) and x > 0.0001]
        
        if len(valid_injs) == 0:
            comment_cell.value = "Discarded: Zero injection dry draw (re-injected in later sequence)"
            flag_cell.value = 4
            discard_fill = PatternFill(fill_type='solid', fgColor='FFFEE2E2')
            flag_cell.fill = discard_fill
            
            stats['zero_inj_cleaned'] += 1
            logs.append(f"[ZERO INJ] Row {r:4d} | Sample: {name_cell.value} | Station: {st_cell.value} Depth: {dp_cell.value} | Flag set to 4 (Discarded)")

# 3. Save as Timestamped Output File
timestamp = datetime.now().strftime('%Y%m%d_%H%M')
out_dir = os.path.dirname(target_excel)
out_filename = f"Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829_Final_QC_Completed_Cleaned_{timestamp}.xlsx"
out_filepath = os.path.join(out_dir, out_filename)

print(f"\n[*] Saving updated workbook to: {out_filepath}")
wb_target.save(out_filepath)

print("\n========================================================================")
print("                        PROCESSING SUMMARY")
print("========================================================================")
print(f"Total Rows Scanned       : {stats['total_rows_scanned']}")
print(f"Qualified Style Cleaned  : {stats['style_cleaned']}")
print(f"Sample IDs Corrected     : {stats['id_corrected']}")
print(f"Zero Injections Cleaned  : {stats['zero_inj_cleaned']}")
print("========================================================================")

log_path = 'scratch/cleaning_execution_log.txt'
with open(log_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(logs))

print(f"Detailed change log written to {log_path}")
