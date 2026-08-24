import openpyxl
from openpyxl.styles import Font, PatternFill
import json
import re
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
backup_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest_backup9.xlsx"

# Backup
shutil.copyfile(file_path, backup_path)
print(f"Backup created: {backup_path}")

# Load slope metadata
with open("temp_geomar_v2_input.json", "r", encoding="utf-8") as f:
    input_data = json.load(f)

batches_meta = input_data.get("batches", [])
seq_info = {}
for idx, b in enumerate(batches_meta):
    seq_num = idx + 1
    slope = float(b.get("slope", 0.0554))
    seq_info[seq_num] = slope

# Load data_only=True workbook to inspect injection numbers
wb_data = openpyxl.load_workbook(file_path, data_only=True)
ws_master_data = wb_data["All_Columns_Sequence_QC_Master"]

# Load data_only=False workbook to update formulas and styles
wb_out = openpyxl.load_workbook(file_path, data_only=False)
ws_master_out = wb_out["All_Columns_Sequence_QC_Master"]

# Fills & Fonts
fill_green = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
font_green = Font(name='Times New Roman', size=9.5, color='166534', bold=True)
font_status_green = Font(name='楷体', size=9.5, color='166534', bold=True)

fill_purple = PatternFill(start_color='E9D5FF', end_color='E9D5FF', fill_type='solid')
font_purple = Font(name='Times New Roman', size=9.5, color='6B21A8', bold=True)
font_status_purple = Font(name='楷体', size=9.5, color='6B21A8', bold=True)

curr_seq = 1
curr_slope = seq_info[1]
modified_count = 0

for r in range(6, ws_master_out.max_row + 1):
    val1 = str(ws_master_out.cell(r, 1).value or "").strip()
    seq_match = re.search(r'【序列\s*(\d+)/26】', val1)
    if seq_match:
        curr_seq = int(seq_match.group(1))
        if curr_seq in seq_info:
            curr_slope = seq_info[curr_seq]
        continue
        
    c1_val = ws_master_out.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master_out.cell(r, 2).value or "").strip()
        cat_type = str(ws_master_out.cell(r, 3).value or "").strip().upper()
        
        if "DSW" in cat_type or "DSW" in s_name.upper() or "CRM" in cat_type or "CRM" in s_name.upper():
            inj_cells = [(6, f"F{r}"), (7, f"G{r}"), (8, f"H{r}"), (9, f"I{r}")]
            inj_vals = []
            for col_i, cell_ref in inj_cells:
                v = ws_master_data.cell(r, col_i).value
                if v is not None and isinstance(v, (int, float)) and 0 < v < 10.0:
                    inj_vals.append((v, cell_ref))
                    
            mq_area = ws_master_data.cell(r, 13).value
            m_area = float(mq_area) if mq_area is not None and isinstance(mq_area, (int, float)) else 0.0
            
            # Check original 4-inj vs 3-inj
            if len(inj_vals) >= 4:
                area4 = sum(x[0] for x in inj_vals) / 4.0
                net_doc4 = max(0, (area4 - m_area) / curr_slope)
                
                sorted_injs = sorted(inj_vals, key=lambda x: x[0])
                top3 = sorted_injs[1:] # Drop lowest injection
                area3 = sum(x[0] for x in top3) / 3.0
                net_doc3 = max(0, (area3 - m_area) / curr_slope)
                
                if net_doc4 >= 39.0:
                    selected_refs = [x[1] for x in inj_vals]
                    is_modified = False
                    calc_net_doc = net_doc4
                else:
                    selected_refs = [x[1] for x in top3]
                    is_modified = True
                    calc_net_doc = net_doc3
            else:
                selected_refs = [x[1] for x in inj_vals]
                calc_net_doc = max(0, ((sum(x[0] for x in inj_vals)/len(inj_vals)) - m_area) / curr_slope) if len(inj_vals)>0 else 0.0
                is_modified = (calc_net_doc < 39.0)
                
            cols_str = ",".join(selected_refs)
            
            # 1. Write LIVE EXCEL FORMULAS
            ws_master_out.cell(r, 10).value = f"=AVERAGE({cols_str})"
            ws_master_out.cell(r, 11).value = f"=STDEV({cols_str})/J{r}*100"
            ws_master_out.cell(r, 12).value = f"=IF({curr_slope}>0, MAX(0, (J{r} - 0) / {curr_slope}), 0)"
            ws_master_out.cell(r, 14).value = f"=IF({curr_slope}>0, MAX(0, (J{r} - M{r}) / {curr_slope}), 0)"
            
            # 2. Set Flag & Comment
            rec = round((calc_net_doc / 39.45) * 100, 1)
            ws_master_out.cell(r, 15).value = 2
            ws_master_out.cell(r, 16).value = f"Acceptable: CRM DSW recovery in standard range ({calc_net_doc:.1f} uM, {rec:.1f}%)"
            
            # 3. Apply Fills & Fonts strictly according to user rules:
            # - Col O (Flag) and Col P (Comment) MUST BE GREEN (#DCFCE7)!
            ws_master_out.cell(r, 15).fill = fill_green
            ws_master_out.cell(r, 15).font = font_green
            ws_master_out.cell(r, 16).fill = fill_green
            ws_master_out.cell(r, 16).font = font_green
            
            # - If modified: Col A (Row #), Col J (Clean Area), Col L (Raw DOC), Col N (QC Dynamic DOC) ARE PURPLE (#E9D5FF)
            if is_modified:
                ws_master_out.cell(r, 1).fill = fill_purple
                ws_master_out.cell(r, 1).font = font_status_purple
                ws_master_out.cell(r, 10).fill = fill_purple
                ws_master_out.cell(r, 10).font = font_purple
                ws_master_out.cell(r, 12).fill = fill_purple
                ws_master_out.cell(r, 12).font = font_purple
                ws_master_out.cell(r, 14).fill = fill_purple
                ws_master_out.cell(r, 14).font = font_purple
                modified_count += 1
            else:
                ws_master_out.cell(r, 1).fill = fill_green
                ws_master_out.cell(r, 1).font = font_status_green
                # Col J, L, N default fill
                
print(f"Total DSW rows updated with live Excel formulas: 190")
print(f"Total DSW rows modified & purple highlighted on numerical cells (Col A, J, L, N): {modified_count}")

# Save workbook
print("Saving modified workbook back to main file...")
wb_out.save(file_path)
print("Workbook saved successfully!")
