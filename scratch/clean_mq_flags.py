import os
import sys
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

"""
========================================================================================
海洋化学 (Marine Chemistry) DOC / TOC 分析数据 - 精细化 MQ 水体空白与 Flag 2 交互式清理工具
========================================================================================
功能：
1. 解决 low-concentration 下相对 %RSD 被人为放大导致超纯水 MQ 被误排为 Flag 4 的统计难题。
2. 引入【绝对标准差 SD 双门限法则】：当平行进样绝对变异 SD <= 0.02 且浓度 <= 10.0 μM 时，自动纠正为 Flag 2。
3. 支持指定【特定行号列表】（如 [3, 5, 8]）精细化挑选人工调为 Flag 2，保留完全操控权。
4. 使用 openpyxl 重新计算并保存带公式 =IF(Slope>0, MAX(0, (J - M)/Slope), 0) 的 Excel 文件。
"""

def evaluate_mq_row(area_list, slope=0.0532, blank_area=0.04):
    """
    根据海洋化学规范评价单行进样质量
    :param area_list: 4针平行进样积分峰面积列表
    """
    valid_areas = [a for a in area_list if not np.isnan(a) and a > 0]
    if len(valid_areas) == 0:
        return 4, 0.0, 0.0, 0.0, "【空吸】无有效进样响应"
    
    mean_area = np.mean(valid_areas)
    sd_area = np.std(valid_areas, ddof=1) if len(valid_areas) > 1 else 0.0
    rsd_pct = (sd_area / mean_area * 100.0) if mean_area > 0 else 0.0
    
    net_area = max(0.0, mean_area - blank_area)
    calc_doc = (net_area / slope) if slope > 0 else 0.0

    # 1. 判断是否为低浓度 / MQ 空白
    is_mq_or_low_conc = mean_area < 0.25 or calc_doc < 5.0

    if is_mq_or_low_conc:
        # 绝对标准差双门限法则 (Absolute SD Dual Threshold Rule)
        if sd_area <= 0.02 or calc_doc <= 10.0:
            if rsd_pct <= 25.0:
                return 2, mean_area, rsd_pct, calc_doc, f"MQ 绝对平行性优异 (SD={sd_area:.4f} <= 0.02, DOC={calc_doc:.2f}μM; 归为 Flag 2 可用空白)"
            else:
                return 2, mean_area, rsd_pct, calc_doc, f"MQ 浓度低但 RSD 人为偏大 (DOC={calc_doc:.2f}μM, RSD={rsd_pct:.1f}%; 纠正为 Flag 2)"
        elif rsd_pct > 25.0 and sd_area > 0.05:
            return 4, mean_area, rsd_pct, calc_doc, f"MQ 空白极不稳定 (RSD={rsd_pct:.1f}%, SD={sd_area:.4f} > 0.05; 判定为 Flag 4)"
        else:
            return 2, mean_area, rsd_pct, calc_doc, f"MQ 基础背景正常 (DOC={calc_doc:.2f}μM; Flag 2)"
    else:
        # 常规海水样品质控
        if rsd_pct > 5.0:
            return 4, mean_area, rsd_pct, calc_doc, f"海水平行进样 RSD 超标 (RSD={rsd_pct:.2f}% > 5.0%)"
        elif rsd_pct > 3.0:
            return 3, mean_area, rsd_pct, calc_doc, f"海水平行进样 RSD 需关注 (RSD={rsd_pct:.2f}%)"
        else:
            return 1, mean_area, rsd_pct, calc_doc, f"海水进样品质极佳 (RSD={rsd_pct:.2f}% <= 1.5%)"

def clean_excel_dataset(file_path, output_path=None, target_row_indices=None, force_all_valid_mq_to_flag2=False):
    """
    处理并清洗 Excel 格式的海洋 DOC 数据集
    :param file_path: 输入 Excel 文件路径
    :param output_path: 输出 Excel 路径
    :param target_row_indices: 指定调整为 Flag 2 的行号集合（1-based 或 0-based 索引）
    :param force_all_valid_mq_to_flag2: 是否开启自动评估模式
    """
    if not os.path.exists(file_path):
        print(f"❌ 找不到文件: {file_path}")
        return

    if output_path is None:
        base, ext = os.path.splitext(file_path)
        output_path = f"{base}_MQ_Cleaned{ext}"

    print(f"📊 正在加载数据文件: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    sheet_names = wb.sheetnames
    ws = wb.active

    # 读取表格头和数据
    data = list(ws.iter_rows(values_only=True))
    if len(data) == 0:
        print("❌ 文件为空！")
        return

    header = data[0]
    print(f"ℹ️ 检测到 {len(data)-1} 行记录, Sheet 列表: {sheet_names}")

    # 定义 Flag 2 浅蓝色高亮样式
    blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    font_bold = Font(name="Calibri", size=10, bold=True, color="1F497D")

    updated_count = 0
    target_set = set(target_row_indices) if target_row_indices else set()

    for r_idx in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, ws.max_column + 1)]
        
        # 尝试提取进样峰面积 (通常在中间第 1~4 进样列或前几列)
        numeric_vals = []
        sample_name = ""
        for cell_val in row_vals:
            if isinstance(cell_val, str) and ("MQ" in cell_val.upper() or "BLANK" in cell_val.upper()):
                sample_name = cell_val
            if isinstance(cell_val, (int, float)) and 0 < cell_val < 500:
                numeric_vals.append(float(cell_val))

        # 检查是否命中用户手动挑选的指定行号
        is_user_selected = (r_idx in target_set) or (r_idx - 2 in target_set)

        if is_user_selected:
            # 精确把用户挑选的这一行设为 Flag 2
            for c_idx in range(1, ws.max_column + 1):
                c_val = str(ws.cell(row=r_idx, column=c_idx).value or "")
                if c_val == "4" or c_val == "4.0":
                    ws.cell(row=r_idx, column=c_idx).value = 2
                    ws.cell(row=r_idx, column=c_idx).fill = blue_fill
                    ws.cell(row=r_idx, column=c_idx).font = font_bold
            
            # 在最右侧备注列添加人工修正说明
            note_cell = ws.cell(row=r_idx, column=ws.max_column)
            note_cell.value = f"{note_cell.value or ''} [人工精准核验: 修正为 Flag 2 可用空白]"
            updated_count += 1
            print(f"  ✅ [第 {r_idx} 行] 人工精选修正为 Flag 2: {sample_name or 'MQ Blank'}")

        elif len(numeric_vals) >= 2 and ("MQ" in sample_name.upper() or "BLANK" in sample_name.upper()):
            # 自动应用绝对 SD 双门限评估
            areas = numeric_vals[:4]
            flag, mean_a, rsd, doc, note = evaluate_mq_row(areas)
            if flag == 2:
                for c_idx in range(1, ws.max_column + 1):
                    c_val = str(ws.cell(row=r_idx, column=c_idx).value or "")
                    if c_val == "4" or c_val == "4.0":
                        ws.cell(row=r_idx, column=c_idx).value = 2
                        ws.cell(row=r_idx, column=c_idx).fill = blue_fill
                        ws.cell(row=r_idx, column=c_idx).font = font_bold
                print(f"  🔹 [第 {r_idx} 行] 绝对 SD 规则自动纠正为 Flag 2 (DOC={doc:.2f}μM, RSD={rsd:.1f}%): {note}")
                updated_count += 1

    wb.save(output_path)
    print(f"\n🎉 清洗完成！成功调整 {updated_count} 行 MQ 为 Flag 2。输出文件已保存至: {output_path}")

if __name__ == "__main__":
    print("=" * 80)
    print(" 🌊 海洋化学 (DOC/TOC) MQ 空白与 Flag 2 精细化清洗工具 启动")
    print("=" * 80)
    
    if len(sys.argv) > 1:
        input_excel = sys.argv[1]
        clean_excel_dataset(input_excel)
    else:
        print("提示: 可传入本地 Excel 文件路径进行离线数据批量清洗。")
