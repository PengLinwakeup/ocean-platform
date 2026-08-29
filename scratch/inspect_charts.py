import os, sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

p = r'F:\印度洋测样\ODV\202608\20260827\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx'
if not os.path.exists(p):
    p = r'F:\印度洋测样\ODV\202608\20260826\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx'

wb = openpyxl.load_workbook(p, data_only=False)
ws = wb['Executive_Dashboard']

print('--- Executive_Dashboard Sections ---')
for r in range(1, 100):
    row_vals = []
    for c in range(1, 20):
        val = ws.cell(r, c).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_vals.append(f"{col_letter}{r}: {str(val)[:40]}")
    if row_vals:
        print(f"Row {r:2d}: " + " | ".join(row_vals[:5]))

print(f"\n--- Total charts: {len(ws._charts)} ---")
for i, c in enumerate(ws._charts):
    print(f"\n[Chart {i+1}]")
    print(f"Width: {c.width} cm, Height: {c.height} cm")
    print(f"Legend: {c.legend.legendPos if c.legend else 'None'}")
    if hasattr(c.anchor, '_from'):
        print(f"Anchor _from: col={c.anchor._from.col}, row={c.anchor._from.row}")
    for s_idx, s in enumerate(c.series):
        title_val = s.title.v if hasattr(s.title, 'v') else (s.title.value if hasattr(s.title, 'value') else str(s.title))
        x_ref = s.xVal.numRef.f if hasattr(s.xVal, 'numRef') and s.xVal.numRef else str(s.xVal)
        y_ref = s.yVal.numRef.f if hasattr(s.yVal, 'numRef') and s.yVal.numRef else str(s.yVal)
        print(f"  Series {s_idx+1}: title='{title_val}'")
        print(f"    xVal: {x_ref}")
        print(f"    yVal: {y_ref}")
        if s.marker:
            print(f"    marker: symbol={s.marker.symbol}, size={s.marker.size}")
        if hasattr(s, 'graphicalProperties') and s.graphicalProperties:
            print(f"    line: {getattr(s.graphicalProperties, 'line', None)}")

