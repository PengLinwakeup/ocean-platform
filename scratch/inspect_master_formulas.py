import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb_formula = openpyxl.load_workbook(file_path, data_only=False)

ws_master = wb_formula["All_Columns_Sequence_QC_Master"]

print("================ INSPECTING FORMULAS IN MASTER SHEET ================")

for r in range(6, 45):
    c1 = ws_master.cell(r, 1).value
    sname = ws_master.cell(r, 2).value
    inj1 = ws_master.cell(r, 6).value
    inj2 = ws_master.cell(r, 7).value
    inj3 = ws_master.cell(r, 8).value
    inj4 = ws_master.cell(r, 9).value
    
    clean_area_formula = ws_master.cell(r, 10).value
    rsd_formula = ws_master.cell(r, 11).value
    raw_doc_formula = ws_master.cell(r, 12).value
    mq_formula = ws_master.cell(r, 13).value
    qc_doc_formula = ws_master.cell(r, 14).value
    
    print(f"Row {r:3d} ({sname}) | CleanArea: {clean_area_formula} | RawDOC: {raw_doc_formula} | QCDoc: {qc_doc_formula}")
