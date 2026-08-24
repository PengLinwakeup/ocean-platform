import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import re
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
backup_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest_backup5.xlsx"

# Backup
shutil.copyfile(file_path, backup_path)
print(f"Backup created: {backup_path}")

# Load metadata JSON for slopes
with open("temp_geomar_v2_input.json", "r", encoding="utf-8") as f:
    input_data = json.load(f)

batches_meta = input_data.get("batches", [])
seq_info = {}
for idx, b in enumerate(batches_meta):
    seq_num = idx + 1
    slope = float(b.get("slope", 0.0554))
    rsq = float(b.get("rsq") or b.get("r2", 0.999))
    seq_info[seq_num] = {"seq_num": seq_num, "slope": slope, "rsq": rsq}

# Step 1: Read evaluated data with data_only=True
wb_data = openpyxl.load_workbook(file_path, data_only=True)

# Build lookup map from ODV_All_Samples_Full_List
ws_all_data = wb_data["ODV_All_Samples_Full_List"]
full_list_map = {}

for r in range(5, ws_all_data.max_row + 1):
    status = ws_all_data.cell(r, 1).value
    seq = ws_all_data.cell(r, 2).value
    st = str(ws_all_data.cell(r, 3).value or "").strip()
    sid = str(ws_all_data.cell(r, 4).value or "").strip()
    stype = str(ws_all_data.cell(r, 5).value or "").strip().upper()
    depth_val = ws_all_data.cell(r, 6).value
    raw_doc = ws_all_data.cell(r, 7).value
    mq_area = ws_all_data.cell(r, 8).value
    clean_area = ws_all_data.cell(r, 9).value
    clean_rsd = ws_all_data.cell(r, 10).value
    qc_doc = ws_all_data.cell(r, 11).value
    flag = ws_all_data.cell(r, 12).value
    comment = ws_all_data.cell(r, 13).value
    
    depth_num = float(depth_val) if depth_val is not None and str(depth_val).replace('.','').isdigit() else 0.0
    area_num = float(clean_area) if clean_area is not None and isinstance(clean_area, (int, float)) else 0.0
    doc_num = float(qc_doc) if qc_doc is not None and isinstance(qc_doc, (int, float)) else 0.0
    rsd_num = float(clean_rsd) if clean_rsd is not None and isinstance(clean_rsd, (int, float)) else 0.0
    
    if sid:
        full_list_map[sid] = {
            "status": status, "seq": seq, "st": st, "sid": sid, "stype": stype,
            "depth": depth_num, "raw_doc": raw_doc, "mq_area": mq_area, "clean_area": area_num,
            "clean_rsd": rsd_num, "qc_doc": doc_num, "flag": flag, "comment": comment
        }

print(f"Loaded {len(full_list_map)} seawater sample evaluations from ODV_All_Samples_Full_List.")

# Evaluate All_Columns_Sequence_QC_Master
ws_master_data = wb_data["All_Columns_Sequence_QC_Master"]
curr_seq = 1
curr_slope = seq_info[1]["slope"]
master_eval = {}

flag_counts = {1: 0, 2: 0, 3: 0, 4: 0}

for r in range(6, ws_master_data.max_row + 1):
    val1 = str(ws_master_data.cell(r, 1).value or "").strip()
    seq_match = re.search(r'【序列\s*(\d+)/26】', val1)
    if seq_match:
        curr_seq = int(seq_match.group(1))
        if curr_seq in seq_info:
            curr_slope = seq_info[curr_seq]["slope"]
        continue
        
    c1_val = ws_master_data.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master_data.cell(r, 2).value or "").strip()
        cat_type = str(ws_master_data.cell(r, 3).value or "").strip().upper()
        station = str(ws_master_data.cell(r, 4).value or "").strip()
        depth_val = ws_master_data.cell(r, 5).value
        
        area = ws_master_data.cell(r, 10).value
        rsd_val = ws_master_data.cell(r, 11).value
        raw_doc = ws_master_data.cell(r, 12).value
        qc_doc = ws_master_data.cell(r, 14).value
        
        area_num = float(area) if area is not None and isinstance(area, (int, float)) else 0.0
        rsd_num = float(rsd_val) if rsd_val is not None and isinstance(rsd_val, (int, float)) else 0.0
        doc_num = float(qc_doc) if qc_doc is not None and isinstance(qc_doc, (int, float)) else (float(raw_doc) if raw_doc is not None and isinstance(raw_doc, (int, float)) else 0.0)
        
        flag = 2
        comment = "Acceptable: Good quality"
        
        # 1. STD
        if "STD" in cat_type or "STD" in s_name.upper():
            flag = 2
            comment = "Acceptable: Calibration standard injection"
            
        # 2. MQ Blank
        elif "MQ" in cat_type or "MQ" in s_name.upper() or "BLANK" in cat_type or "BLANK" in s_name.upper():
            if area_num <= 0.15 or doc_num <= 0.9:
                flag = 2
                comment = "Acceptable: Low absolute area MQ blank (Area <= 0.15 / DOC <= 0.9 uM)"
            else:
                if rsd_num > 5.0:
                    flag = 4
                    comment = f"Rejected: Contaminated MQ blank (Area {area_num:.4f} > 0.15, DOC {doc_num:.2f} uM)"
                else:
                    flag = 2
                    comment = "Acceptable: Normal MQ blank"
                    
        # 3. DSW CRM
        elif "DSW" in cat_type or "DSW" in s_name.upper() or "CRM" in cat_type or "CRM" in s_name.upper():
            if doc_num >= 39.0:
                rec = round((doc_num / 39.45) * 100, 1)
                flag = 2
                comment = f"Acceptable: CRM DSW recovery in standard range ({doc_num:.1f} uM, {rec:.1f}%)"
            else:
                rec = round((doc_num / 39.45) * 100, 1)
                flag = 3
                comment = f"Questionable: Low CRM DSW recovery ({doc_num:.1f} uM < 39.0 uM, {rec:.1f}%)"
                
        # 4. Field Seawater Sample
        else:
            if s_name in full_list_map:
                item = full_list_map[s_name]
                flag = item["flag"]
                comment = item["comment"]
            else:
                if rsd_num > 5.0 and doc_num > 10.0:
                    flag = 4
                    comment = f"Rejected: High injection RSD ({rsd_num:.1f}% > 5.0%)"
                elif rsd_num > 3.0:
                    flag = 3
                    comment = f"Questionable: Moderate injection RSD ({rsd_num:.1f}% > 3.0%)"
                else:
                    flag = 2
                    comment = f"Acceptable: Low injection RSD ({rsd_num:.1f} <= 3.0%)"

        flag_counts[flag] += 1
        master_eval[r] = {"flag": flag, "comment": comment}

print("Master Sheet Restored Flag Distribution:")
print(flag_counts)

# Step 2: Write back to openpyxl workbook (data_only=False)
wb_out = openpyxl.load_workbook(file_path, data_only=False)

fill_green = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
fill_yellow = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid')
fill_red = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

font_green = Font(name='Times New Roman', size=9.5, color='166534', bold=True)
font_yellow = Font(name='Times New Roman', size=9.5, color='9A3412', bold=True)
font_red = Font(name='Times New Roman', size=9.5, color='991B1B', bold=True)

font_status_green = Font(name='楷体', size=9.5, color='166534', bold=True)
font_status_yellow = Font(name='楷体', size=9.5, color='9A3412', bold=True)
font_status_red = Font(name='楷体', size=9.5, color='991B1B', bold=True)

# 1. Write back to All_Columns_Sequence_QC_Master
ws_master_out = wb_out["All_Columns_Sequence_QC_Master"]
print("Writing restored flags & comments to All_Columns_Sequence_QC_Master...")

for r, eval_data in master_eval.items():
    flag = eval_data["flag"]
    comment = eval_data["comment"]
    
    if ws_master_out.max_column >= 15:
        ws_master_out.cell(r, 15).value = flag
    if ws_master_out.max_column >= 16:
        ws_master_out.cell(r, 16).value = comment
        
    c1 = ws_master_out.cell(r, 1)
    co = ws_master_out.cell(r, 15)
    
    if flag in [1, 2]:
        c1.fill = fill_green
        c1.font = font_green if isinstance(c1.value, (int, float)) else font_status_green
        co.fill = fill_green
        co.font = font_green
    elif flag == 3:
        c1.fill = fill_yellow
        c1.font = font_yellow if isinstance(c1.value, (int, float)) else font_status_yellow
        co.fill = fill_yellow
        co.font = font_yellow
    elif flag == 4:
        c1.fill = fill_red
        c1.font = font_red if isinstance(c1.value, (int, float)) else font_status_red
        co.fill = fill_red
        co.font = font_red

# 2. Write back to ODV_All_Samples_Full_List
ws_all_out = wb_out["ODV_All_Samples_Full_List"]
print("Writing restored flags & fills to ODV_All_Samples_Full_List...")

for r in range(5, ws_all_out.max_row + 1):
    sid = str(ws_all_out.cell(r, 4).value or "").strip()
    if sid in full_list_map:
        d = full_list_map[sid]
        flag = d["flag"]
        comment = d["comment"]
        status = d["status"]
        
        ws_all_out.cell(r, 1).value = status
        ws_all_out.cell(r, 12).value = flag
        ws_all_out.cell(r, 13).value = comment
        
        c1 = ws_all_out.cell(r, 1)
        cl = ws_all_out.cell(r, 12)
        
        if flag in [1, 2]:
            c1.fill = fill_green; c1.font = font_status_green
            cl.fill = fill_green; cl.font = font_green
        elif flag == 3:
            c1.fill = fill_yellow; c1.font = font_status_yellow
            cl.fill = fill_yellow; cl.font = font_yellow
        elif flag == 4:
            c1.fill = fill_red; c1.font = font_status_red
            cl.fill = fill_red; cl.font = font_red

# 3. Rebuild ODV_Clean_Export_Only
if "ODV_Clean_Export_Only" in wb_out.sheetnames:
    ws_clean = wb_out["ODV_Clean_Export_Only"]
    print("Rebuilding ODV_Clean_Export_Only sheet...")
    
    clean_list = [d for sid, d in full_list_map.items() if d["status"] == "保留 (Included)" and d["flag"] in [1, 2, 3]]
    
    for r in range(ws_clean.max_row, 4, -1):
        ws_clean.delete_rows(r)
        
    font_main = Font(name='Times New Roman', size=9.5, color='1E293B')
    font_chinese = Font(name='楷体', size=9.5, color='1E293B')
    fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    border_thin = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    for idx, d in enumerate(clean_list):
        r = 5 + idx
        ws_clean.row_dimensions[r].height = 20
        row_bg = fill_zebra if idx % 2 == 1 else fill_white
        
        ws_clean.cell(r, 1, d["seq"])
        ws_clean.cell(r, 2, d["st"])
        ws_clean.cell(r, 3, d["sid"])
        ws_clean.cell(r, 4, d["stype"])
        ws_clean.cell(r, 5, d["depth"])
        ws_clean.cell(r, 6, d["raw_doc"])
        ws_clean.cell(r, 7, d["mq_area"])
        ws_clean.cell(r, 8, d["clean_area"])
        ws_clean.cell(r, 9, d["clean_rsd"])
        ws_clean.cell(r, 10, d["qc_doc"])
        ws_clean.cell(r, 11, d["flag"])
        ws_clean.cell(r, 12, d["comment"])
        
        for c in range(1, 13):
            cell = ws_clean.cell(r, c)
            cell.fill = row_bg
            cell.border = border_thin
            if c in [1, 2, 3, 4, 12]: cell.font = font_chinese
            else: cell.font = font_main

# 4. Rebuild Flag4_Discarded_Audit_List
if "Flag4_Discarded_Audit_List" in wb_out.sheetnames:
    ws_flag4 = wb_out["Flag4_Discarded_Audit_List"]
    print("Rebuilding Flag4_Discarded_Audit_List sheet...")
    
    flag4_list = [d for sid, d in full_list_map.items() if d["status"] == "被丢弃 (Discarded)" or d["flag"] == 4]
    
    for r in range(ws_flag4.max_row, 5, -1):
        ws_flag4.delete_rows(r)
        
    for idx, d in enumerate(flag4_list):
        r = 6 + idx
        ws_flag4.row_dimensions[r].height = 20
        row_bg = fill_zebra if idx % 2 == 1 else fill_white
        
        ws_flag4.cell(r, 1, idx + 1)
        ws_flag4.cell(r, 2, d["seq"])
        ws_flag4.cell(r, 3, d["st"])
        ws_flag4.cell(r, 4, d["depth"])
        ws_flag4.cell(r, 5, d["sid"])
        ws_flag4.cell(r, 10, d["clean_area"])
        ws_flag4.cell(r, 11, d["clean_rsd"])
        ws_flag4.cell(r, 12, d["qc_doc"])
        ws_flag4.cell(r, 13, 4)
        ws_flag4.cell(r, 14, d["comment"])
        
        for c in range(1, 15):
            cell = ws_flag4.cell(r, c)
            cell.fill = row_bg
            cell.border = border_thin
            if c in [2, 3, 5, 14]: cell.font = font_chinese
            else: cell.font = font_main
            
        ws_flag4.cell(r, 1).fill = fill_red; ws_flag4.cell(r, 1).font = font_red
        ws_flag4.cell(r, 13).fill = fill_red; ws_flag4.cell(r, 13).font = font_red

# 5. Update Executive_Dashboard KPI summary counts
if "Executive_Dashboard" in wb_out.sheetnames:
    ws_dash = wb_out["Executive_Dashboard"]
    print("Updating Executive_Dashboard KPI summary counts...")
    
    clean_list = [d for sid, d in full_list_map.items() if d["status"] == "保留 (Included)" and d["flag"] in [1, 2, 3]]
    flag4_list = [d for sid, d in full_list_map.items() if d["status"] == "被丢弃 (Discarded)" or d["flag"] == 4]
    
    flag2_3_count = len(clean_list)
    flag4_count = len(flag4_list)
    total_eval = flag2_3_count + flag4_count
    
    pass_pct = round((flag2_3_count / total_eval) * 100, 1) if total_eval > 0 else 100.0
    discard_pct = round((flag4_count / total_eval) * 100, 1) if total_eval > 0 else 0.0
    
    ws_dash.cell(5, 5).value = f"{flag2_3_count} ({pass_pct}%)"
    ws_dash.cell(5, 7).value = f"{flag4_count} ({discard_pct}%)"

print("Saving modified workbook back to file...")
wb_out.save(file_path)
print("Workbook saved successfully!")
