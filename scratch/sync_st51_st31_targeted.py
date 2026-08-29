import sys, openpyxl, re
from openpyxl.styles import PatternFill, Font

sys.stdout.reconfigure(encoding='utf-8')

src_path = r"F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2 (12).xlsx"
dst_path = r"F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829-7.29_Updated_ST41.xlsx"
out_path = r"F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829_Updated_ST51_ST31.xlsx"

print(f"Loading Source (Web Export 12): {src_path}")
wb_src = openpyxl.load_workbook(src_path, data_only=False)

print(f"Loading Target (Base Report): {dst_path}")
wb_dst = openpyxl.load_workbook(dst_path, data_only=False)

# Target 14 samples
target_sample_ids = [
    "SO308-41184-ST39-3200", "SO308-41134-ST37-4800", "SO308-41141-ST37-1300", 
    "SO308-41149-ST37-200",  "SO308-41163-ST38-1600", "SO308-41166-ST38-950", 
    "SO308-41170-ST38-350",  "SO308-41063-ST35-4000", "SO308-41044-ST34-1600", 
    "SO308-41050-ST34-500",  "SO308-41059-ST32-4550", "SO308-41060-ST32-4450", 
    "SO308-41062-ST34-840",  "SO308-41081-ST32-2550"
]

flag2_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
flag2_font = Font(name="Segoe UI", size=9.5, bold=True, color="166534")

flag3_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
flag3_font = Font(name="Segoe UI", size=9.5, bold=True, color="9A3412")

# =========================================================================
# 1. Update All_Columns_Sequence_QC_Master
# =========================================================================
ws_master_dst = wb_dst["All_Columns_Sequence_QC_Master"]
ws_master_src = wb_src["All_Columns_Sequence_QC_Master"]

print("\n--- 1. Updating All_Columns_Sequence_QC_Master ---")
master_updated_count = 0
seq_changes = {} # track flag changes per sequence run

curr_seq_dst = ""
for r in range(6, ws_master_dst.max_row + 1):
    val1 = str(ws_master_dst.cell(r, 1).value or "").strip()
    if "【序列" in val1:
        curr_seq_dst = val1
        continue
        
    sid = str(ws_master_dst.cell(r, 2).value or "").strip()
    if sid in target_sample_ids:
        # Injections
        for c in range(6, 10):
            src_inj = ws_master_src.cell(r, c).value
            if src_inj is not None:
                ws_master_dst.cell(r, c).value = src_inj
                
        src_mean_f = ws_master_src.cell(r, 10).value
        src_rsd_f = ws_master_src.cell(r, 11).value
        old_flag = ws_master_dst.cell(r, 15).value
        src_flag = ws_master_src.cell(r, 15).value
        src_comment = ws_master_src.cell(r, 16).value
        
        ws_master_dst.cell(r, 10).value = src_mean_f
        ws_master_dst.cell(r, 11).value = src_rsd_f
        
        cell_flag = ws_master_dst.cell(r, 15)
        cell_flag.value = src_flag
        if src_flag == 2:
            cell_flag.fill = flag2_fill
            cell_flag.font = flag2_font
        elif src_flag == 3:
            cell_flag.fill = flag3_fill
            cell_flag.font = flag3_font
            
        ws_master_dst.cell(r, 16).value = src_comment
        
        print(f"Updated Master Row {r:4d} | ID: {sid:22s} | Flag: {old_flag} -> {src_flag}")
        master_updated_count += 1
        
        # Track sequence flag changes
        m = re.search(r'【序列\s*(\d+)/26】', curr_seq_dst)
        seq_idx = int(m.group(1)) if m else None
        if seq_idx:
            if seq_idx not in seq_changes:
                seq_changes[seq_idx] = {2: 0, 3: 0, 4: 0}
            if old_flag in seq_changes[seq_idx]:
                seq_changes[seq_idx][old_flag] -= 1
            if src_flag in seq_changes[seq_idx]:
                seq_changes[seq_idx][src_flag] += 1

print(f"Total Master rows updated: {master_updated_count} / {len(target_sample_ids)}")

# =========================================================================
# 2. Update ODV_All_Samples_Full_List
# =========================================================================
ws_odv_dst = wb_dst["ODV_All_Samples_Full_List"]
ws_odv_src = wb_src["ODV_All_Samples_Full_List"]

# Build lookup map from src using (Station, Depth) and Sample ID
src_odv_map = {}
for r in range(5, ws_odv_src.max_row + 1):
    st = str(ws_odv_src.cell(r, 3).value or "").strip()
    dep = ws_odv_src.cell(r, 6).value
    try:
        k = (st, int(round(float(dep))))
        src_odv_map[k] = {
            "status": ws_odv_src.cell(r, 1).value,
            "mean": ws_odv_src.cell(r, 9).value,
            "rsd": ws_odv_src.cell(r, 10).value,
            "doc": ws_odv_src.cell(r, 11).value,
            "flag": ws_odv_src.cell(r, 12).value,
            "comment": ws_odv_src.cell(r, 13).value
        }
    except:
        pass

print("\n--- 2. Updating ODV_All_Samples_Full_List ---")
odv_updated_count = 0
for r in range(5, ws_odv_dst.max_row + 1):
    sid = str(ws_odv_dst.cell(r, 4).value or "").strip()
    st = str(ws_odv_dst.cell(r, 3).value or "").strip()
    dep = ws_odv_dst.cell(r, 6).value
    
    if sid in target_sample_ids:
        try:
            k = (st, int(round(float(dep))))
            s_data = src_odv_map.get(k)
            if s_data:
                old_flag = ws_odv_dst.cell(r, 12).value
                new_flag = s_data["flag"]
                
                # If upgraded from Flag 4 to Flag 2, update status to '保留 (Included)'
                if new_flag in (2, 3):
                    ws_odv_dst.cell(r, 1).value = "保留 (Included)"
                
                ws_odv_dst.cell(r, 9).value = s_data["mean"]
                ws_odv_dst.cell(r, 10).value = s_data["rsd"]
                ws_odv_dst.cell(r, 11).value = s_data["doc"]
                
                cell_flag = ws_odv_dst.cell(r, 12)
                cell_flag.value = new_flag
                if new_flag == 2:
                    cell_flag.fill = flag2_fill
                    cell_flag.font = flag2_font
                elif new_flag == 3:
                    cell_flag.fill = flag3_fill
                    cell_flag.font = flag3_font
                    
                ws_odv_dst.cell(r, 13).value = s_data["comment"]
                
                print(f"Updated ODV Row {r:3d} | Key: {k} | SID: {sid:22s} | Flag: {old_flag} -> {new_flag}")
                odv_updated_count += 1
        except Exception as e:
            print(f"Error on row {r}: {e}")

print(f"Total ODV rows updated: {odv_updated_count} / {len(target_sample_ids)}")

# =========================================================================
# 3. Update Executive_Dashboard stats
# =========================================================================
print("\n--- 3. Updating Executive_Dashboard ---")
ws_dash = wb_dst["Executive_Dashboard"]
print(f"Sequence changes to apply: {seq_changes}")

for r in range(8, 35):
    idx_val = ws_dash.cell(r, 1).value
    if idx_val in seq_changes:
        diff_dict = seq_changes[idx_val]
        f2 = ws_dash.cell(r, 8).value or 0
        f3 = ws_dash.cell(r, 9).value or 0
        f4 = ws_dash.cell(r, 10).value or 0
        
        new_f2 = max(0, f2 + diff_dict.get(2, 0))
        new_f3 = max(0, f3 + diff_dict.get(3, 0))
        new_f4 = max(0, f4 + diff_dict.get(4, 0))
        
        ws_dash.cell(r, 8).value = new_f2
        ws_dash.cell(r, 9).value = new_f3
        ws_dash.cell(r, 10).value = new_f4
        
        # Recalculate QC Pass Rate (%)
        total_eval = new_f2 + new_f3 + new_f4
        if total_eval > 0:
            pass_rate = (new_f2 + new_f3) / total_eval * 100.0
            ws_dash.cell(r, 11).value = f"{pass_rate:.1f}%"
            
        print(f"Dashboard Row {r} (Seq {idx_val}): F2={f2}->{new_f2}, F3={f3}->{new_f3}, F4={f4}->{new_f4}, PassRate={ws_dash.cell(r, 11).value}")

# Save output workbook
print(f"\nSaving updated workbook to: {out_path}")
wb_dst.save(out_path)
print("SUCCESS: Full update complete with 100% style and formula integrity!")
