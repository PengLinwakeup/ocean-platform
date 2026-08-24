import openpyxl
import sys

# Ensure stdout uses utf-8
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"

print("--- Loading without data_only (Formulas) ---")
wb_formula = openpyxl.load_workbook(file_path, data_only=False)
print("Sheet names in workbook:")
for name in wb_formula.sheetnames:
    print(f" - {name}")

print("\n--- Loading with data_only (Evaluated Values) ---")
wb_val = openpyxl.load_workbook(file_path, data_only=True)

for name in wb_val.sheetnames:
    ws = wb_val[name]
    print(f"\n================ SHEET: {name} (Dimensions: {ws.max_row} rows x {ws.max_column} cols) ================")
    for r in range(1, min(40, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column + 1))]
        if any(v is not None for v in row_vals):
            # truncate long strings for readability
            short_vals = [str(v)[:30] if v is not None else "" for v in row_vals]
            print(f"Row {r:2d}: {short_vals}")
