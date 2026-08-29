import sys
import openpyxl
from openpyxl.styles import PatternFill, Font

sys.stdout.reconfigure(encoding='utf-8')

src_path = r"F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2 (11).xlsx"
dst_path = r"F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829-7.29.xlsx"
out_path = r"F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829-7.29_Updated_ST41.xlsx"

print(f"Loading Source (Web Export): {src_path}")
wb_src = openpyxl.load_workbook(src_path, data_only=False)

print(f"Loading Target (Base Report): {dst_path}")
wb_dst = openpyxl.load_workbook(dst_path, data_only=False)

# Target sample to update: ST-41, depth 490, SO308-41255-ST41-490
target_sample_id = "SO308-41255-ST41-490"
target_st = "ST-41"
target_depth = 490

# Standard Green Fill and Font for Flag 2 (WOCE Good)
flag2_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
flag2_font = Font(name="Segoe UI", size=9.5, bold=True, color="166534")

# =========================================================================
# 1. Update Sheet: All_Columns_Sequence_QC_Master
# =========================================================================
ws_master_dst = wb_dst["All_Columns_Sequence_QC_Master"]
ws_master_src = wb_src["All_Columns_Sequence_QC_Master"]

print("\n--- 1. Updating All_Columns_Sequence_QC_Master ---")
master_updated = False
for r in range(6, ws_master_dst.max_row + 1):
    sid = str(ws_master_dst.cell(r, 2).value or "").strip()
    st = str(ws_master_dst.cell(r, 4).value or "").strip()
    dep = ws_master_dst.cell(r, 5).value
    
    if sid == target_sample_id or (target_st in st and dep == target_depth):
        # Extract new values from src
        src_mean_formula = ws_master_src.cell(r, 10).value
        src_rsd_formula = ws_master_src.cell(r, 11).value
        src_flag = ws_master_src.cell(r, 15).value
        src_comment = ws_master_src.cell(r, 16).value
        
        # Log before
        print(f"Matched Master Row {r}: {sid}")
        print(f"  [BEFORE] Mean: {ws_master_dst.cell(r, 10).value} | RSD: {ws_master_dst.cell(r, 11).value} | Flag: {ws_master_dst.cell(r, 15).value}")
        print(f"           Comment: {ws_master_dst.cell(r, 16).value}")
        
        # Update cells in-place
        ws_master_dst.cell(r, 10).value = src_mean_formula
        ws_master_dst.cell(r, 11).value = src_rsd_formula
        
        cell_flag = ws_master_dst.cell(r, 15)
        cell_flag.value = src_flag
        cell_flag.fill = flag2_fill
        cell_flag.font = flag2_font
        
        ws_master_dst.cell(r, 16).value = src_comment
        
        # Log after
        print(f"  [AFTER]  Mean: {ws_master_dst.cell(r, 10).value} | RSD: {ws_master_dst.cell(r, 11).value} | Flag: {ws_master_dst.cell(r, 15).value}")
        print(f"           Comment: {ws_master_dst.cell(r, 16).value}")
        master_updated = True
        break

if not master_updated:
    print(f"WARNING: Target sample {target_sample_id} not found in All_Columns_Sequence_QC_Master!")

# =========================================================================
# 2. Update Sheet: ODV_All_Samples_Full_List
# =========================================================================
ws_odv_dst = wb_dst["ODV_All_Samples_Full_List"]
ws_odv_src = wb_src["ODV_All_Samples_Full_List"]

print("\n--- 2. Updating ODV_All_Samples_Full_List ---")
odv_updated = False
for r in range(5, ws_odv_dst.max_row + 1):
    sid = str(ws_odv_dst.cell(r, 4).value or "").strip()
    st = str(ws_odv_dst.cell(r, 3).value or "").strip()
    dep = ws_odv_dst.cell(r, 6).value
    
    if sid == target_sample_id or (target_st in st and dep == target_depth):
        src_mean = ws_odv_src.cell(r, 9).value
        src_rsd = ws_odv_src.cell(r, 10).value
        src_qc_doc = ws_odv_src.cell(r, 11).value
        src_flag = ws_odv_src.cell(r, 12).value
        src_comment = ws_odv_src.cell(r, 13).value
        
        print(f"Matched ODV Row {r}: {sid}")
        print(f"  [BEFORE] Mean Area: {ws_odv_dst.cell(r, 9).value} | RSD: {ws_odv_dst.cell(r, 10).value}% | DOC: {ws_odv_dst.cell(r, 11).value} | Flag: {ws_odv_dst.cell(r, 12).value}")
        print(f"           Comment: {ws_odv_dst.cell(r, 13).value}")
        
        ws_odv_dst.cell(r, 9).value = src_mean
        ws_odv_dst.cell(r, 10).value = src_rsd
        ws_odv_dst.cell(r, 11).value = src_qc_doc
        
        cell_flag = ws_odv_dst.cell(r, 12)
        cell_flag.value = src_flag
        cell_flag.fill = flag2_fill
        cell_flag.font = flag2_font
        
        ws_odv_dst.cell(r, 13).value = src_comment
        
        print(f"  [AFTER]  Mean Area: {ws_odv_dst.cell(r, 9).value} | RSD: {ws_odv_dst.cell(r, 10).value}% | DOC: {ws_odv_dst.cell(r, 11).value} | Flag: {ws_odv_dst.cell(r, 12).value}")
        print(f"           Comment: {ws_odv_dst.cell(r, 13).value}")
        odv_updated = True
        break

if not odv_updated:
    print(f"WARNING: Target sample {target_sample_id} not found in ODV_All_Samples_Full_List!")

# =========================================================================
# 3. Update Sheet: Executive_Dashboard (Sequence 5 stats: Flag 2 +1, Flag 3 -1)
# =========================================================================
print("\n--- 3. Updating Executive_Dashboard ---")
ws_dash = wb_dst["Executive_Dashboard"]
dash_updated = False
for r in range(8, 35):
    seq_name = str(ws_dash.cell(r, 2).value or "")
    if "序列 5" in seq_name or "ST-41" in seq_name:
        f2 = ws_dash.cell(r, 8).value
        f3 = ws_dash.cell(r, 9).value
        print(f"Matched Dashboard Row {r}: {seq_name}")
        print(f"  [BEFORE] Flag 2: {f2} | Flag 3: {f3}")
        
        if isinstance(f2, (int, float)):
            ws_dash.cell(r, 8).value = f2 + 1
        if isinstance(f3, (int, float)) and f3 > 0:
            ws_dash.cell(r, 9).value = f3 - 1
            
        print(f"  [AFTER]  Flag 2: {ws_dash.cell(r, 8).value} | Flag 3: {ws_dash.cell(r, 9).value}")
        dash_updated = True
        break

# Save to output file
print(f"\nSaving updated workbook to: {out_path}")
wb_dst.save(out_path)
print("SUCCESS: Target file updated in-place with zero style loss!")
