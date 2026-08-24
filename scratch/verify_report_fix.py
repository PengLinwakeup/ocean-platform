import sys
import os
sys.path.append(os.path.abspath('.'))

import json
import openpyxl
import run_geomar_qc_processor_20260820 as proc

json_path = r"temp_geomar_v2_input.json"
out_path = r"F:\印度洋测样\ODV\202608\20260822\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2 .xlsx"

print("1. Parsing JSON batches with updated processor...")
with open(json_path, 'r', encoding='utf-8') as f:
    jdata = json.load(f)

batches = proc.parse_json_batches(jdata)

print(f"Total batches processed: {len(batches)}")
for b in batches[:6]:
    print(f"Batch {b.index} ({b.sheet_name}): {len(b.samples)} samples (slope={b.slope:.5f}, R²={b.rsq:.5f})")

print("\n2. Rebuilding GEOMAR Master Excel report...")
proc.build_geomar_master_excel(batches, out_path)
print("Excel saved successfully.")

print("\n3. Verifying Master Sheet in generated Excel...")
wb = openpyxl.load_workbook(out_path, data_only=True)
ws = wb["All_Columns_Sequence_QC_Master"]

seq_info = []
for r in range(1, ws.max_row + 1):
    val = ws.cell(r, 1).value
    if val and "【序列" in str(val):
        # count rows in this sequence
        r_curr = r + 3
        cnt = 0
        while r_curr <= ws.max_row and ws.cell(r_curr, 1).value is not None:
            cnt += 1
            r_curr += 1
        seq_info.append((r, str(val), cnt))

print(f"Total sequences in Master: {len(seq_info)}")
for r, title, cnt in seq_info[:10]:
    clean_title = str(title).encode('ascii', errors='ignore').decode('ascii')
    print(f"  Row {r:4d}: {clean_title} --> {cnt} samples")

# Check for duplicate sample names in sequence 1 vs sequence 2 vs sequence 3
def get_seq_samples(start_r):
    samples = []
    dr = start_r + 3
    while dr <= ws.max_row and ws.cell(dr, 1).value is not None:
        samples.append(ws.cell(dr, 2).value)
        dr += 1
    return samples

seq1_s = get_seq_samples(seq_info[0][0])
seq2_s = get_seq_samples(seq_info[1][0])
seq3_s = get_seq_samples(seq_info[2][0])

print("\nSequence 1 sample count:", len(seq1_s), "first 3:", seq1_s[:3])
print("Sequence 2 sample count:", len(seq2_s), "first 3:", seq2_s[:3])
print("Sequence 3 sample count:", len(seq3_s), "first 3:", seq3_s[:3])
