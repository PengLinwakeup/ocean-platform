import openpyxl

file_path = r'F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829_Final_QC_Completed_Cleaned_20260829_2218.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb['All_Columns_Sequence_QC_Master']

out_lines = []
out_lines.append("=== DETAILED INSPECTION OF DSW PURPLE FONT CELLS ===")

dsw_purple_count = 0
non_dsw_purple_count = 0

for r in range(7, ws.max_row + 1):
    c2_val = str(ws.cell(r, 2).value or '').strip().upper()
    c3_val = str(ws.cell(r, 3).value or '').strip().upper()
    
    is_dsw = ('DSW' in c2_val or 'DSW' in c3_val)
    
    row_purple_cols = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c)
        if cell.font and cell.font.color:
            color_str = str(cell.font.color.rgb) if cell.font.color.rgb else ''
            if '6B21A8' in color_str.upper() or 'PURPLE' in color_str.upper() or '7E22CE' in color_str.upper() or '9333EA' in color_str.upper():
                row_purple_cols.append((c, cell.value, color_str))
                
    if row_purple_cols:
        if is_dsw:
            dsw_purple_count += len(row_purple_cols)
            cols_summary = ", ".join([f"Col {c} ({val!r})" for c, val, col_str in row_purple_cols])
            out_lines.append(f"DSW Row {r:4d} | Name: {ws.cell(r,2).value!r} | Category: {ws.cell(r,3).value!r} | Purple Cols: {cols_summary}")
        else:
            non_dsw_purple_count += len(row_purple_cols)
            cols_summary = ", ".join([f"Col {c} ({val!r})" for c, val, col_str in row_purple_cols])
            out_lines.append(f"NON-DSW Row {r:4d} | Name: {ws.cell(r,2).value!r} | Category: {ws.cell(r,3).value!r} | Purple Cols: {cols_summary}")

out_lines.append("\n=== SUMMARY ===")
out_lines.append(f"Total DSW purple font cells: {dsw_purple_count}")
out_lines.append(f"Total NON-DSW purple font cells: {non_dsw_purple_count}")

with open('scratch/dsw_purple_inspection_detail.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out_lines))

print("Saved inspection detail to scratch/dsw_purple_inspection_detail.txt")
