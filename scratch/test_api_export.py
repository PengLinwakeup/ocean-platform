import json
import urllib.request
import openpyxl

json_path = r"temp_geomar_v2_input.json"
out_test_path = r"scratch/api_exported_test.xlsx"

with open(json_path, 'r', encoding='utf-8') as f:
    jdata = json.load(f)

req = urllib.request.Request(
    'http://localhost:3000/api/export-geomar-v2',
    data=json.dumps(jdata).encode('utf-8'),
    headers={'Content-Type': 'application/json'}
)

try:
    with urllib.request.urlopen(req) as resp:
        content = resp.read()
        with open(out_test_path, 'wb') as out_f:
            out_f.write(content)
        print("Successfully exported Excel from /api/export-geomar-v2!")
        
        wb = openpyxl.load_workbook(out_test_path, data_only=True)
        ws = wb["All_Columns_Sequence_QC_Master"]
        
        seq_info = []
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and "【序列" in str(val):
                r_curr = r + 3
                cnt = 0
                while r_curr <= ws.max_row and ws.cell(r_curr, 1).value is not None:
                    cnt += 1
                    r_curr += 1
                seq_info.append((r, str(val), cnt))
        
        print(f"Total sequences in Master: {len(seq_info)}")
        for r, title, cnt in seq_info[:6]:
            clean_title = str(title).encode('ascii', errors='ignore').decode('ascii')
            print(f"  Row {r:4d}: {clean_title} --> {cnt} samples")

except Exception as e:
    print("Failed to reach endpoint:", e)
