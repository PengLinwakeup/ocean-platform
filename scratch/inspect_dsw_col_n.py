import openpyxl
import json
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb_data = openpyxl.load_workbook(file_path, data_only=True)

ws_master = wb_data["All_Columns_Sequence_QC_Master"]

# Load slopes from temp_geomar_v2_input.json
with open("temp_geomar_v2_input.json", "r", encoding="utf-8") as f:
    input_data = json.load(f)

batches_meta = input_data.get("batches", [])
seq_info = {}
for idx, b in enumerate(batches_meta):
    seq_num = idx + 1
    slope = float(b.get("slope", 0.0554))
    seq_info[seq_num] = slope

curr_seq = 1
curr_slope = seq_info[1]

dsw_rows_info = []

for r in range(6, ws_master.max_row + 1):
    val1 = str(ws_master.cell(r, 1).value or "").strip()
    seq_match = re.search(r'【序列\s*(\d+)/26】', val1)
    if seq_match:
        curr_seq = int(seq_match.group(1))
        if curr_seq in seq_info:
            curr_slope = seq_info[curr_seq]
        continue
        
    c1_val = ws_master.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master.cell(r, 2).value or "").strip()
        cat_type = str(ws_master.cell(r, 3).value or "").strip().upper()
        
        if "DSW" in cat_type or "DSW" in s_name.upper() or "CRM" in cat_type or "CRM" in s_name.upper():
            inj1 = ws_master.cell(r, 6).value
            inj2 = ws_master.cell(r, 7).value
            inj3 = ws_master.cell(r, 8).value
            inj4 = ws_master.cell(r, 9).value
            
            clean_area = ws_master.cell(r, 10).value
            raw_doc = ws_master.cell(r, 12).value
            mq_area = ws_master.cell(r, 13).value
            qc_doc = ws_master.cell(r, 14).value
            
            injs = [float(x) for x in [inj1, inj2, inj3, inj4] if x is not None and isinstance(x, (int, float)) and x > 0 and x < 10.0]
            
            c_area = float(clean_area) if clean_area is not None and isinstance(clean_area, (int, float)) else (sum(injs)/len(injs) if len(injs)>0 else 0)
            m_area = float(mq_area) if mq_area is not None and isinstance(mq_area, (int, float)) else 0.0
            
            # QC Dynamic DOC = max(0, (c_area - m_area) / curr_slope)
            net_doc = round(max(0, (c_area - m_area) / curr_slope), 2) if curr_slope > 0 else 0.0
            
            dsw_rows_info.append({
                "row": r, "seq": curr_seq, "slope": curr_slope, "name": s_name,
                "injections": injs, "clean_area": c_area, "mq_area": m_area,
                "net_doc": net_doc, "raw_doc": round(c_area / curr_slope, 2)
            })

print(f"Total DSW rows found: {len(dsw_rows_info)}")
below_39_rows = [d for d in dsw_rows_info if d["net_doc"] < 39.0]
print(f"DSW rows where QC Dynamic DOC (Col N) < 39.0: {len(below_39_rows)}")

print("\nSample DSW rows with QC Dynamic DOC < 39.0:")
for d in below_39_rows[:20]:
    print(f"Row {d['row']:4d} | Seq {d['seq']:2d} | Injections: {d['injections']} | CleanArea: {d['clean_area']:.4f} | MQ: {d['mq_area']:.4f} | Raw DOC: {d['raw_doc']:.2f} | QC Dynamic DOC: {d['net_doc']:.2f}")
