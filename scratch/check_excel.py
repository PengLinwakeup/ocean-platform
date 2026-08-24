import openpyxl

excel_path = r"F:\印度洋测样\ODV\202608\20260822\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2 .xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

ws = wb["All_Columns_Sequence_QC_Master"]

print("Master Sheet Max Row:", ws.max_row)

seq_info = []
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v and "【序列" in str(v):
        seq_info.append((r, str(v)))

print(f"Total sequences in Master: {len(seq_info)}")
for r, title in seq_info:
    print(f"Row {r:4d}: {title}")

# Let's inspect rows for Seq 1 (row 4) and Seq 2 (row 115) in detail
def get_seq_rows(start_r):
    rows = []
    # Data headers are at start_r + 2, data rows start at start_r + 3
    dr = start_r + 3
    while dr <= ws.max_row:
        val1 = ws.cell(dr, 1).value # Seq Order
        if val1 is None:
            break
        # sample_name, cat, station, depth, raw_doc, dynamic_blank, qc_doc
        row_data = [
            ws.cell(dr, 1).value, # order
            ws.cell(dr, 2).value, # sample_name
            ws.cell(dr, 3).value, # cat
            ws.cell(dr, 4).value, # station
            ws.cell(dr, 5).value, # depth
            ws.cell(dr, 12).value, # raw_doc
            ws.cell(dr, 14).value, # qc_doc
        ]
        rows.append(row_data)
        dr += 1
    return rows

seq1_data = get_seq_rows(4)
seq2_data = get_seq_rows(115)

print("\n--- Seq 1 Data (first 10 rows) ---")
for r in seq1_data[:10]:
    print(r)

print("\n--- Seq 2 Data (first 10 rows) ---")
for r in seq2_data[:10]:
    print(r)

# Compare seq1 and seq2 data
is_identical = (seq1_data == seq2_data)
print(f"\nAre Seq 1 and Seq 2 data identical? {is_identical}")
