import sys
import openpyxl

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"F:\印度洋测样\ODV\202608\20260822\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2 .xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

ws = wb["All_Columns_Sequence_QC_Master"]

seq_starts = []
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v and "【序列" in str(v):
        seq_starts.append(r)

print(f"Total sequences found: {len(seq_starts)}")

for i in range(min(5, len(seq_starts))):
    r_start = seq_starts[i]
    seq_title = ws.cell(r_start, 1).value
    seq_param = ws.cell(r_start + 1, 1).value
    
    # Read first 5 samples
    samples = []
    for r in range(r_start + 3, r_start + 8):
        s_name = ws.cell(r, 2).value
        raw_area1 = ws.cell(r, 6).value
        raw_area2 = ws.cell(r, 7).value
        clean_mean = ws.cell(r, 10).value
        samples.append((s_name, raw_area1, raw_area2, clean_mean))
        
    print(f"\n--- Sequence {i+1} ---")
    print("Title:", seq_title)
    print("Param:", seq_param)
    print("Samples (first 5):")
    for s in samples:
        print("  ", s)
