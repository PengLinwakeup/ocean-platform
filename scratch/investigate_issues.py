import json
import openpyxl

json_path = r"temp_geomar_v2_input.json"
excel_path = r"F:\印度洋测样\ODV\202608\20260822\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2 .xlsx"

print("--- Inspecting temp_geomar_v2_input.json ---")
with open(json_path, 'r', encoding='utf-8') as f:
    jdata = json.load(f)

batches_json = jdata.get("batches", [])
print(f"Total batches in JSON: {len(batches_json)}")

for i, b in enumerate(batches_json):
    curve_id = b.get("curveId")
    curve_name = b.get("curveName")
    file_name = b.get("fileName")
    col_idx = b.get("fileColIdx")
    slope = b.get("slope")
    rsq = b.get("rsq")
    samples = b.get("samples", [])
    print(f"Batch {i+1}: fileColIdx={col_idx}, curveId='{curve_id}', curveName='{curve_name}', fileName='{file_name}', slope={slope}, rsq={rsq}, samples_len={len(samples)}")

print("\n--- Inspecting Excel All_Columns_Sequence_QC_Master ---")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb["All_Columns_Sequence_QC_Master"]

seq_rows = []
for r in range(1, ws.max_row + 1):
    val = ws.cell(r, 1).value
    if val and "【序列" in str(val):
        sub = ws.cell(r+1, 1).value
        seq_rows.append((r, str(val), str(sub)))

for r, title, sub in seq_rows:
    print(f"Row {r:4d}: {title} | {sub}")
