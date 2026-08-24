import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"

# Load formula mode
wb_formula = openpyxl.load_workbook(file_path, data_only=False)
ws_master_formula = wb_formula["All_Columns_Sequence_QC_Master"]

# Load data_only mode
wb_data = openpyxl.load_workbook(file_path, data_only=True)
ws_master_data = wb_data["All_Columns_Sequence_QC_Master"]

print("================ DEEP VERIFYING EXCEL FORMULAS & CELL FILLS ================")

# Test Row 245 (from user's screenshot)
r = 245
print(f"\nInspecting Row {r} (User's Screenshot Row):")
for col in [1, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]:
    c_f = ws_master_formula.cell(r, col)
    c_d = ws_master_data.cell(r, col)
    fill = c_f.fill
    fill_type = fill.fill_type if fill else None
    fg = fill.fgColor.value if (fill and fill.fgColor) else None
    print(f"Col {col:2d} | Formula/Val: {str(c_f.value):35s} | Eval Value: {str(c_d.value):10s} | FillType: {fill_type} | FgColor: {fg}")

print("\nVerifying Col P (Comment) background fills across all rows:")
has_fill_comments = 0
for row_i in range(6, ws_master_formula.max_row + 1):
    cell_p = ws_master_formula.cell(row_i, 16)
    fill_p = cell_p.fill
    if fill_p and fill_p.fill_type is not None and fill_p.fill_type != 'none':
        has_fill_comments += 1

print(f"Total Col P (Comment) cells with background fill: {has_fill_comments} (Expected: 0)")
