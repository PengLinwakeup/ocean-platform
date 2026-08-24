import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
backup_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest_backup4.xlsx"

# Backup
shutil.copyfile(file_path, backup_path)
print(f"Backup created: {backup_path}")

wb = openpyxl.load_workbook(file_path)

fill_green = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
fill_yellow = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid')
fill_red = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

font_green = Font(name='Times New Roman', size=9.5, color='166534', bold=True)
font_yellow = Font(name='Times New Roman', size=9.5, color='9A3412', bold=True)
font_red = Font(name='Times New Roman', size=9.5, color='991B1B', bold=True)

font_status_green = Font(name='楷体', size=9.5, color='166534', bold=True)
font_status_yellow = Font(name='楷体', size=9.5, color='9A3412', bold=True)
font_status_red = Font(name='楷体', size=9.5, color='991B1B', bold=True)

# 1. Update All_Columns_Sequence_QC_Master
ws_master = wb["All_Columns_Sequence_QC_Master"]
print("Synchronizing fills in All_Columns_Sequence_QC_Master...")

master_updated_count = 0
for r in range(6, ws_master.max_row + 1):
    c1 = ws_master.cell(r, 1)
    co = ws_master.cell(r, 15)
    flag_val = co.value
    
    if flag_val is not None and isinstance(flag_val, (int, float)):
        flag_int = int(flag_val)
        if flag_int in [1, 2]:
            c1.fill = fill_green
            c1.font = font_green if isinstance(c1.value, (int, float)) else font_status_green
            co.fill = fill_green
            co.font = font_green
            master_updated_count += 1
        elif flag_int == 3:
            c1.fill = fill_yellow
            c1.font = font_yellow if isinstance(c1.value, (int, float)) else font_status_yellow
            co.fill = fill_yellow
            co.font = font_yellow
            master_updated_count += 1
        elif flag_int == 4:
            c1.fill = fill_red
            c1.font = font_red if isinstance(c1.value, (int, float)) else font_status_red
            co.fill = fill_red
            co.font = font_red
            master_updated_count += 1

print(f"Master Sheet: {master_updated_count} cells synchronized with Flag fill colors!")

# 2. Update ODV_All_Samples_Full_List
if "ODV_All_Samples_Full_List" in wb.sheetnames:
    ws_all = wb["ODV_All_Samples_Full_List"]
    print("Synchronizing fills in ODV_All_Samples_Full_List...")
    
    full_list_count = 0
    for r in range(5, ws_all.max_row + 1):
        c1 = ws_all.cell(r, 1)
        cl = ws_all.cell(r, 12)
        flag_val = cl.value
        
        if flag_val is not None and isinstance(flag_val, (int, float)):
            flag_int = int(flag_val)
            if flag_int in [1, 2]:
                c1.fill = fill_green
                c1.font = font_status_green
                cl.fill = fill_green
                cl.font = font_green
                full_list_count += 1
            elif flag_int == 3:
                c1.fill = fill_yellow
                c1.font = font_status_yellow
                cl.fill = fill_yellow
                cl.font = font_yellow
                full_list_count += 1
            elif flag_int == 4:
                c1.fill = fill_red
                c1.font = font_status_red
                cl.fill = fill_red
                cl.font = font_red
                full_list_count += 1

    print(f"Full List Sheet: {full_list_count} cells synchronized with Flag fill colors!")

# 3. Update Flag4_Discarded_Audit_List
if "Flag4_Discarded_Audit_List" in wb.sheetnames:
    ws_flag4 = wb["Flag4_Discarded_Audit_List"]
    print("Synchronizing fills in Flag4_Discarded_Audit_List...")
    
    for r in range(6, ws_flag4.max_row + 1):
        c1 = ws_flag4.cell(r, 1)
        c13 = ws_flag4.cell(r, 13)
        c1.fill = fill_red
        c1.font = font_red
        c13.fill = fill_red
        c13.font = font_red

print("Saving updated workbook back to file...")
wb.save(file_path)
print("Workbook saved successfully!")
