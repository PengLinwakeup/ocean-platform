import openpyxl
import os
import glob
import json

target_excel = r'F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829_Final_QC_Completed.xlsx'
sample_list_file = r'F:\印度洋测样\Indian Ocean_SO308_DOC_Sample List(1) 的副本.xlsx'

out_lines = []

def log(msg):
    out_lines.append(str(msg))

log("=== 1. TARGET EXCEL SHEET HEADERS AND COLUMNS ===")
wb_target = openpyxl.load_workbook(target_excel, data_only=False)
ws_master = wb_target['All_Columns_Sequence_QC_Master']

log(f"Master Sheet Max Row: {ws_master.max_row}, Max Col: {ws_master.max_column}")

for r in range(1, 7):
    row_vals = [ws_master.cell(r, c).value for c in range(1, ws_master.max_column + 1)]
    log(f"Row {r}: {row_vals}")

log("\n=== 2. SAMPLE LIST SHEET HEADERS AND COLUMNS ===")
if os.path.exists(sample_list_file):
    wb_sample = openpyxl.load_workbook(sample_list_file, data_only=False)
    log(f"Sample List Sheets: {wb_sample.sheetnames}")
    ws_sample = wb_sample.active
    for r in range(1, 6):
        row_vals = [ws_sample.cell(r, c).value for c in range(1, ws_sample.max_column + 1)]
        log(f"Sample Row {r}: {row_vals}")
else:
    log(f"Sample list file NOT found at {sample_list_file}")
    # find alternative
    cand = glob.glob(r'F:\印度洋测样\**\*Sample List*.xlsx', recursive=True)
    log(f"Candidates found: {cand}")

log("\n=== 3. MASTER SHEET CELL FORMATTING & FLAGS (Rows 6 to 50) ===")
# Inspect cell styles for qualified vs warning rows
for r in range(6, 60):
    row_vals = [ws_master.cell(r, c).value for c in range(1, ws_master.max_column + 1)]
    if not any(row_vals):
        continue
    c1 = ws_master.cell(r, 1)
    font_color = c1.font.color.rgb if (c1.font and c1.font.color) else (c1.font.color.theme if c1.font and c1.font.color else None)
    fill_fg = c1.fill.fgColor.rgb if (c1.fill and c1.fill.fgColor) else None
    fill_type = c1.fill.fill_type
    
    # inspect other columns in row r
    # sample ID, station, depth, peak areas, flag
    log(f"Row {r:3d} | Col1: val={c1.value!r}, font_color={font_color}, fill_type={fill_type}, fgColor={fill_fg} | Row snippet: {row_vals[:6]} ... {row_vals[15:]}")

with open('scratch/inspection_result.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out_lines))

print("Saved to scratch/inspection_result.txt")
