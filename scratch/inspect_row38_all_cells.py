import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path)

ws_master = wb["All_Columns_Sequence_QC_Master"]

print("================ MASTER SHEET ROW 38 & 39 CELL FILLS ================")
for r in [38, 39, 40]:
    row_fills = []
    for c in range(1, ws_master.max_column + 1):
        cell = ws_master.cell(r, c)
        fill_hex = cell.fill.fgColor.value if (cell.fill and cell.fill.fgColor) else "None"
        row_fills.append(f"Col{c}:{fill_hex}")
    print(f"Row {r:2d}: " + " | ".join(row_fills))

print("\n================ FULL LIST ROW 38 & 39 CELL FILLS ================")
ws_all = wb["ODV_All_Samples_Full_List"]
for r in [38, 39, 40]:
    row_fills = []
    for c in range(1, ws_all.max_column + 1):
        cell = ws_all.cell(r, c)
        fill_hex = cell.fill.fgColor.value if (cell.fill and cell.fill.fgColor) else "None"
        row_fills.append(f"Col{c}:{fill_hex}")
    print(f"Row {r:2d}: " + " | ".join(row_fills))
