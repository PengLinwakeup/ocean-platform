import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb_data = openpyxl.load_workbook(file_path, data_only=True)

ws_master = wb_data["All_Columns_Sequence_QC_Master"]

print("Master sheet rows with 41446:")
for r in range(6, ws_master.max_row + 1):
    c2 = str(ws_master.cell(r, 2).value or "")
    if "41446" in c2:
        print(f"Row {r:4d} | Col 2: '{c2}' | Col 3 (Cat): '{ws_master.cell(r,3).value}' | Col 4 (ST): '{ws_master.cell(r,4).value}' | Col 5 (Depth): {ws_master.cell(r,5).value} | Col 12 (Raw DOC): {ws_master.cell(r,12).value} | Col 14 (QC DOC): {ws_master.cell(r,14).value}")

ws_all = wb_data["ODV_All_Samples_Full_List"]
print("\nODV_All_Samples_Full_List rows with 41446:")
for r in range(5, ws_all.max_row + 1):
    c4 = str(ws_all.cell(r, 4).value or "")
    if "41446" in c4:
        print(f"Row {r:4d} | Col 1 (Status): '{ws_all.cell(r,1).value}' | Col 3 (ST): '{ws_all.cell(r,3).value}' | Col 4 (ID): '{c4}' | Col 5 (Type): '{ws_all.cell(r,5).value}' | Col 6 (Depth): {ws_all.cell(r,6).value} | Col 11 (QC DOC): {ws_all.cell(r,11).value} | Col 12 (Flag): {ws_all.cell(r,12).value}")
