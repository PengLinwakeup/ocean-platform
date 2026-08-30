import openpyxl
import os
import re

target_excel = r'F:\印度洋测样\ODV\202608\20260829\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829_Final_QC_Completed.xlsx'
sample_list_file = r'F:\印度洋测样\Indian Ocean_SO308_DOC_Sample List(1) 的副本.xlsx'

wb_target = openpyxl.load_workbook(target_excel, data_only=True)
ws_master = wb_target['All_Columns_Sequence_QC_Master']

wb_sample = openpyxl.load_workbook(sample_list_file, data_only=True)
ws_sample = wb_sample['Sheet1']

# Build mapping from Sample List: (Station, Depth) -> Lable ID_DOC
sample_map = {}
for r in range(2, ws_sample.max_row + 1):
    st = ws_sample.cell(r, 3).value
    dp = ws_sample.cell(r, 6).value
    lbl = ws_sample.cell(r, 7).value
    if st and dp is not None and lbl:
        st_norm = str(st).strip().upper().replace(' ', '')
        try:
            dp_norm = int(round(float(dp)))
        except:
            dp_norm = str(dp).strip()
        sample_map[(st_norm, dp_norm)] = str(lbl).strip()

out = []

out.append("=== ANALYSIS OF ZERO INJECTION / DRY DRAW ROWS ===")
zero_rows = []
for r in range(7, ws_master.max_row + 1):
    ctype = ws_master.cell(r, 3).value
    if ctype == 'SAMPLE':
        inj1 = ws_master.cell(r, 6).value or 0
        inj2 = ws_master.cell(r, 7).value or 0
        inj3 = ws_master.cell(r, 8).value or 0
        inj4 = ws_master.cell(r, 9).value or 0
        mean_area = ws_master.cell(r, 10).value or 0
        if inj1 == 0 and inj2 == 0 and inj3 == 0 and inj4 == 0:
            sname = ws_master.cell(r, 2).value
            st = ws_master.cell(r, 4).value
            dp = ws_master.cell(r, 5).value
            flag = ws_master.cell(r, 15).value
            comment = ws_master.cell(r, 16).value
            zero_rows.append((r, sname, st, dp, flag, comment))
            out.append(f"Row {r:4d} | Name: {sname} | Station: {st} | Depth: {dp} | Flag: {flag} | Comment: {comment}")

out.append(f"\nTotal Zero-Injection Sample Rows Found: {len(zero_rows)}")

out.append("\n=== ANALYSIS OF MAPPED ID MISMATCHES (Correction Candidates) ===")
mismatches = []
for r in range(7, ws_master.max_row + 1):
    ctype = ws_master.cell(r, 3).value
    if ctype == 'SAMPLE':
        sname = str(ws_master.cell(r, 2).value or '')
        st = ws_master.cell(r, 4).value
        dp = ws_master.cell(r, 5).value
        if st and dp is not None:
            st_norm = str(st).strip().upper().replace(' ', '')
            try:
                dp_norm = int(round(float(dp)))
            except:
                dp_norm = str(dp).strip()
            
            correct_lbl = sample_map.get((st_norm, dp_norm))
            if correct_lbl:
                # Extract original ID prefix e.g. SO308-41183 from SO308-41183-ST39-4000
                m = re.match(r'^(SO308-\d+)', sname)
                curr_id = m.group(1) if m else sname
                if curr_id != correct_lbl:
                    # Construct full formatted corrected sample name
                    if '-' in sname and len(sname.split('-')) >= 4:
                        # e.g., SO308-41183-ST39-4000 -> correct_lbl + '-ST39-4000'
                        parts = sname.split('-')
                        new_sname = f"{correct_lbl}-{''.join(parts[2:])}" if len(parts)==4 else f"{correct_lbl}-{st_norm}-{dp_norm}"
                    else:
                        new_sname = correct_lbl
                    mismatches.append((r, sname, correct_lbl, new_sname, st, dp))
                    out.append(f"Row {r:4d} | Current: {sname:<30s} | Station: {st} | Depth: {dp} | Correct ID: {correct_lbl:<12s}")

out.append(f"\nTotal Sample ID Mismatches Found: {len(mismatches)}")

with open('scratch/analyze_zeros_and_remappings.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))

print("Saved to scratch/analyze_zeros_and_remappings.txt")
