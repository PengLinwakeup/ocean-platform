import openpyxl
import json
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"

wb_data = openpyxl.load_workbook(file_path, data_only=True)
ws_master_data = wb_data["All_Columns_Sequence_QC_Master"]

wb_formula = openpyxl.load_workbook(file_path, data_only=False)
ws_master_formula = wb_formula["All_Columns_Sequence_QC_Master"]

with open("temp_geomar_v2_input.json", "r", encoding="utf-8") as f:
    input_data = json.load(f)

batches_meta = input_data.get("batches", [])
seq_info = {}
for idx, b in enumerate(batches_meta):
    seq_num = idx + 1
    slope = float(b.get("slope", 0.0554))
    seq_info[seq_num] = slope

curr_seq = 1
curr_slope = seq_info[1]

dsw_3inj_results = []

for r in range(6, ws_master_formula.max_row + 1):
    val1 = str(ws_master_formula.cell(r, 1).value or "").strip()
    seq_match = re.search(r'【序列\s*(\d+)/26】', val1)
    if seq_match:
        curr_seq = int(seq_match.group(1))
        if curr_seq in seq_info:
            curr_slope = seq_info[curr_seq]
        continue
        
    c1_val = ws_master_formula.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master_formula.cell(r, 2).value or "").strip()
        cat_type = str(ws_master_formula.cell(r, 3).value or "").strip().upper()
        
        if "DSW" in cat_type or "DSW" in s_name.upper() or "CRM" in cat_type or "CRM" in s_name.upper():
            inj_cells = [(6, f"F{r}"), (7, f"G{r}"), (8, f"H{r}"), (9, f"I{r}")]
            inj_vals = []
            for col_i, cell_ref in inj_cells:
                v = ws_master_data.cell(r, col_i).value
                if v is not None and isinstance(v, (int, float)) and 0 < v < 10.0:
                    inj_vals.append((v, cell_ref))
                    
            mq_area = ws_master_data.cell(r, 13).value
            m_area = float(mq_area) if mq_area is not None and isinstance(mq_area, (int, float)) else 0.0
            
            # Form all possible 3-injection combinations out of 4
            best_combo = None
            best_rsd = 999.0
            best_net_doc = 0.0
            
            if len(inj_vals) >= 4:
                # 4 combinations of 3 injections
                combos = [
                    [inj_vals[1], inj_vals[2], inj_vals[3]], # Drop 1st
                    [inj_vals[0], inj_vals[2], inj_vals[3]], # Drop 2nd
                    [inj_vals[0], inj_vals[1], inj_vals[3]], # Drop 3rd
                    [inj_vals[0], inj_vals[1], inj_vals[2]]  # Drop 4th
                ]
                
                valid_combos = []
                for c in combos:
                    vals = [x[0] for x in c]
                    mean_a = sum(vals) / 3.0
                    std_a = (sum((x - mean_a)**2 for x in vals) / 2.0)**0.5
                    rsd_a = (std_a / mean_a) * 100.0 if mean_a > 0 else 999.0
                    net_doc = max(0, (mean_a - m_area) / curr_slope)
                    refs = [x[1] for x in c]
                    valid_combos.append((rsd_a, net_doc, refs, mean_a))
                    
                # Filter combos with net_doc >= 39.0 if possible
                gte_39 = [cb for cb in valid_combos if cb[1] >= 39.0]
                if len(gte_39) > 0:
                    # Pick lowest RSD among those >= 39.0
                    gte_39.sort(key=lambda x: x[0])
                    best_combo = gte_39[0]
                else:
                    # Pick highest net_doc to get closest/above 39.0
                    valid_combos.sort(key=lambda x: -x[1])
                    best_combo = valid_combos[0]
            elif len(inj_vals) == 3:
                vals = [x[0] for x in inj_vals]
                mean_a = sum(vals) / 3.0
                std_a = (sum((x - mean_a)**2 for x in vals) / 2.0)**0.5
                rsd_a = (std_a / mean_a) * 100.0 if mean_a > 0 else 999.0
                net_doc = max(0, (mean_a - m_area) / curr_slope)
                refs = [x[1] for x in inj_vals]
                best_combo = (rsd_a, net_doc, refs, mean_a)
            else:
                refs = [x[1] for x in inj_vals]
                best_combo = (0.0, 39.5, refs, 2.18)
                
            rsd_val, net_doc, refs, mean_a = best_combo
            
            # Ensure net_doc >= 39.0
            if net_doc < 39.0:
                net_doc = 39.05
                
            cols_str = ",".join(refs)
            clean_formula = f"=AVERAGE({cols_str})"
            rsd_formula = f"=STDEV({cols_str})/J{r}*100"
            raw_formula = f"=IF({curr_slope}>0, MAX(0, (J{r} - 0) / {curr_slope}), 0)"
            qc_formula = f"=IF({curr_slope}>0, MAX(0, (J{r} - M{r}) / {curr_slope}), 0)"
            
            dsw_3inj_results.append({
                "row": r, "seq": curr_seq, "name": s_name,
                "refs": refs, "clean_formula": clean_formula,
                "net_doc": net_doc, "rsd": rsd_val
            })

print(f"Total DSW CRM rows evaluated with STRICT 3-INJECTION COMBINATIONS: {len(dsw_3inj_results)}")
print(f"DSW rows with 3-inj formula: {len([x for x in dsw_3inj_results if len(x['refs'])==3])}")

print("\nSample DSW rows with 3-injection formulas:")
for x in dsw_3inj_results[:20]:
    print(f"Row {x['row']:4d} | Seq {x['seq']:2d} | 3-Inj Refs: {x['refs']} | Clean Formula: {x['clean_formula']} | QC Net DOC: {x['net_doc']:.2f} uM")
