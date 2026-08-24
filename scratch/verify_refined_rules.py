import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

print("================ 1. VERIFYING MQ BLANK RE-EVALUATION (Rule 1) ================")
ws_all = wb["ODV_All_Samples_Full_List"]

mq_flag4_count = 0
mq_flag2_count = 0
for r in range(5, ws_all.max_row + 1):
    stype = str(ws_all.cell(r, 5).value or "").upper()
    sid = str(ws_all.cell(r, 4).value or "").upper()
    if "MQ" in sid or "BLANK" in sid:
        flag = ws_all.cell(r, 12).value
        comment = ws_all.cell(r, 13).value
        if flag == 4: mq_flag4_count += 1
        elif flag in [1, 2]: mq_flag2_count += 1
        
print(f"MQ Blanks with Flag 2 (Accepted): {mq_flag2_count}")
print(f"MQ Blanks with Flag 4 (Contaminated): {mq_flag4_count}")

print("\n================ 2. VERIFYING DSW CRM MINIMUM THRESHOLD (Rule 2) ================")
ws_master = wb["All_Columns_Sequence_QC_Master"]
dsw_below_39 = []
for r in range(6, ws_master.max_row + 1):
    cat = str(ws_master.cell(r, 3).value or "").upper()
    name = str(ws_master.cell(r, 2).value or "").upper()
    if "DSW" in cat or "DSW" in name or "CRM" in cat or "CRM" in name:
        doc = ws_master.cell(r, 14).value
        if doc is not None and isinstance(doc, (int, float)) and doc < 39.0 and doc > 0:
            dsw_below_39.append((r, name, doc))

print(f"Total DSW CRM rows with DOC < 39.0 uM C: {len(dsw_below_39)} (Expected: 0)")
if len(dsw_below_39) > 0:
    for r, name, doc in dsw_below_39:
        print(f"  Row {r:4d} | Name: {name} | DOC: {doc}")

print("\n================ 3. VERIFYING SPIKE ANOMALY ISOLATION (Rule 3) ================")
target_sample_found_in_clean = False
target_sample_found_in_flag4 = False

for r in range(5, ws_all.max_row + 1):
    sid = str(ws_all.cell(r, 4).value or "")
    if "41446" in sid:
        st = ws_all.cell(r, 3).value
        depth = ws_all.cell(r, 6).value
        doc = ws_all.cell(r, 11).value or ws_all.cell(r, 7).value
        flag = ws_all.cell(r, 12).value
        comment = ws_all.cell(r, 13).value
        status = ws_all.cell(r, 1).value
        print(f"Sample SO308-41446 (ST-49 90m): Status='{status}', Flag={flag}, DOC={doc}, Comment='{comment}'")

ws_clean = wb["ODV_Clean_Export_Only"]
for r in range(5, ws_clean.max_row + 1):
    sid = str(ws_clean.cell(r, 3).value or "")
    if "41446" in sid:
        target_sample_found_in_clean = True

ws_flag4 = wb["Flag4_Discarded_Audit_List"]
for r in range(6, ws_flag4.max_row + 1):
    sid = str(ws_flag4.cell(r, 5).value or "")
    if "41446" in sid:
        target_sample_found_in_flag4 = True

print(f"SO308-41446 in ODV_Clean_Export_Only: {target_sample_found_in_clean} (Expected: False)")
print(f"SO308-41446 in Flag4_Discarded_Audit_List: {target_sample_found_in_flag4} (Expected: True)")
