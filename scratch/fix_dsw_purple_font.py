import openpyxl
from openpyxl.styles import Font
import os
import glob
from datetime import datetime

print("========================================================================")
print("        DSW Purple Font Calibration & Reset to Standard Black Tool")
print("========================================================================")

cleaned_files = glob.glob(r'F:\印度洋测样\ODV\202608\20260829\*Cleaned*.xlsx')
if not cleaned_files:
    raise FileNotFoundError("No cleaned input file found!")

latest_input_file = max(cleaned_files, key=os.path.getmtime)
print(f"[*] Loading workbook (formulas preserved): {latest_input_file}")

wb = openpyxl.load_workbook(latest_input_file, data_only=False)
ws_master = wb['All_Columns_Sequence_QC_Master']

black_color = 'FF000000'
dsw_purple_fixed_count = 0
logs = []

for r in range(7, ws_master.max_row + 1):
    c2_val = str(ws_master.cell(r, 2).value or '').strip().upper()
    c3_val = str(ws_master.cell(r, 3).value or '').strip().upper()
    
    is_dsw = ('DSW' in c2_val or 'DSW' in c3_val)
    
    if is_dsw:
        for c in range(1, ws_master.max_column + 1):
            cell = ws_master.cell(r, c)
            if cell.font and cell.font.color:
                color_str = str(cell.font.color.rgb) if cell.font.color.rgb else ''
                # Check for purple font hues
                if any(p in color_str.upper() for p in ['6B21A8', 'PURPLE', '7E22CE', '9333EA', '581C87']):
                    # Preserve original font parameters (name, size, bold, italic), only change color to black
                    old_font = cell.font
                    cell.font = Font(
                        name=old_font.name or 'Calibri',
                        size=old_font.size or 11,
                        bold=old_font.bold,
                        italic=old_font.italic,
                        color=black_color
                    )
                    dsw_purple_fixed_count += 1
                    logs.append(f"[DSW FONT FIX] Row {r:4d} Col {c:2d} | Val: {cell.value!r:<30s} | Color: '{color_str}' -> '{black_color}'")

timestamp = datetime.now().strftime('%Y%m%d_%H%M')
out_dir = os.path.dirname(latest_input_file)
out_filename = f"Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed-20260829_Final_QC_Completed_Cleaned_{timestamp}.xlsx"
out_filepath = os.path.join(out_dir, out_filename)

print(f"\n[*] Saving updated workbook to: {out_filepath}")
wb.save(out_filepath)

print("\n========================================================================")
print("                        PROCESSING SUMMARY")
print("========================================================================")
print(f"DSW Purple Font Cells Converted to Black : {dsw_purple_fixed_count}")
print("========================================================================")

log_path = 'scratch/dsw_font_fix_log.txt'
with open(log_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(logs))

print(f"Detailed log saved to {log_path}")
