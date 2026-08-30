import openpyxl
import os
import glob

cleaned_files = glob.glob(r'F:\印度洋测样\ODV\202608\20260829\*Cleaned*.xlsx')
if not cleaned_files:
    raise FileNotFoundError("No cleaned excel file found!")

latest_cleaned_file = max(cleaned_files, key=os.path.getmtime)
print(f"[*] Verifying latest output file: {latest_cleaned_file}")

wb = openpyxl.load_workbook(latest_cleaned_file, data_only=False)
ws_master = wb['All_Columns_Sequence_QC_Master']

checks = {
    'dsw_purple_font_remaining': 0,
    'dsw_black_font_count': 0,
    'total_dsw_rows': 0,
    'formulas_intact_count': 0
}

for r in range(7, ws_master.max_row + 1):
    c2_val = str(ws_master.cell(r, 2).value or '').strip().upper()
    c3_val = str(ws_master.cell(r, 3).value or '').strip().upper()
    
    is_dsw = ('DSW' in c2_val or 'DSW' in c3_val)
    
    if is_dsw:
        checks['total_dsw_rows'] += 1
        for c in range(1, ws_master.max_column + 1):
            cell = ws_master.cell(r, c)
            if cell.font and cell.font.color:
                color_str = str(cell.font.color.rgb) if cell.font.color.rgb else ''
                if any(p in color_str.upper() for p in ['6B21A8', 'PURPLE', '7E22CE', '9333EA', '581C87']):
                    checks['dsw_purple_font_remaining'] += 1
                elif '000000' in color_str or 'FF000000' in color_str:
                    checks['dsw_black_font_count'] += 1

    for c in [21, 23]:
        v = ws_master.cell(r, c).value
        if isinstance(v, str) and v.startswith('='):
            checks['formulas_intact_count'] += 1

print("\n========================================================================")
print("                        VERIFICATION RESULTS")
print("========================================================================")
print(f"Total DSW Rows Verified                : {checks['total_dsw_rows']}")
print(f"DSW Purple Font Cells Remaining        : {checks['dsw_purple_font_remaining']} (Expected: 0)")
print(f"DSW Converted Black Font Cells Count   : {checks['dsw_black_font_count']} (Expected: 87)")
print(f"Intact Excel Formulas Count           : {checks['formulas_intact_count']} (Expected: >= 701)")
print("========================================================================")

if checks['dsw_purple_font_remaining'] == 0 and checks['dsw_black_font_count'] == 87 and checks['formulas_intact_count'] >= 701:
    print("[SUCCESS] All 100% verification checks passed flawlessly!")
else:
    print("[WARNING] Verification check mismatch.")
