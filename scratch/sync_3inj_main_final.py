import openpyxl
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
inj3_file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest_3INJ_DSW.xlsx"
backup11_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest_backup11.xlsx"

# 1. Create backup11
shutil.copyfile(file_path, backup11_path)
print(f"Backup created: {backup11_path}")

# 2. Copy 3-injection DSW file directly to main file
shutil.copyfile(inj3_file_path, file_path)
print(f"Successfully copied 3-injection DSW file to main file: {file_path}")

# 3. Verify main file
wb_formula = openpyxl.load_workbook(file_path, data_only=False)
ws_master_formula = wb_formula["All_Columns_Sequence_QC_Master"]

wb_data = openpyxl.load_workbook(file_path, data_only=True)
ws_master_data = wb_data["All_Columns_Sequence_QC_Master"]

print("\n================ DEEP VERIFYING 3-INJECTION DSW FORMULAS ON MAIN FILE ================")

# Test Row 34 (from user's screenshot)
r = 34
print(f"\nInspecting Row {r} (User's Screenshot Row 34 DSW):")
for col in [1, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]:
    c_f = ws_master_formula.cell(r, col)
    c_d = ws_master_data.cell(r, col)
    fill = c_f.fill
    fill_type = fill.fill_type if fill else None
    fg = fill.fgColor.value if (fill and fill.fgColor) else None
    print(f"Col {col:2d} | Formula/Val: {str(c_f.value):35s} | Eval Value: {str(c_d.value):10s} | FillType: {fill_type} | FgColor: {fg}")

# Count 3-injection formulas
three_inj_count = 0
for row_i in range(6, ws_master_formula.max_row + 1):
    c1_val = ws_master_formula.cell(row_i, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master_formula.cell(row_i, 2).value or "").strip()
        cat_type = str(ws_master_formula.cell(row_i, 3).value or "").strip().upper()
        if "DSW" in cat_type or "DSW" in s_name.upper() or "CRM" in cat_type or "CRM" in s_name.upper():
            f_val = str(ws_master_formula.cell(row_i, 10).value or "")
            # Check number of commas in =AVERAGE(F34,G34,H34)
            if "=AVERAGE(" in f_val and f_val.count(",") == 2:
                three_inj_count += 1

print(f"\nTotal DSW CRM rows verified with STRICT 3-CELL AVERAGE formulas `=AVERAGE(cell1,cell2,cell3)`: {three_inj_count} (Expected: 190)")
