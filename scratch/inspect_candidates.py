import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

ws_master = wb["All_Columns_Sequence_QC_Master"]

print("================ 1. MQ BLANKS CANDIDATES (Area <= 0.15 or DOC <= 0.9 uM) ================")
mq_flag4_to_flag2 = []
for r in range(6, ws_master.max_row + 1):
    cat = str(ws_master.cell(r, 3).value or "").upper()
    name = str(ws_master.cell(r, 2).value or "").upper()
    if "MQ" in cat or "MQ" in name or "BLANK" in cat or "BLANK" in name:
        area = ws_master.cell(r, 10).value
        rsd = ws_master.cell(r, 11).value
        doc = ws_master.cell(r, 14).value
        flag = ws_master.cell(r, 15).value if ws_master.max_column >= 15 else None
        diag = ws_master.cell(r, 16).value if ws_master.max_column >= 16 else None
        
        area_val = float(area) if area is not None and isinstance(area, (int, float)) else 0.0
        doc_val = float(doc) if doc is not None and isinstance(doc, (int, float)) else 0.0
        
        if (area_val <= 0.15 or doc_val <= 0.9) and flag == 4:
            mq_flag4_to_flag2.append((r, name, area_val, rsd, doc_val, diag))

print(f"MQ Blank rows to correct from Flag 4 -> Flag 2: {len(mq_flag4_to_flag2)}")
for r, name, area_val, rsd, doc_val, diag in mq_flag4_to_flag2:
    print(f"  Row {r:4d} | Name: {name:20s} | Area: {area_val:.4f} | RSD%: {rsd} | DOC: {doc_val:.2f} | Diag: {diag}")

print("\n================ 2. DSW CANDIDATES (DOC < 39.0 uM) ================")
dsw_low_candidates = []
for r in range(6, ws_master.max_row + 1):
    cat = str(ws_master.cell(r, 3).value or "").upper()
    name = str(ws_master.cell(r, 2).value or "").upper()
    if "DSW" in cat or "DSW" in name or "CRM" in cat or "CRM" in name:
        inj1 = ws_master.cell(r, 6).value
        inj2 = ws_master.cell(r, 7).value
        inj3 = ws_master.cell(r, 8).value
        inj4 = ws_master.cell(r, 9).value
        area = ws_master.cell(r, 10).value
        doc = ws_master.cell(r, 14).value
        
        doc_val = float(doc) if doc is not None and isinstance(doc, (int, float)) else 0.0
        if doc_val < 39.0 and doc_val > 0:
            injs = [inj1, inj2, inj3, inj4]
            injs_clean = [float(x) for x in injs if x is not None and isinstance(x, (int, float)) and x > 0]
            dsw_low_candidates.append((r, name, doc_val, area, injs_clean))

print(f"DSW rows with DOC < 39.0 uM: {len(dsw_low_candidates)}")
for r, name, doc_val, area, injs in dsw_low_candidates[:15]:
    print(f"  Row {r:4d} | Name: {name:25s} | DOC: {doc_val:.2f} | Area: {area} | Injections: {injs}")

print("\n================ 3. SEAWATER SPIKE ANOMALY CANDIDATES ================")
ws_all = wb["ODV_All_Samples_Full_List"]

all_samples = []
for r in range(5, ws_all.max_row + 1):
    status = ws_all.cell(r, 1).value
    seq = ws_all.cell(r, 2).value
    st = ws_all.cell(r, 3).value
    sid = ws_all.cell(r, 4).value
    stype = ws_all.cell(r, 5).value
    depth = ws_all.cell(r, 6).value
    doc = ws_all.cell(r, 10).value if ws_all.cell(r, 10).value is not None else ws_all.cell(r, 7).value
    flag = ws_all.cell(r, 11).value or ws_all.cell(r, 12).value
    comment = ws_all.cell(r, 12).value or ws_all.cell(r, 13).value
    
    if stype == "SAMPLE" and doc is not None and isinstance(doc, (int, float)):
        all_samples.append({
            "row": r,
            "status": status,
            "seq": seq,
            "st": st,
            "sid": sid,
            "depth": float(depth) if depth is not None and str(depth).replace('.','').isdigit() else 0.0,
            "doc": float(doc),
            "flag": flag,
            "comment": comment
        })

# Group samples by station
st_map = {}
for s in all_samples:
    st = s["st"]
    if st not in st_map: st_map[st] = []
    st_map[st].append(s)

spikes = []
for st, s_list in st_map.items():
    s_list.sort(key=lambda x: x["depth"])
    for i, s in enumerate(s_list):
        doc = s["doc"]
        # Find neighbors at same station
        neighbors = []
        if i > 0: neighbors.append(s_list[i-1]["doc"])
        if i < len(s_list) - 1: neighbors.append(s_list[i+1]["doc"])
        
        if len(neighbors) > 0:
            avg_n = sum(neighbors) / len(neighbors)
            if doc > 100.0 and doc > 1.5 * avg_n and avg_n < 90.0:
                spikes.append((s, avg_n, neighbors))

print(f"Isolated High-DOC Spike Anomaly candidates: {len(spikes)}")
for s, avg_n, neighbors in spikes:
    print(f"  Row {s['row']:4d} | ST: {s['st']:6s} | ID: {s['sid']:22s} | Depth: {s['depth']:6.1f}m | DOC: {s['doc']:6.2f} uM | Neighbors Avg: {avg_n:6.2f} uM | Current Flag: {s['flag']}")
