import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import re
import shutil
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest.xlsx"
backup_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest_backup.xlsx"

# 0. Backup
shutil.copyfile(file_path, backup_path)
print(f"Created backup: {backup_path}")

# 1. Load batch metadata from temp_geomar_v2_input.json
with open("temp_geomar_v2_input.json", "r", encoding="utf-8") as f:
    input_data = json.load(f)

batches_meta = input_data.get("batches", [])

# Construct metadata mapping for each batch (index 1 to 26)
seq_info = {}

for idx, b in enumerate(batches_meta):
    seq_num = idx + 1
    col_idx = b.get("fileColIdx", seq_num)
    file_name = b.get("fileName", "")
    slope = float(b.get("slope", 0.0554))
    rsq = float(b.get("rsq") or b.get("r2", 0.999))
    samples = b.get("samples", [])
    
    # Extract unique station names
    stations = []
    for s in samples:
        st = (s.get("station") or "").strip()
        if st and st != "-" and not st.upper().startswith("STD") and "工作曲线" not in st and st not in stations:
            stations.append(st)
            
    if seq_num == 26:
        st_summary = "ST-1~7,ST-19"
    elif len(stations) > 4:
        st_summary = ",".join(stations[:3]) + f"等{len(stations)}个站位"
    elif len(stations) > 0:
        st_summary = ",".join(stations)
    else:
        st_summary = "STD/Blank"
        
    clean_seq_name = f"【序列 {seq_num}/26】柱{col_idx}_工作曲线-站位{st_summary}"
    
    seq_info[seq_num] = {
        "seq_num": seq_num,
        "col_idx": col_idx,
        "file_name": file_name,
        "slope": slope,
        "rsq": rsq,
        "clean_seq_name": clean_seq_name,
        "stations": stations
    }

print(f"Parsed metadata for {len(seq_info)} sequences.")

# 2. Load openpyxl workbook
wb = openpyxl.load_workbook(file_path)

font_header = Font(name='楷体', size=11, bold=True, color='FFFFFF')
font_chinese = Font(name='楷体', size=10.5, color='1E293B')
font_number = Font(name='Times New Roman', size=10.5, color='1E293B')
font_bold_number = Font(name='Times New Roman', size=10.5, bold=True, color='1E3A8A')
fill_navy = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

border_thin = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# ---------------------------------------------------------
# A. Executive_Dashboard Update
# ---------------------------------------------------------
if "Executive_Dashboard" in wb.sheetnames:
    ws_dash = wb["Executive_Dashboard"]
    print("Updating Executive_Dashboard sheet...")
    
    # Format header row 7
    ws_dash.row_dimensions[7].height = 28
    for c in range(1, 12):
        cell = ws_dash.cell(7, c)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    # Update rows 8 to 33
    for seq_num in range(1, 27):
        r = 7 + seq_num
        info = seq_info[seq_num]
        ws_dash.row_dimensions[r].height = 22
        
        # Col A: Index
        ws_dash.cell(r, 1).value = seq_num
        # Col B: Sequence / Sheet Name
        ws_dash.cell(r, 2).value = info["clean_seq_name"]
        # Col D: Linearity R2
        cell_d = ws_dash.cell(r, 4)
        cell_d.value = round(info["rsq"], 5)
        cell_d.number_format = '0.00000'
        
        # Col E: Slope (m)
        cell_e = ws_dash.cell(r, 5)
        cell_e.value = round(info["slope"], 6)
        cell_e.number_format = '0.000000'
        
        # Format styling for row cells
        is_even = seq_num % 2 == 0
        row_fill = fill_zebra if is_even else fill_white
        
        for c in range(1, 12):
            cell = ws_dash.cell(r, c)
            cell.fill = row_fill
            cell.border = border_thin
            if c == 2:
                cell.font = font_chinese
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif c in [4, 5]:
                cell.font = font_bold_number
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = font_number
                cell.alignment = Alignment(horizontal='center', vertical='center')

# ---------------------------------------------------------
# B. All_Columns_Sequence_QC_Master Update
# ---------------------------------------------------------
if "All_Columns_Sequence_QC_Master" in wb.sheetnames:
    ws_master = wb["All_Columns_Sequence_QC_Master"]
    print("Updating All_Columns_Sequence_QC_Master sheet...")
    
    curr_seq_num = 0
    curr_slope = 0.0554
    curr_rsq = 0.999
    
    for r in range(1, ws_master.max_row + 1):
        val1 = str(ws_master.cell(r, 1).value or "").strip()
        
        # Check sequence title row
        seq_match = re.search(r'【序列\s*(\d+)/26】', val1)
        if seq_match:
            curr_seq_num = int(seq_match.group(1))
            if curr_seq_num in seq_info:
                info = seq_info[curr_seq_num]
                curr_slope = info["slope"]
                curr_rsq = info["rsq"]
                ws_master.cell(r, 1).value = f"{info['clean_seq_name']}  {info['clean_seq_name']}"
            continue
            
        # Check summary stats line right below title
        if "数据源:" in val1 or "R²:" in val1 or "斜率:" in val1:
            if curr_seq_num in seq_info:
                # Update text preserving MQ drift, DSW recovery, QC pass rate
                old_text = val1
                # Replace R²: x.xxxxx and 斜率: x.xxxxx
                new_text = re.sub(r'R²:\s*[\d\.]+', f'R²: {curr_rsq:.5f}', old_text)
                new_text = re.sub(r'斜率:\s*[\d\.]+', f'斜率: {curr_slope:.5f}', new_text)
                ws_master.cell(r, 1).value = new_text
            continue
            
        # Check data row (starts with integer seq order)
        c1_val = ws_master.cell(r, 1).value
        if c1_val is not None and isinstance(c1_val, (int, float)):
            # Col L (Raw DOC) = IF(slope > 0, MAX(0, (J{r} - 0) / slope), 0)
            raw_doc_cell = ws_master.cell(r, 12)
            raw_doc_cell.value = f"=IF({curr_slope:.6f}>0, MAX(0, (J{r} - 0) / {curr_slope:.6f}), 0)"
            raw_doc_cell.number_format = '0.00'
            
            # Col N (QC Dynamic DOC) = IF(slope > 0, MAX(0, (J{r} - M{r}) / slope), 0)
            qc_doc_cell = ws_master.cell(r, 14)
            qc_doc_cell.value = f"=IF({curr_slope:.6f}>0, MAX(0, (J{r} - M{r}) / {curr_slope:.6f}), 0)"
            qc_doc_cell.number_format = '0.00'

# ---------------------------------------------------------
# C. ODV_All_Samples_Full_List Update
# ---------------------------------------------------------
if "ODV_All_Samples_Full_List" in wb.sheetnames:
    ws_all = wb["ODV_All_Samples_Full_List"]
    print("Updating ODV_All_Samples_Full_List sheet...")
    
    for r in range(5, ws_all.max_row + 1):
        seq_cell = ws_all.cell(r, 2)
        seq_text = str(seq_cell.value or "").strip()
        seq_match = re.search(r'【序列\s*(\d+)/26】', seq_text)
        if seq_match:
            seq_num = int(seq_match.group(1))
            if seq_num in seq_info:
                info = seq_info[seq_num]
                # Update Sequence Run name
                seq_cell.value = info["clean_seq_name"]
                
                # Update Raw DOC formula (Col 7 / G)
                raw_doc_cell = ws_all.cell(r, 7)
                raw_doc_cell.value = f"=IF({info['slope']:.6f}>0, MAX(0, (I{r} - 0) / {info['slope']:.6f}), 0)"
                raw_doc_cell.number_format = '0.00'

# ---------------------------------------------------------
# D. ODV_Clean_Export_Only Update
# ---------------------------------------------------------
if "ODV_Clean_Export_Only" in wb.sheetnames:
    ws_clean = wb["ODV_Clean_Export_Only"]
    print("Updating ODV_Clean_Export_Only sheet...")
    
    for r in range(5, ws_clean.max_row + 1):
        seq_cell = ws_clean.cell(r, 1)
        seq_text = str(seq_cell.value or "").strip()
        seq_match = re.search(r'【序列\s*(\d+)/26】', seq_text)
        if seq_match:
            seq_num = int(seq_match.group(1))
            if seq_num in seq_info:
                info = seq_info[seq_num]
                # Update Sequence Run name
                seq_cell.value = info["clean_seq_name"]
                
                # Update Raw DOC formula (Col 6 / F)
                raw_doc_cell = ws_clean.cell(r, 6)
                raw_doc_cell.value = f"=IF({info['slope']:.6f}>0, MAX(0, (H{r} - 0) / {info['slope']:.6f}), 0)"
                raw_doc_cell.number_format = '0.00'

# ---------------------------------------------------------
# E. Flag4_Discarded_Audit_List Update
# ---------------------------------------------------------
if "Flag4_Discarded_Audit_List" in wb.sheetnames:
    ws_flag4 = wb["Flag4_Discarded_Audit_List"]
    print("Updating Flag4_Discarded_Audit_List sheet...")
    
    for r in range(6, ws_flag4.max_row + 1):
        seq_cell = ws_flag4.cell(r, 2)
        seq_text = str(seq_cell.value or "").strip()
        seq_match = re.search(r'【序列\s*(\d+)/26】', seq_text)
        if seq_match:
            seq_num = int(seq_match.group(1))
            if seq_num in seq_info:
                info = seq_info[seq_num]
                col_num = info["col_idx"]
                # Update Sequence Run name
                seq_cell.value = f"第 {col_num} 柱 ({info['clean_seq_name']})"
                
                # Update Raw DOC value if numeric formula needed
                area_val = ws_flag4.cell(r, 10).value
                if area_val is not None and isinstance(area_val, (int, float)):
                    calc_doc = round(max(0, area_val / info['slope']), 2)
                    ws_flag4.cell(r, 12).value = calc_doc

# Save updated workbook
print("Saving modified workbook back to original path...")
wb.save(file_path)
print("Workbook saved successfully!")
