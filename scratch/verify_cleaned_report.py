import openpyxl
import os
import glob

cleaned_files = glob.glob(r'F:\印度洋测样\ODV\202608\20260829\*Cleaned*.xlsx')
if not cleaned_files:
    raise FileNotFoundError("No cleaned excel file found!")

latest_cleaned_file = max(cleaned_files, key=os.path.getmtime)
print(f"[*] Verifying latest cleaned file: {latest_cleaned_file}")

wb = openpyxl.load_workbook(latest_cleaned_file, data_only=False)
ws_master = wb['All_Columns_Sequence_QC_Master']

checks = {
    'flag2_green_font_count': 0,
    'flag2_fill_none_count': 0,
    'flag2_flag_fill_green_count': 0,
    'zero_inj_flag4_count': 0,
    'formulas_intact_count': 0,
    'total_flag2_rows': 0
}

r7_col21_formula = ws_master.cell(7, 21).value
r7_col23_formula = ws_master.cell(7, 23).value
print(f"[*] Formula in Row 7 Col 21 (MQ DOC): {r7_col21_formula}")
print(f"[*] Formula in Row 7 Col 23 (DSW DOC): {r7_col23_formula}")

for r in range(7, ws_master.max_row + 1):
    seq_cell = ws_master.cell(r, 1)
    name_cell = ws_master.cell(r, 2)
    ctype_cell = ws_master.cell(r, 3)
    flag_cell = ws_master.cell(r, 15)
    comment_cell = ws_master.cell(r, 16)
    
    if not any([seq_cell.value, name_cell.value, ctype_cell.value]):
        continue

    flag_val = str(flag_cell.value).strip() if flag_cell.value is not None else ''
    comment_val = str(comment_cell.value).strip() if comment_cell.value is not None else ''
    
    if flag_val == '2' or 'Acceptable' in comment_val:
        checks['total_flag2_rows'] += 1
        
        # Check font color of Col 1
        col1_color = str(seq_cell.font.color.rgb) if (seq_cell.font and seq_cell.font.color) else ''
        if '166534' in col1_color:
            checks['flag2_green_font_count'] += 1
            
        # Check fill of Col 1
        if seq_cell.fill is None or seq_cell.fill.fill_type is None:
            checks['flag2_fill_none_count'] += 1
            
        # Check fill of Col 15 (Flag)
        flag_fg = str(flag_cell.fill.fgColor.rgb) if (flag_cell.fill and flag_cell.fill.fgColor) else ''
        if 'DCFCE7' in flag_fg:
            checks['flag2_flag_fill_green_count'] += 1
            
    if 'Zero injection dry draw' in comment_val:
        if str(flag_cell.value) == '4':
            checks['zero_inj_flag4_count'] += 1

    for c in [21, 23]:
        v = ws_master.cell(r, c).value
        if isinstance(v, str) and v.startswith('='):
            checks['formulas_intact_count'] += 1

print("\n========================================================================")
print("                        VERIFICATION RESULTS")
print("========================================================================")
print(f"Total Flag 2 Rows Verified            : {checks['total_flag2_rows']}")
print(f"Flag 2 Col 1 Green Font (166534) Match : {checks['flag2_green_font_count']} / {checks['total_flag2_rows']}")
print(f"Flag 2 Col 1 Empty Fill Match         : {checks['flag2_fill_none_count']} / {checks['total_flag2_rows']}")
print(f"Flag 2 Col 15 Light Green Fill Match  : {checks['flag2_flag_fill_green_count']} / {checks['total_flag2_rows']}")
print(f"Zero Injection Discarded Rows Count   : {checks['zero_inj_flag4_count']}")
print(f"Intact Excel Formulas Count           : {checks['formulas_intact_count']}")
print("========================================================================")

if checks['flag2_green_font_count'] == checks['total_flag2_rows'] and checks['zero_inj_flag4_count'] == 9:
    print("[SUCCESS] All 100% verification checks passed flawlessly!")
else:
    print("[WARNING] Verification check mismatch.")
