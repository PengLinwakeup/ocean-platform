import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Load backup3 (which has pre-overwritten data) or backup2 to check original master flags & RSDs
file_path = r"F:\印度洋测样\ODV\202608\20260824\Ocean_DOC_MultiColumn_QC_Report_GEOMAR_Validated_v2_Processed_latest_backup3.xlsx"
wb_data = openpyxl.load_workbook(file_path, data_only=True)

ws_master = wb_data["All_Columns_Sequence_QC_Master"]
ws_all = wb_data["ODV_All_Samples_Full_List"]

# Build map from Full List for seawater sample flags & comments
full_list_map = {}
for r in range(5, ws_all.max_row + 1):
    sid = str(ws_all.cell(r, 4).value or "").strip()
    st = str(ws_all.cell(r, 3).value or "").strip()
    flag = ws_all.cell(r, 12).value
    comment = ws_all.cell(r, 13).value
    qc_doc = ws_all.cell(r, 11).value
    rsd = ws_all.cell(r, 10).value
    if sid:
        full_list_map[sid] = {"flag": flag, "comment": comment, "qc_doc": qc_doc, "rsd": rsd}

master_eval_counts = {1: 0, 2: 0, 3: 0, 4: 0, "other": 0}
sample_details = []

for r in range(6, ws_master.max_row + 1):
    c1_val = ws_master.cell(r, 1).value
    if c1_val is not None and isinstance(c1_val, (int, float)):
        s_name = str(ws_master.cell(r, 2).value or "").strip()
        cat_type = str(ws_master.cell(r, 3).value or "").strip().upper()
        area = ws_master.cell(r, 10).value
        rsd_val = ws_master.cell(r, 11).value
        raw_doc = ws_master.cell(r, 12).value
        qc_doc = ws_master.cell(r, 14).value
        
        area_num = float(area) if area is not None and isinstance(area, (int, float)) else 0.0
        rsd_num = float(rsd_val) if rsd_val is not None and isinstance(rsd_val, (int, float)) else 0.0
        doc_num = float(qc_doc) if qc_doc is not None and isinstance(qc_doc, (int, float)) else (float(raw_doc) if raw_doc is not None and isinstance(raw_doc, (int, float)) else 0.0)
        
        flag = 2
        comment = "Acceptable: Good quality"
        
        # 1. STD
        if "STD" in cat_type or "STD" in s_name.upper():
            flag = 2
            comment = "Acceptable: Calibration standard injection"
            
        # 2. MQ Blank
        elif "MQ" in cat_type or "MQ" in s_name.upper() or "BLANK" in cat_type or "BLANK" in s_name.upper():
            if area_num <= 0.15 or doc_num <= 0.9:
                flag = 2
                comment = "Acceptable: Low absolute area MQ blank (Area <= 0.15 / DOC <= 0.9 uM)"
            else:
                if rsd_num > 5.0:
                    flag = 4
                    comment = f"Rejected: Contaminated MQ blank (Area {area_num:.4f} > 0.15, DOC {doc_num:.2f} uM)"
                else:
                    flag = 2
                    comment = "Acceptable: Normal MQ blank"
                    
        # 3. DSW CRM
        elif "DSW" in cat_type or "DSW" in s_name.upper() or "CRM" in cat_type or "CRM" in s_name.upper():
            if doc_num >= 39.0:
                rec = round((doc_num / 39.45) * 100, 1)
                flag = 2
                comment = f"Acceptable: CRM DSW recovery in standard range ({doc_num:.1f} uM, {rec:.1f}%)"
            else:
                rec = round((doc_num / 39.45) * 100, 1)
                flag = 3
                comment = f"Questionable: Low CRM DSW recovery ({doc_num:.1f} uM < 39.0 uM, {rec:.1f}%)"
                
        # 4. Field Seawater Sample
        else:
            if s_name in full_list_map:
                item = full_list_map[s_name]
                flag = item["flag"]
                comment = item["comment"]
            else:
                if rsd_num > 5.0 and doc_num > 10.0:
                    flag = 4
                    comment = f"Rejected: High injection RSD ({rsd_num:.1f}% > 5.0%)"
                elif rsd_num > 3.0:
                    flag = 3
                    comment = f"Questionable: Moderate injection RSD ({rsd_num:.1f}% > 3.0%)"
                else:
                    flag = 2
                    comment = f"Acceptable: Low injection RSD ({rsd_num:.1f}% <= 3.0%)"

        if flag in master_eval_counts:
            master_eval_counts[flag] += 1
        sample_details.append((r, s_name, cat_type, flag, comment))

print("Master Sheet Proper Evaluation Flag Counts:")
print(master_eval_counts)

print("\nSample rows of proper evaluation:")
for r, s_name, cat_type, flag, comment in sample_details[30:50]:
    print(f"Row {r:4d} | Name: {s_name:25s} | Type: {cat_type:10s} | Flag: {flag} | Comment: {comment}")
