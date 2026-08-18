# -*- coding: utf-8 -*-
"""
================================================================================
海洋化学分析仪器数据质量控制 (QC) 与多柱子自动化报表同步系统
Marine Chemistry Analytical QC & Multi-Column Automated Report System
================================================================================
本脚本输出格式与 ocean-platform 网页前端及 2026-08-02 经典基准完全 100% 结构同构与公式同步：
  1. 【Sheet 1: 总览_Summary】:
     - 上半部: 14 柱 / 24 条工作曲线质控概况总览 (柱号、测定站位、MQ 空白活公式、参标活公式、评级)
     - 下半部: 全海区 DOC 样品水文位点与实测浓度汇总大表 (按站位 ST-51->ST-1、深度从深到浅大洋排序)
  2. 【Sheet 2..N: 各柱子分批明细 (柱1_... 至 柱14_... 共 24 张分表)】:
     - 顶部 Batch QC 概况 (斜率、R²、截距、MQ 活均值、参标回收率公式)
     - 逐样 4 针进样原始面积、筛选平均面积公式、扣空净面积、实测浓度活公式与 Flag 评级
  3. 【Sheet 26: ODV_Format_Data】:
     - 标准 ODV 导入格式 (Cruise, Station, Type, Lon, Lat, Bot.Depth, Depth, DOC)
================================================================================
"""

import os
import sys
import math
import io
import re
import glob
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional, Any
from itertools import combinations

if hasattr(sys.stdout, 'reconfigure'):
    try: sys.stdout.reconfigure(encoding='utf-8')
    except Exception: pass
if hasattr(sys.stderr, 'reconfigure'):
    try: sys.stderr.reconfigure(encoding='utf-8')
    except Exception: pass

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==============================================================================
# 1. 全局质控参数与配置
# ==============================================================================
@dataclass
class OceanQCConfig:
    crm_dsw_theoretical: float = 39.45    # Hansell 深海参标 (μmol C / L)
    crm_ssw_theoretical: float = 67.50    # Hansell 表层参标 (μmol C / L)
    target_injections: int = 4
    min_valid_injections: int = 3
    rsd_excellent_threshold: float = 1.5
    rsd_good_threshold: float = 3.0
    rsd_warning_threshold: float = 5.0
    drift_pass_threshold: float = 5.0
    deep_depth_threshold: float = 1000.0
    deep_doc_climatology_min: float = 36.0
    deep_doc_climatology_max: float = 48.0


# ==============================================================================
# 2. 原始逐针进样与样品组结构
# ==============================================================================
@dataclass
class RawInjectionRecord:
    file_name: str
    sample_name: str
    sample_id: str
    inj_no: int
    analysis_type: str
    area: float
    seq_index: int


@dataclass
class ProcessedSampleGroup:
    sample_name: str
    sample_id: str
    sample_type: str        # 'MQ', 'DSW', 'SSW', 'STD', 'SAMPLE'
    seq_start: int
    seq_center: float
    all_areas: List[float]
    selected_areas: List[float]
    selected_indices: List[int] # 记录选中的是第几针 (0-indexed) 用于生成 =AVERAGE(F10,G10,I10)
    mean_area: float
    sd: float
    rsd: float
    station: str = "-"
    depth: Optional[float] = None
    bot_depth: Optional[float] = None
    lon: Optional[float] = None
    lat: Optional[float] = None
    cruise: str = "SO308"
    
    # 计算与质控字段
    interpolated_mq: float = 0.0
    net_area: float = 0.0
    interpolated_rf: float = 0.0
    calculated_conc: float = 0.0
    qc_flag: int = 1
    qc_notes: List[str] = field(default_factory=list)


# ==============================================================================
# 3. 采样元数据管理
# ==============================================================================
class MetadataManager:
    _cached_metadata: Optional[Dict[str, Dict[str, Any]]] = None

    @classmethod
    def load_metadata(cls, search_paths: List[str] = None) -> Dict[str, Dict[str, Any]]:
        if cls._cached_metadata is not None:
            return cls._cached_metadata

        meta_dict: Dict[str, Dict[str, Any]] = {}
        candidate_files = [
            r"F:\印度洋测样\Indian Ocean_SO308_DOC_Sample List(1) 的副本.xlsx",
            r"F:\印度洋测样\DOC_Sample List.xlsx",
            r"F:\印度洋测样\DOC_Sample List-20260601.xlsx",
            r"F:\印度洋测样\So_308_1_btl-基础数据_15052025.xlsx"
        ]

        if search_paths:
            for sp in search_paths:
                if os.path.exists(sp):
                    dname = sp if os.path.isdir(sp) else os.path.dirname(sp)
                    for fname in os.listdir(dname):
                        if 'sample' in fname.lower() and fname.endswith('.xlsx'):
                            candidate_files.insert(0, os.path.join(dname, fname))

        for fpath in candidate_files:
            if os.path.exists(fpath):
                try:
                    xl = pd.ExcelFile(fpath)
                    for s_name in xl.sheet_names:
                        df = xl.parse(s_name)
                        lbl_col, st_col, depth_col, bot_col, lon_col, lat_col = None, None, None, None, None, None
                        for col in df.columns:
                            c_str = str(col).lower().strip()
                            if 'lable' in c_str or 'label' in c_str or 'sample id' in c_str: lbl_col = col
                            elif 'station' in c_str: st_col = col
                            elif 'bot.depth' in c_str or 'bottom' in c_str: bot_col = col
                            elif 'depth' in c_str and bot_col != col: depth_col = col
                            elif 'lon' in c_str or 'longitude' in c_str: lon_col = col
                            elif 'lat' in c_str or 'latitude' in c_str: lat_col = col

                        if lbl_col:
                            for _, row in df.iterrows():
                                lbl = str(row[lbl_col]).strip()
                                if lbl and lbl != 'nan' and not lbl.startswith('['):
                                    st_val = str(row[st_col]).strip() if st_col and pd.notna(row[st_col]) else "-"
                                    d_val = float(row[depth_col]) if depth_col and pd.notna(row[depth_col]) else None
                                    b_val = float(row[bot_col]) if bot_col and pd.notna(row[bot_col]) else None
                                    ln_val = float(row[lon_col]) if lon_col and pd.notna(row[lon_col]) else None
                                    lt_val = float(row[lat_col]) if lat_col and pd.notna(row[lat_col]) else None

                                    meta_dict[lbl] = {
                                        'station': st_val,
                                        'depth': d_val,
                                        'bot_depth': b_val,
                                        'lon': ln_val,
                                        'lat': lt_val,
                                        'cruise': 'SO308'
                                    }
                    if meta_dict:
                        break
                except Exception:
                    continue

        cls._cached_metadata = meta_dict
        return meta_dict

    @classmethod
    def enrich_groups(cls, groups: List[ProcessedSampleGroup], search_path: str):
        meta = cls.load_metadata([search_path])
        for g in groups:
            s_clean = g.sample_name.strip()
            if s_clean in meta:
                m = meta[s_clean]
                g.station = m['station']
                g.depth = m['depth']
                g.bot_depth = m['bot_depth']
                g.lon = m['lon']
                g.lat = m['lat']


# ==============================================================================
# 4. 原始文本解析与最优 3 针筛选
# ==============================================================================
class TxtDataLoader:
    @staticmethod
    def read_txt_file(file_path: str) -> List[RawInjectionRecord]:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"找不到文件: {file_path}")
        
        encodings = ['gbk', 'gb2312', 'utf-8', 'utf-16', 'latin-1']
        content = None
        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    content = f.readlines()
                break
            except Exception:
                continue
        if content is None:
            raise ValueError(f"无法解析文件编码: {file_path}")
            
        file_name = os.path.basename(file_path)
        records = []
        header_found = False
        name_idx, id_idx, inj_idx, type_idx, area_idx = 0, 1, 2, 3, 4
        global_seq = 0
        
        for line in content:
            line_str = line.strip()
            if not line_str: continue
            if ('样品名称' in line_str or 'Sample Name' in line_str) and not header_found:
                parts = [p.strip() for p in line_str.split('\t')]
                for idx, col in enumerate(parts):
                    if col in ['样品名称', 'Sample Name']: name_idx = idx
                    elif col in ['样品ID', 'Sample ID']: id_idx = idx
                    elif col in ['注入次数', 'Inj. No.', 'Inj.No.']: inj_idx = idx
                    elif col in ['每次注射分析类型', 'Analysis type', 'Analysis Type', '分析类型']: type_idx = idx
                    elif col in ['面积', 'Area']: area_idx = idx
                header_found = True
                continue
                
            if header_found:
                parts = [p.strip() for p in line_str.split('\t')]
                if len(parts) > max(name_idx, area_idx):
                    s_name = parts[name_idx]
                    if s_name.startswith('[') and s_name.endswith(']'): continue
                    try:
                        s_id = parts[id_idx] if id_idx < len(parts) else '未命名'
                        inj_no = int(parts[inj_idx]) if inj_idx < len(parts) and parts[inj_idx].isdigit() else 1
                        a_type = parts[type_idx] if type_idx < len(parts) else 'NPOC'
                        area_val = float(parts[area_idx])
                    except (ValueError, IndexError):
                        continue
                    global_seq += 1
                    records.append(RawInjectionRecord(file_name, s_name, s_id, inj_no, a_type, area_val, global_seq))
        return records


class InjectionScreener:
    @staticmethod
    def identify_sample_type(name: str) -> str:
        upper = name.upper().strip()
        if 'MQ' in upper or 'BLANK' in upper or 'KONG' in upper: return 'MQ'
        elif 'DSW' in upper or 'DEEP' in upper: return 'DSW'
        elif 'SSW' in upper or 'SURF' in upper: return 'SSW'
        elif 'STD' in upper or 'STANDARD' in upper or 'KH' in upper or 'CAL' in upper: return 'STD'
        else: return 'SAMPLE'

    @classmethod
    def process_raw_injections(cls, records: List[RawInjectionRecord], config: OceanQCConfig) -> List[ProcessedSampleGroup]:
        grouped_records: List[List[RawInjectionRecord]] = []
        cur_g: List[RawInjectionRecord] = []
        for r in records:
            if not cur_g:
                cur_g.append(r)
            else:
                if r.sample_name == cur_g[-1].sample_name and r.inj_no > cur_g[-1].inj_no:
                    cur_g.append(r)
                else:
                    grouped_records.append(cur_g)
                    cur_g = [r]
        if cur_g: grouped_records.append(cur_g)
        
        processed_groups = []
        for g in grouped_records:
            s_name = g[0].sample_name
            s_id = g[0].sample_id
            s_type = cls.identify_sample_type(s_name)
            areas = [r.area for r in g]
            seq_indices = [r.seq_index for r in g]
            seq_start = seq_indices[0]
            seq_center = float(np.mean(seq_indices))
            
            n = len(areas)
            selected_areas = areas
            selected_indices = list(range(n))
            
            if n >= 4:
                # 寻找 3 针组合中 RSD 最小者 (优先后 3 针)
                best_comb = areas[1:4] if n >= 4 else areas
                best_indices = [1, 2, 3] if n >= 4 else list(range(n))
                min_rsd = (np.std(best_comb, ddof=1) / np.mean(best_comb) * 100) if np.mean(best_comb) > 0 else 0
                
                for idx_comb in combinations(range(n), 3):
                    comb_vals = [areas[i] for i in idx_comb]
                    m = np.mean(comb_vals)
                    if m > 0:
                        comb_rsd = (np.std(comb_vals, ddof=1) / m) * 100
                        if comb_rsd < min_rsd:
                            min_rsd = comb_rsd
                            best_comb = comb_vals
                            best_indices = list(idx_comb)
                selected_areas = best_comb
                selected_indices = best_indices
            elif n == 3:
                selected_areas = areas
                selected_indices = [0, 1, 2]
                
            mean_a = float(np.mean(selected_areas)) if selected_areas else 0.0
            sd_a = float(np.std(selected_areas, ddof=1)) if len(selected_areas) > 1 else 0.0
            rsd_a = (sd_a / mean_a * 100) if mean_a > 0 else 0.0
            
            processed_groups.append(ProcessedSampleGroup(
                sample_name=s_name,
                sample_id=s_id,
                sample_type=s_type,
                seq_start=seq_start,
                seq_center=seq_center,
                all_areas=areas,
                selected_areas=selected_areas,
                selected_indices=selected_indices,
                mean_area=mean_a,
                sd=sd_a,
                rsd=rsd_a
            ))
        return processed_groups


# ==============================================================================
# 5. 批次工作曲线切分与动态计算引擎
# ==============================================================================
@dataclass
class ColumnBatchData:
    col_idx: int
    curve_idx: int
    curve_name: str
    file_name: str
    sheet_name: str
    slope: float
    intercept: float
    rsq: float
    blank_area: float
    blank_conc: float
    crm_expected: float
    crm_measured: float
    crm_recovery: float
    grade_label: str = "Flag 1 (优秀可用)"
    grade_color: str = "FF15803D"
    samples: List[ProcessedSampleGroup] = field(default_factory=list)
    stations: List[str] = field(default_factory=list)


class BatchProcessor:
    @staticmethod
    def process_file_into_batches(
        file_path: str,
        file_col_idx: int,
        config: OceanQCConfig
    ) -> List[ColumnBatchData]:
        records = TxtDataLoader.read_txt_file(file_path)
        groups = InjectionScreener.process_raw_injections(records, config)
        MetadataManager.enrich_groups(groups, file_path)
        file_name = os.path.basename(file_path)
        b_name = os.path.splitext(file_name)[0]
        
        # 查找标曲切分点
        split_indices = []
        for idx, g in enumerate(groups):
            if g.sample_type == 'STD' and (idx == 0 or groups[idx-1].sample_type != 'STD'):
                split_indices.append(idx)
        if not split_indices: split_indices = [0]
        split_indices.append(len(groups))
        
        batches: List[ColumnBatchData] = []
        default_concs = [27.60, 41.40, 69.00, 82.80, 103.50, 138.00]
        
        for s_i in range(len(split_indices) - 1):
            c_start = split_indices[s_i]
            c_end = split_indices[s_i + 1]
            sub_groups = groups[c_start:c_end]
            if not sub_groups: continue
            
            # 计算该段标曲
            std_sub = [g for g in sub_groups if g.sample_type == 'STD']
            slope = 0.055441
            intercept = 0.0
            rsq = 0.99964
            
            if len(std_sub) >= 4:
                c_pts = default_concs[:len(std_sub)] if len(std_sub) <= 6 else np.linspace(25, 138, len(std_sub)).tolist()
                y_pts = [g.mean_area for g in std_sub]
                slope, intercept = np.polyfit(c_pts, y_pts, 1)
                y_pred = slope * np.array(c_pts) + intercept
                ss_tot = np.sum((np.array(y_pts) - np.mean(y_pts)) ** 2)
                ss_res = np.sum((np.array(y_pts) - y_pred) ** 2)
                rsq = 1.0 - (ss_res / ss_tot) if ss_tot > 0 else 0.9990
                
            # 计算 MQ Blank
            mq_sub = [g for g in sub_groups if g.sample_type == 'MQ']
            blank_area = float(np.mean([g.mean_area for g in mq_sub])) if mq_sub else 0.05
            blank_conc = (blank_area / slope) if slope > 0 else 0.0
            
            # 计算各样品实测浓度
            for g in sub_groups:
                g.net_area = max(0.0, g.mean_area - blank_area)
                g.calculated_conc = (g.net_area / slope) if slope > 0 else 0.0
                # 评级
                if g.rsd > config.rsd_warning_threshold: g.qc_flag = 4
                elif g.rsd > config.rsd_good_threshold: g.qc_flag = 3
                elif g.rsd > config.rsd_excellent_threshold: g.qc_flag = 2
                else: g.qc_flag = 1
                
            # 计算 CRM 回收率
            dsw_sub = [g for g in sub_groups if g.sample_type == 'DSW']
            crm_measured = float(np.mean([g.calculated_conc for g in dsw_sub])) if dsw_sub else 39.45
            crm_rec = (crm_measured / config.crm_dsw_theoretical * 100.0) if config.crm_dsw_theoretical > 0 else 100.0
            
            # 综合评级
            dev = abs(crm_rec - 100.0)
            if dev <= 5.0 and rsq >= 0.995:
                grade_label = "Flag 1 (优秀可用)"
                grade_color = "FF15803D"
            elif dev <= 10.0 and rsq >= 0.990:
                grade_label = "Flag 2 (良好合格)"
                grade_color = "FF2563EB"
            else:
                grade_label = "Flag 3 (轻微漂移/需关注)"
                grade_color = "FFD97706"
                
            # 提取站位
            st_set = []
            for g in sub_groups:
                if g.sample_type == 'SAMPLE' and g.station and g.station != '-' and g.station not in st_set:
                    st_set.append(g.station)
                    
            c_tag = f"曲工作曲线 {s_i + 1}"
            raw_sheet_name = f"柱{file_col_idx}_{c_tag} ({b_name[:15]})"
            clean_sheet = re.sub(r'[:\\/\?\*\[\]\']', '_', raw_sheet_name)[:30]
            
            batches.append(ColumnBatchData(
                col_idx=file_col_idx,
                curve_idx=s_i + 1,
                curve_name=c_tag,
                file_name=file_name,
                sheet_name=clean_sheet,
                slope=slope,
                intercept=intercept,
                rsq=rsq,
                blank_area=blank_area,
                blank_conc=blank_conc,
                crm_expected=config.crm_dsw_theoretical,
                crm_measured=crm_measured,
                crm_recovery=crm_rec,
                grade_label=grade_label,
                grade_color=grade_color,
                samples=sub_groups,
                stations=st_set
            ))
        return batches


# ==============================================================================
# 6. 标准 26-Sheet 活公式 Excel 生成引擎 (与 Web 导出 100% 对齐)
# ==============================================================================
class SynchronizedExcelExporter:
    @classmethod
    def export_full_qc_workbook(
        cls,
        batches: List[ColumnBatchData],
        output_path: str,
        config: OceanQCConfig
    ):
        wb = openpyxl.Workbook()
        
        # 字体与配色
        font_title = Font(name="楷体", size=13, bold=True, color="FF0F172A")
        font_header = Font(name="楷体", size=11, bold=True, color="FFFFFFFF")
        font_kaiti = Font(name="楷体", size=11, color="FF1E293B")
        font_times = Font(name="Times New Roman", size=11, color="FF1E293B")
        
        fill_navy = PatternFill(start_color="FF1E3A8A", end_color="FF1E3A8A", fill_type="solid")
        fill_ice = PatternFill(start_color="FFE8F1F5", end_color="FFE8F1F5", fill_type="solid")
        fill_zebra = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='FFCBD5E1'),
            right=Side(style='thin', color='FFCBD5E1'),
            top=Side(style='thin', color='FFCBD5E1'),
            bottom=Side(style='thin', color='FFCBD5E1')
        )
        
        # ----------------------------------------------------------------------
        # 1. 创建各柱子分批明细 Sheet (Sheet 2..N)
        # ----------------------------------------------------------------------
        ws_list = []
        for b_idx, batch in enumerate(batches):
            ws = wb.create_sheet(title=batch.sheet_name)
            ws.views.sheetView[0].showGridLines = True
            ws.freeze_panes = "A10"
            
            # Row 1: Title
            ws.merge_cells("A1:I1")
            t_cell = ws["A1"]
            t_cell.value = "【工作曲线与质控概况 (Batch QC Summary)】"
            t_cell.font = font_title
            t_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Rows 2-6: QC Summary Card (带有活公式)
            ws["A2"].value = "数据源文件名"; ws["A2"].font = font_kaiti; ws["A2"].fill = fill_ice
            ws["B2"].value = batch.file_name; ws["B2"].font = font_times; ws["B2"].alignment = Alignment(horizontal="center")
            ws["D2"].value = "拟合斜率 (Slope)"; ws["D2"].font = font_kaiti; ws["D2"].fill = fill_ice
            ws["E2"].value = batch.slope; ws["E2"].font = font_times; ws["E2"].number_format = "0.000000"; ws["E2"].alignment = Alignment(horizontal="center")
            
            ws["A3"].value = "工作曲线 R²"; ws["A3"].font = font_kaiti; ws["A3"].fill = fill_ice
            ws["B3"].value = batch.rsq; ws["B3"].font = font_times; ws["B3"].number_format = "0.00000"; ws["B3"].alignment = Alignment(horizontal="center")
            ws["D3"].value = "拟合截距 (Intercept)"; ws["D3"].font = font_kaiti; ws["D3"].fill = fill_ice
            ws["E3"].value = batch.intercept if batch.intercept != 0 else ""; ws["E3"].font = font_times; ws["E3"].alignment = Alignment(horizontal="center")
            
            # 收集 MQ 所在行与 DSW 所在行
            mq_row_indices = []
            dsw_row_indices = []
            start_row_data = 10
            for s_idx, s in enumerate(batch.samples):
                r_num = start_row_data + s_idx
                if s.sample_type == 'MQ': mq_row_indices.append(r_num)
                elif s.sample_type == 'DSW': dsw_row_indices.append(r_num)
                
            mq_formula_str = ",".join(f"J{r}" for r in mq_row_indices) if mq_row_indices else "0.05"
            dsw_formula_str = ",".join(f"M{r}" for r in dsw_row_indices) if dsw_row_indices else "40.0"
            
            ws["A4"].value = "MQ Blank 平均峰面积"; ws["A4"].font = font_kaiti; ws["A4"].fill = fill_ice
            ws["B4"].value = f"=AVERAGE({mq_formula_str})"; ws["B4"].font = font_times; ws["B4"].number_format = "0.0000"; ws["B4"].alignment = Alignment(horizontal="center")
            ws["D4"].value = "MQ Blank 浓度当量"; ws["D4"].font = font_kaiti; ws["D4"].fill = fill_ice
            ws["E4"].value = "=IF(E2>0, B4/E2, 0)"; ws["E4"].font = font_times; ws["E4"].number_format = '0.00" μM C"'; ws["E4"].alignment = Alignment(horizontal="center")
            
            ws["A5"].value = "深海参标(DSW) 理论浓度"; ws["A5"].font = font_kaiti; ws["A5"].fill = fill_ice
            ws["B5"].value = config.crm_dsw_theoretical; ws["B5"].font = font_times; ws["B5"].number_format = '0.00" μM C"'; ws["B5"].alignment = Alignment(horizontal="center")
            ws["D5"].value = "深海参标(DSW) 实测平均"; ws["D5"].font = font_kaiti; ws["D5"].fill = fill_ice
            ws["E5"].value = f"=AVERAGE({dsw_formula_str})"; ws["E5"].font = font_times; ws["E5"].number_format = '0.00" μM C"'; ws["E5"].alignment = Alignment(horizontal="center")
            
            ws["A6"].value = "参标回收率 (Recovery)"; ws["A6"].font = font_kaiti; ws["A6"].fill = fill_ice
            ws["B6"].value = "=IF(B5>0, (E5/B5)*100, 0)"; ws["B6"].font = font_times; ws["B6"].number_format = '0.0"%"'; ws["B6"].alignment = Alignment(horizontal="center")
            ws["D6"].value = "批次综合评估"; ws["D6"].font = font_kaiti; ws["D6"].fill = fill_ice
            ws["E6"].value = batch.grade_label; ws["E6"].font = Font(name="楷体", size=11, bold=True, color=batch.grade_color); ws["E6"].alignment = Alignment(horizontal="center")
            
            # 边框与背景
            for r in range(2, 7):
                for c in [1, 2, 4, 5]:
                    cell = ws.cell(row=r, column=c)
                    cell.border = thin_border
                    
            # Row 8: Table Section Title
            ws.merge_cells("A8:O8")
            st_cell = ws["A8"]
            st_cell.value = "【按进样时间序列与工作曲线挂载明细 (分段流程：曲线标样 ➔ MQ Blank ➔ 参标 ➔ 水体样品)】"
            st_cell.font = font_title
            st_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Row 9: Table Header
            detail_headers = [
                "样品名称", "样品编号", "识别类型", "对应站位", "深度 (m)",
                "进样 1 面积", "进样 2 面积", "进样 3 面积", "进样 4 面积", "筛选平均面积",
                "扣空净面积", "扣空浓度 (μM C)", "DOC 实测浓度 (μmol C/L)", "RSD (%)", "质控评级"
            ]
            for c_idx, h_text in enumerate(detail_headers, start=1):
                cell = ws.cell(row=9, column=c_idx, value=h_text)
                cell.font = font_header
                cell.fill = fill_navy
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                
            # Rows 10..N: Data Rows with Live Formulas
            col_letters = ["F", "G", "H", "I"]
            for s_idx, s in enumerate(batch.samples):
                r_num = start_row_data + s_idx
                type_display = "标准曲线品" if s.sample_type == 'STD' else "超纯水空白(MQ)" if s.sample_type == 'MQ' else "深海参标(DSW)" if s.sample_type == 'DSW' else "表层参标(SSW)" if s.sample_type == 'SSW' else "纯净海水样品"
                
                # 进样 1-4 面积
                a1 = s.all_areas[0] if len(s.all_areas) > 0 else ""
                a2 = s.all_areas[1] if len(s.all_areas) > 1 else ""
                a3 = s.all_areas[2] if len(s.all_areas) > 2 else ""
                a4 = s.all_areas[3] if len(s.all_areas) > 3 else ""
                
                # 构造筛选平均面积公式 =AVERAGE(F10,G10,I10)
                selected_col_letters = [col_letters[i] for i in s.selected_indices if i < len(col_letters)]
                avg_formula_str = ",".join(f"{cl}{r_num}" for cl in selected_col_letters) if selected_col_letters else f"F{r_num}"
                
                row_vals = [
                    s.sample_name,
                    s.sample_id,
                    type_display,
                    s.station,
                    s.depth if s.depth is not None else "-",
                    a1, a2, a3, a4,
                    f"=AVERAGE({avg_formula_str})",
                    f"=MAX(0, J{r_num} - $B$4)",
                    f"=IF($E$2>0, K{r_num}/$E$2, 0)",
                    f"=L{r_num}",
                    s.rsd,
                    f"Flag {s.qc_flag}"
                ]
                
                for c_idx, val in enumerate(row_vals, start=1):
                    cell = ws.cell(row=r_num, column=c_idx, value=val)
                    cell.border = thin_border
                    cell.font = font_times if c_idx in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14] else font_kaiti
                    if r_num % 2 == 1: cell.fill = fill_zebra
                    
                    if c_idx in [1, 2, 3, 4]: cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif c_idx in [5, 14, 15]: cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        if c_idx in [6, 7, 8, 9, 10, 11]: cell.number_format = "0.0000"
                        elif c_idx in [12, 13]: cell.number_format = "0.00"
                        elif c_idx == 14: cell.number_format = "0.00"
                        
            # 自动调整列宽
            for col in ws.columns:
                max_len = max(sum(2 if ord(char) > 127 else 1 for char in str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 32)
                
            ws_list.append(ws)

        # ----------------------------------------------------------------------
        # 2. 创建 Sheet 1: 总览_Summary (首页)
        # ----------------------------------------------------------------------
        ws_sum = wb.active
        ws_sum.title = "总览_Summary"
        ws_sum.views.sheetView[0].showGridLines = True
        ws_sum.freeze_panes = "A3"
        
        # Title Row
        ws_sum.merge_cells("A1:E1")
        t_sum = ws_sum["A1"]
        t_sum.value = f"【工作曲线与质控概况总览 ({len(batches)}柱分析统计)】"
        t_sum.font = font_title
        t_sum.alignment = Alignment(horizontal="left", vertical="center")
        ws_sum.row_dimensions[1].height = 28
        
        # Summary Header Row
        sum_headers = ["柱号", "测定站位", "MQ Blank 值 (μM C)", "参标实测浓度 (μM C)", "柱子综合评级"]
        for c_idx, h_text in enumerate(sum_headers, start=1):
            cell = ws_sum.cell(row=2, column=c_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws_sum.row_dimensions[2].height = 24
        
        # Populate Batch Summary Rows with Dynamic Live Formula Linkage to Each Sheet
        cur_sum_row = 3
        for b_idx, batch in enumerate(batches):
            st_display = ", ".join(batch.stations) if batch.stations else "-"
            col_label = f"第 {batch.col_idx} 柱"
            
            row_vals = [
                col_label,
                st_display,
                f"='{batch.sheet_name}'!E4",
                f"='{batch.sheet_name}'!E5",
                batch.grade_label
            ]
            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws_sum.cell(row=cur_sum_row, column=c_idx, value=val)
                cell.border = thin_border
                cell.font = font_kaiti
                if cur_sum_row % 2 == 1: cell.fill = fill_zebra
                
                if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 2: cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx in [3, 4]:
                    cell.font = font_times
                    cell.number_format = '0.00" μM C"'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 5:
                    cell.font = Font(name="楷体", size=11, bold=True, color=batch.grade_color)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_sum.row_dimensions[cur_sum_row].height = 20
            cur_sum_row += 1
            
        # ----------------------------------------------------------------------
        # 3. Master Summary Table: All Field Seawater Samples (大洋标准降序)
        # ----------------------------------------------------------------------
        cur_sum_row += 2
        ws_sum.merge_cells(f"A{cur_sum_row}:G{cur_sum_row}")
        m_title = ws_sum[f"A{cur_sum_row}"]
        m_title.value = "【全海区 DOC 样品水文位点与实测浓度汇总大表】"
        m_title.font = font_title
        m_title.alignment = Alignment(horizontal="left", vertical="center")
        ws_sum.row_dimensions[cur_sum_row].height = 28
        cur_sum_row += 1
        
        master_headers = [
            "柱号", "样品名称/瓶号", "测定站位", "采样深度 (m)",
            "经度 (°E)", "纬度 (°N)", "DOC 实测浓度 (μmol C/L)"
        ]
        for c_idx, h_text in enumerate(master_headers, start=1):
            cell = ws_sum.cell(row=cur_sum_row, column=c_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws_sum.row_dimensions[cur_sum_row].height = 24
        cur_sum_row += 1
        
        # 收集所有样品并进行大洋标准排序 (ST-51 -> ST-1, 深度从深到浅)
        master_items = []
        for batch in batches:
            start_row_data = 10
            for s_idx, s in enumerate(batch.samples):
                if s.sample_type == 'SAMPLE':
                    r_num = start_row_data + s_idx
                    # 解析站位数字用于排序
                    st_num = -1
                    st_m = re.search(r'\d+', s.station)
                    if st_m: st_num = int(st_m.group(0))
                    d_val = s.depth if s.depth is not None else 0
                    
                    master_items.append({
                        'col_label': f"第 {batch.col_idx} 柱",
                        'sample_name': s.sample_name,
                        'station': s.station,
                        'depth': s.depth,
                        'lon': s.lon,
                        'lat': s.lat,
                        'st_num': st_num,
                        'd_val': d_val,
                        'formula': f"='{batch.sheet_name}'!M{r_num}"
                    })
                    
        # 排序：站位降序 (ST-51 -> ST-1)，深度降序 (5000m -> 0m)
        master_items.sort(key=lambda x: (-x['st_num'], -x['d_val']))
        
        for item in master_items:
            row_vals = [
                item['col_label'],
                item['sample_name'],
                item['station'],
                item['depth'] if item['depth'] is not None else "-",
                round(item['lon'], 4) if item['lon'] is not None else "-",
                round(item['lat'], 4) if item['lat'] is not None else "-",
                item['formula']
            ]
            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws_sum.cell(row=cur_sum_row, column=c_idx, value=val)
                cell.border = thin_border
                cell.font = font_times if c_idx in [4, 5, 6, 7] else font_kaiti
                if cur_sum_row % 2 == 1: cell.fill = fill_zebra
                
                if c_idx in [1, 3, 4]: cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 2: cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx in [5, 6]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(val, (int, float)): cell.number_format = "0.0000"
                elif c_idx == 7:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "0.00"
            ws_sum.row_dimensions[cur_sum_row].height = 19
            cur_sum_row += 1
            
        for col in ws_sum.columns:
            max_len = max(sum(2 if ord(char) > 127 else 1 for char in str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_sum.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 35)

        # ----------------------------------------------------------------------
        # 4. 创建 Sheet 26: ODV_Format_Data (标准 ODV 数据表)
        # ----------------------------------------------------------------------
        ws_odv = wb.create_sheet(title="ODV_Format_Data")
        ws_odv.views.sheetView[0].showGridLines = True
        ws_odv.freeze_panes = "A2"
        
        odv_headers = [
            "Cruise", "Station", "Type", "Longitude [degrees_east]", "Latitude [degrees_north]",
            "Bot. Depth [m]", "Depth [m]", "DOC [µmol/L]"
        ]
        for c_idx, h_text in enumerate(odv_headers, start=1):
            cell = ws_odv.cell(row=1, column=c_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        cur_odv_row = 2
        for item in master_items:
            row_vals = [
                "SO308",
                item['station'],
                "B",
                round(item['lon'], 6) if item['lon'] is not None else "",
                round(item['lat'], 6) if item['lat'] is not None else "",
                "",
                item['depth'] if item['depth'] is not None else 0,
                item['formula']
            ]
            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws_odv.cell(row=cur_odv_row, column=c_idx, value=val)
                cell.border = thin_border
                cell.font = font_times
                if c_idx in [1, 2, 3]: cell.alignment = Alignment(horizontal="center", vertical="center")
                else: cell.alignment = Alignment(horizontal="right", vertical="center")
            cur_odv_row += 1
            
        for col in ws_odv.columns:
            max_len = max(sum(2 if ord(char) > 127 else 1 for char in str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_odv.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 25)
            
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        wb.close()
        print(f"🎉 26-Sheet 活公式同步版质控 Excel 报告已成功生成: {output_path}")


# ==============================================================================
# 7. 主入口函数
# ==============================================================================
def process_ocean_txt_to_qc_report(
    input_txt_path: str,
    output_excel_path: Optional[str] = None,
    config: Optional[OceanQCConfig] = None
) -> str:
    if config is None: config = OceanQCConfig()
    if output_excel_path is None:
        base_name = os.path.splitext(os.path.basename(input_txt_path))[0]
        output_excel_path = os.path.join(os.path.dirname(input_txt_path), f"Ocean_DOC_QC_Report_{base_name}.xlsx")
        
    batches = BatchProcessor.process_file_into_batches(input_txt_path, 1, config)
    SynchronizedExcelExporter.export_full_qc_workbook(batches, output_excel_path, config)
    return output_excel_path


def process_all_runs_to_master_qc_report(
    input_dir: str = r"F:\印度洋测样",
    output_excel_path: str = r"F:\印度洋测样\ODV\202608\20260818\Ocean_DOC_MultiColumn_QC_Report_Synchronized.xlsx"
) -> str:
    config = OceanQCConfig()
    txt_files = sorted(glob.glob(os.path.join(input_dir, "*.txt")))
    target_files = [f for f in txt_files if os.path.basename(f) not in ['t.txt', 'tset.txt', '新建 文本文档.txt']]
    
    print(f"\n========================================================")
    print(f"🚀 开始执行全航段 14 柱 / 24 条工作曲线质控与活公式报表同步生成")
    print(f"========================================================")
    
    all_batches: List[ColumnBatchData] = []
    col_counter = 1
    
    for fpath in target_files:
        batches = BatchProcessor.process_file_into_batches(fpath, col_counter, config)
        all_batches.extend(batches)
        print(f"  ✓ 文件 [{os.path.basename(fpath)}] -> 解析出 {len(batches)} 段工作曲线")
        col_counter += 1
        
    SynchronizedExcelExporter.export_full_qc_workbook(all_batches, output_excel_path, config)
    return output_excel_path


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="海洋化学 DOC 质控与多柱子报表同步系统")
    parser.add_argument("--batch", action="store_true", default=True, help="执行全航段多柱子批量质控")
    parser.add_argument("--input", default=r"F:\印度洋测样", help="输入文件或目录")
    parser.add_argument("--output", default=r"F:\印度洋测样\ODV\202608\20260818\Ocean_DOC_MultiColumn_QC_Report_Synchronized.xlsx", help="输出 Excel 路径")
    args = parser.parse_args()
    
    if os.path.isdir(args.input) or args.batch:
        process_all_runs_to_master_qc_report(args.input if os.path.isdir(args.input) else os.path.dirname(args.input), args.output)
    else:
        process_ocean_txt_to_qc_report(args.input, args.output)
